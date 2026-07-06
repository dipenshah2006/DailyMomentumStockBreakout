"""added ATH
╔═════════════════════════════════════════════════════════════════════════════╗
║   RSI MULTI-TIMEFRAME BREAKOUT HTML REPORT  v2.0                            ║
║   Daily · Weekly · Monthly RSI/SMA Crossover | Phase | Entry/Exit           ║
║   NEW v2.0:                                                                  ║
║    • Exception logging → error_log.txt  (ticker + company + full traceback) ║
║    • Ranking vs Nifty50  (relative strength percentile)                     ║
║    • Ranking vs all NSE stocks  (universe percentile)                       ║
║    • Lightweight HTML — charts lazy-loaded on expand, never hangs browser   ║
║    • Native <details> expand/collapse — no JS needed, instant               ║
╚═════════════════════════════════════════════════════════════════════════════╝

INSTALL:  pip install yfinance pandas numpy matplotlib requests openpyxl
RUN:      python rsi_mtf_report_v2.py
OUTPUTS:  rsi_mtf_report_YYYYMMDD_HHMM.html  +  error_log_YYYYMMDD_HHMM.txt
"""

import os
import glob as _glob

# ═════════════════════════════════════════════════════════════════════════════
# USER CONFIG
# ═════════════════════════════════════════════════════════════════════════════

LOCAL_NSE_CSV       = "india/NSE/NSECash/EQUITY_L.csv"
NSE_CSV_URL         = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
SERIES_FILTER       = ["EQ"]       # NSE equity series (EQ = cash equities)

LOCAL_SME_CSV       = "india/NSE/NSESME/MW-SME-05-May-2026.csv"
SME_SERIES_FILTER   = ["ST", "SM"] # NSE SME series (ST = SME T, SM = SME M)

# Nifty Indices Master Excel — multi-sheet workbook with *_Stocks sheets
# Layout: sheets named Broad_Based_Stocks, Sectoral_Stocks, Thematic_Stocks,
#         Strategy_Stocks, All_Stocks_Combined
# Columns: #, Index Name, Category, Symbol, Company Name  (header on row 3)
LOCAL_INDICES_XLSX  = "india/NSE/NIFTY_Indices_Master.xlsx"

# Optional: folder containing individual NSE index constituent CSVs
# e.g. india/NSE/NseIndice/ind_nifty50list.csv, ind_niftybanklist.csv, etc.
# Download from: https://www.niftyindices.com/indices/equity/broad-based-indices
# Supports filenames like: ind_nifty50list.csv, ind_niftyAlpha_Index.csv, ind_nifty_alpha_lowvol30list.csv
# Name extraction strips: leading "ind_" / "Ind_", trailing "list" / "_list", then replaces "_" with spaces
LOCAL_INDICES_DIR   = "india/NSE/NseIndice"
LOCAL_FO_CSV        = "india/NSE/nse_fo_list.csv"   # NSE F&O securities list

# ASX (Australian Securities Exchange) stocks
# CSV columns expected: "ACT Symbol", "Company name", "GICS industry group"
LOCAL_ASX_CSV       = "ASX/nyse-listed.csv"
ASX_CHART_OUTPUT_DIR = "charts/asx"            # folder for ASX chart PNGs

DATA_PERIOD         = "max"
MIN_CANDLES         = 1          # include all stocks regardless of history length
MAX_CHART_STOCKS    = 0         # 0 = generate charts for all stocks; otherwise top N stocks
# Allow GitHub Actions (or any CI) to cap chart count via env variable
_chart_override = os.environ.get("MAX_CHART_STOCKS_OVERRIDE", "")
if _chart_override.isdigit():
    MAX_CHART_STOCKS = int(_chart_override)
GITHUB_CHARTS_BASE  = "charts"   # "https://raw.githubusercontent.com/dipenshah2006/DailyMomentumStockBreakout/main/charts"
CHART_OUTPUT_DIR    = "charts"   # folder for generated PNG chart files
CHART_BARS          = 90        # bars per chart — 90 days is plenty; 120 adds ~25% render time
CHART_DPI           = 120       # Good quality; 200 is overkill and ~3x slower

FRESH_DAYS_D        = 3
FRESH_WEEKS_W       = 2

RSI_P               = 14
RSI_SMA_P           = 34
CCI_P               = 20
MACD_F, MACD_S, MACD_SIG_P = 12, 26, 9
ATR_P               = 14

BATCH_SIZE          = 25
BATCH_PAUSE         = 1.0

# HTML report pagination constants
PAGE_TBL            = 100   # table rows per page
PAGE_CARDS          = 50    # cards per "Load More" batch

SCORE_STRONG_BUY    = 16
SCORE_BUY           = 12
SCORE_WATCH         = 8

# Nifty 50 tickers (used for ranking vs index)
NIFTY50 = [
    "RELIANCE","TCS","HDFCBANK","INFY","ICICIBANK","HINDUNILVR","ITC","SBIN",
    "BAJFINANCE","BHARTIARTL","KOTAKBANK","LT","AXISBANK","ASIANPAINT","MARUTI",
    "SUNPHARMA","TITAN","WIPRO","ULTRACEMCO","NTPC","POWERGRID","ONGC","JSWSTEEL",
    "TATASTEEL","COALINDIA","TECHM","HCLTECH","DRREDDY","CIPLA","DIVISLAB",
    "ADANIENT","ADANIPORTS","BAJAJ-AUTO","EICHERMOT","HEROMOTOCO","NESTLEIND",
    "BRITANNIA","TATACONSUM","TATAMOTORS","M&M","HINDALCO","GRASIM","JSWSTEEL",
    "APOLLOHOSP","BPCL","INDUSINDBK","LTIM","HDFCLIFE","SBILIFE","COALINDIA",
]

# ═════════════════════════════════════════════════════════════════════════════
# AUTO-INSTALL MISSING LIBRARIES
# ═════════════════════════════════════════════════════════════════════════════

import sys
import subprocess

def install_missing_packages():
    required = {
        'yfinance': 'yfinance',
        'pandas': 'pandas',
        'numpy': 'numpy',
        'matplotlib': 'matplotlib',
        'requests': 'requests',
        'openpyxl': 'openpyxl',
    }

    missing = []
    for pkg_name, import_name in required.items():
        try:
            __import__(import_name)
        except ImportError:
            missing.append(pkg_name)

    if missing:
        print(f"Installing missing packages: {', '.join(missing)}")
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing)
        print("✓ Packages installed successfully\n")

if __name__ == "__main__":
    install_missing_packages()

# ═════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ═════════════════════════════════════════════════════════════════════════════

import csv
import hashlib
import io
import json
import logging
import os
import pickle
import sys
import time
import traceback
import warnings
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from datetime import datetime

import argparse
import matplotlib
matplotlib.use("Agg")
import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import requests
import yfinance as yf

warnings.filterwarnings("ignore")

CHART_WORKERS = min(12, max(2, (os.cpu_count() or 4)))

RUN_TS      = datetime.now().strftime("%d %b %Y  %H:%M")
_STAMP      = datetime.now().strftime("%d%m%Y_%H%M")
START_TS    = RUN_TS
START_TIME  = time.time()
OUTPUT_HTML = "asx_report_NSE.html"        # fixed name — avoids conflict with main NSE report
ERROR_LOG   = f"error_log_{_STAMP}.txt"   # timestamped so each run's errors are separate
CACHE_FILE  = "asx_stock_cache.pkl"
CHART_CACHE_META = os.path.join(CHART_OUTPUT_DIR, "chart_cache.json")

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 0 — ERROR LOGGER
# ═════════════════════════════════════════════════════════════════════════════

# Configure a dedicated file logger — separate from print output
_logger = logging.getLogger("rsi_scanner")
_logger.setLevel(logging.DEBUG)
_fh = logging.FileHandler(ERROR_LOG, encoding="utf-8")
_fh.setFormatter(logging.Formatter(
    "%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
))
_logger.addHandler(_fh)

def log_error(ticker: str, company: str, stage: str, exc: Exception):
    """
    Log a structured error entry to ERROR_LOG with:
    - Timestamp
    - Ticker symbol
    - Company name
    - Stage where it failed (download / indicators / chart / html)
    - Full traceback
    """
    tb = traceback.format_exc()
    _logger.error(
        f"TICKER='{ticker}.NS' | COMPANY={company!r:35s} | STAGE={stage}\n"
        f"  ERROR : {type(exc).__name__}: {exc}\n"
        f"  TRACE :\n{tb}"
    )

def log_info(msg: str):
    _logger.info(msg)

def log_warn(ticker: str, company: str, msg: str):
    _logger.warning(f"TICKER='{ticker}.NS' | COMPANY={company!r:35s} | {msg}")


def format_timespan(seconds: float) -> str:
    mins, secs = divmod(int(seconds), 60)
    hours, mins = divmod(mins, 60)
    if hours:
        return f"{hours}h {mins}m {secs}s"
    if mins:
        return f"{mins}m {secs}s"
    return f"{secs}s"


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1 — COMPANY NAME LOOKUP
# ═════════════════════════════════════════════════════════════════════════════

_COMPANY_MAP: dict[str, str] = {}   # populated by load_universe()
_LISTING_DATE_MAP: dict[str, str] = {}   # symbol → listing date string
_SME_STOCKS: set[str] = set()   # set of SME stock symbols
_SECTOR_MAP: dict[str, str] = {}   # symbol → sector/industry name
_INDEX_MAP: dict[str, set[str]] = {}   # index name → set of symbols in that index
_INDUSTRY_MAP: dict[str, str] = {}   # symbol → industry (from index CSV "Industry" column)
_FO_SET:     set[str]            = {}
_MARKETCAP_MAP: dict[str, float] = {}   # symbol → market cap in rupees
_ASX_STOCKS: set[str] = set()   # set of ASX stock symbols (bare code, no .AX suffix)

def get_company_name(ticker: str) -> str:
    return _COMPANY_MAP.get(ticker, ticker)

def is_sme_stock(ticker: str) -> bool:
    return ticker in _SME_STOCKS

def is_asx_stock(ticker: str) -> bool:
    """Return True if this ticker is an ASX-listed stock (bare code, no .AX suffix)."""
    return ticker in _ASX_STOCKS


    """Return list of sector labels for this ticker (e.g. ['NIFTY AUTO - Sectoral',
    'NIFTY MOBILITY - Thematic']). Empty list if none."""
    v = _SECTOR_MAP.get(ticker, [])
    return v if isinstance(v, list) else [v] if v else []

def get_indices(ticker: str) -> list[str]:
    """Get list of indices this stock belongs to."""
    result = []
    for idx_name, symbols in _INDEX_MAP.items():
        if ticker in symbols:
            result.append(idx_name)
    return result

def get_industry(ticker: str) -> str:
    """Return industry string from index CSV Industry column, or ''."""
    return _INDUSTRY_MAP.get(ticker, "")


def is_fo_stock(ticker: str) -> bool:
    """Return True if this ticker is F&O eligible on NSE."""
    return ticker in _FO_SET

# Note: get_marketcap() is defined in Section 2 (cache) with TTL-aware logic

def categorize_marketcap(marketcap: float | None) -> tuple[str, str]:
    """Categorize stock by market cap. Returns (category, css_class).
    Market cap in INR:
    - Large Cap: > 20,000 crore
    - Mid Cap: 5,000 - 20,000 crore
    - Small Cap: 500 - 5,000 crore
    - Micro Cap: < 500 crore
    """
    if marketcap is None:
        return "Unknown", "cap-unknown"
    cap_crore = marketcap / 1e7  # Convert to crores
    if cap_crore > 20000:        # > 20,000 cr  → Large Cap
        return "Large Cap", "cap-large"
    elif cap_crore > 5000:       # 5,000-20,000 cr → Mid Cap
        return "Mid Cap", "cap-mid"
    elif cap_crore > 500:        # 500-5,000 cr  → Small Cap
        return "Small Cap", "cap-small"
    else:                        # < 500 cr
        return "Micro Cap", "cap-micro"

def get_listing_date(ticker: str) -> str | None:
    return _LISTING_DATE_MAP.get(ticker)

def get_min_candles_required(ticker: str) -> int:
    """Minimum candles required — set to 1 to include all stocks."""
    return MIN_CANDLES  # MIN_CANDLES=1: accept any stock with at least 1 candle

def _build_company_map(text: str):
    """
    Parse NSE CSV and build symbol → company name and listing date dicts.
    Also captures INDUSTRY/SECTOR column if present (NSE EQUITY_L.csv does NOT
    have this column — sector data comes from _populate_indices instead).
    Robust to column names with leading/trailing spaces.
    """
    reader = csv.DictReader(io.StringIO(text))
    # Normalise column names once: strip whitespace from all header keys
    for row in reader:
        clean = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
        sym = clean.get("SYMBOL", "").strip()
        name = (clean.get("NAME OF COMPANY") or clean.get("COMPANY NAME") or "").strip()
        date_str = (clean.get("DATE OF LISTING") or "").strip()
        # INDUSTRY column: present in some NSE data exports, absent in EQUITY_L.csv
        sector = (clean.get("INDUSTRY") or clean.get("SECTOR") or clean.get("MACRO SECTOR")
                  or clean.get("BASIC INDUSTRY") or "").strip()
        if sym:
            _COMPANY_MAP[sym] = name or sym
            if date_str:
                _LISTING_DATE_MAP[sym] = date_str
            if sector:
                _SECTOR_MAP[sym] = sector


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 2 — SMART CACHE + INCREMENTAL DOWNLOAD ENGINE
# ═════════════════════════════════════════════════════════════════════════════
#
#  Cache layout (v2):
#  {
#    "__version__": 2,
#    "TICKER": {
#        "df":         pd.DataFrame,   # full OHLCV history, daily
#        "last_date":  "YYYY-MM-DD",   # most recent bar date
#        "marketcap":  float | None,   # latest known market cap (INR)
#        "mcap_ts":    float,          # unix timestamp when mcap was cached
#    },
#    ...
#  }
#
#  On every run:
#   • Cached + fresh (last_date >= yesterday): used as-is → 0 downloads
#   • Cached + stale (last_date < yesterday): batch-download last 60d, append
#   • New (never seen): batch-download "max"
#   • Market cap: re-fetched only if missing or >MCAP_TTL_DAYS old
#
# ═════════════════════════════════════════════════════════════════════════════

import concurrent.futures as _cf
from datetime import date as _date, timedelta as _td

MCAP_TTL_DAYS  = 7          # refresh market cap if older than this many days
DL_BATCH_SIZE  = 50         # tickers per yf.download() batch call
DL_MAX_WORKERS = 8          # parallel threads for batch downloads
CACHE_SAVE_INT = 200        # save cache to disk every N tickers processed
STALE_BUCKET_DAYS = 7       # group stale tickers whose missing ranges are close

_CACHE: dict = {}           # module-level, loaded once by _load_cache_v2()
_CACHE_DIRTY  = False       # set True whenever _CACHE is modified

# ── Helpers ────────────────────────────────────────────────────────────────

def _today_str() -> str:
    return _date.today().isoformat()            # "YYYY-MM-DD"


def _last_trading_day_str() -> str:
    today = _date.today()
    if today.weekday() == 0:      # Monday → last trading day was Friday
        return (today - _td(days=3)).isoformat()
    if today.weekday() == 6:      # Sunday → last trading day was Friday
        return (today - _td(days=2)).isoformat()
    if today.weekday() == 5:      # Saturday → last trading day was Friday
        return (today - _td(days=1)).isoformat()
    return (today - _td(days=1)).isoformat()


def _is_fresh(entry: dict) -> bool:
    """True if cached data already has the most recent likely market bar."""
    last = entry.get("last_date", "")
    return last >= _last_trading_day_str()

# ── Load / save ────────────────────────────────────────────────────────────

def _load_cache_v2() -> dict:
    """Load cache, migrating old v1 format (plain dict of DataFrames) if needed."""
    if not os.path.exists(CACHE_FILE):
        return {}
    try:
        with open(CACHE_FILE, "rb") as fh:
            raw = pickle.load(fh)
    except Exception as e:
        print(f"  [!] Cache load error: {e} — starting fresh")
        return {}

    # ── Migrate old format (dict of DataFrames, no __version__) ──────────────
    if isinstance(raw, dict) and raw.get("__version__") != 2:
        migrated = {"__version__": 2}
        migrated_count = 0
        for k, v in raw.items():
            if k.startswith("__"):
                continue
            if isinstance(v, pd.DataFrame) and not v.empty:
                last = v.index[-1].date().isoformat() if len(v) else ""
                migrated[k] = {"df": v, "last_date": last,
                                "marketcap": None, "mcap_ts": 0.0}
                migrated_count += 1
        if migrated_count:
            print(f"  🔄 Migrated {migrated_count} cached stocks from old cache format → v2")
        return migrated

    # ── Repair mixed-state v2 cache (some entries may still be raw DataFrames) ─
    fixed_count = 0
    for k, v in list(raw.items()):
        if k.startswith("__"):
            continue
        if isinstance(v, pd.DataFrame):
            last = v.index[-1].date().isoformat() if len(v) else ""
            raw[k] = {"df": v, "last_date": last, "marketcap": None, "mcap_ts": 0.0}
            fixed_count += 1
    if fixed_count:
        print(f"  🔧 Repaired {fixed_count} unmigrated DataFrame entries in v2 cache")

    return raw

def _save_cache_v2():
    """Write current _CACHE to disk."""
    global _CACHE_DIRTY
    try:
        with open(CACHE_FILE, "wb") as fh:
            pickle.dump(_CACHE, fh, protocol=pickle.HIGHEST_PROTOCOL)
        _CACHE_DIRTY = False
    except Exception as e:
        print(f"  [!] Cache save error: {e}")

def _maybe_save_cache():
    """Save only if dirty."""
    if _CACHE_DIRTY:
        _save_cache_v2()

# ── In-memory accessors ────────────────────────────────────────────────────

def _get_df(ticker: str) -> pd.DataFrame | None:
    entry = _CACHE.get(ticker)
    if isinstance(entry, dict):
        return entry.get("df")
    # Old v1 format: entry itself is a DataFrame
    if isinstance(entry, pd.DataFrame):
        return entry if not entry.empty else None
    return None

def _set_df(ticker: str, df: pd.DataFrame,
            marketcap: float | None = None, mcap_ts: float | None = None):
    """Update in-memory cache only. Caller must trigger _maybe_save_cache()."""
    global _CACHE_DIRTY
    # Guard: if existing entry is a raw DataFrame (old v1 format not fully migrated),
    # replace it with a fresh dict to avoid pandas __setitem__ crash
    existing = _CACHE.get(ticker)
    if not isinstance(existing, dict):
        _CACHE[ticker] = {}
    entry = _CACHE[ticker]
    # Flatten MultiIndex columns if yf.download returned them for single ticker
    if isinstance(df.columns, pd.MultiIndex):
        try:
            df = df.droplevel(1, axis=1)
        except Exception:
            pass
        df = df.dropna(subset=["Close"]) if "Close" in df.columns else df
    entry["df"]        = df
    entry["last_date"] = df.index[-1].date().isoformat() if len(df) else ""
    if marketcap is not None:
        entry["marketcap"] = marketcap
        entry["mcap_ts"]   = mcap_ts if mcap_ts is not None else time.time()
    _CACHE_DIRTY = True

def get_marketcap(ticker: str) -> float | None:
    entry = _CACHE.get(ticker)
    if isinstance(entry, dict):
        mc  = entry.get("marketcap")
        ts  = entry.get("mcap_ts", 0.0)
        age = (time.time() - ts) / 86400
        if mc and age < MCAP_TTL_DAYS:
            return mc
    return _MARKETCAP_MAP.get(ticker)

# ── DataFrame merger ────────────────────────────────────────────────────────

def _merge_df(old_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """Append new bars to old_df, deduplicate by date index, return sorted."""
    if new_df is None or new_df.empty:
        return old_df
    if old_df is None or old_df.empty:
        return new_df
    combined = pd.concat([old_df, new_df])
    combined = combined[~combined.index.duplicated(keep="last")]
    return combined.sort_index()

# ── Single-ticker clean download ────────────────────────────────────────────

def _clean_df(raw) -> pd.DataFrame:
    """Flatten MultiIndex columns safely regardless of yfinance version."""
    if raw is None or raw.empty:
        return pd.DataFrame()
    df = raw.copy()
    if isinstance(df.columns, pd.MultiIndex):
        l0_vals = df.columns.get_level_values(0).tolist()
        if any(c in l0_vals for c in ("Open", "High", "Low", "Close", "Volume")):
            df.columns = df.columns.droplevel(1)   # yfinance >= 0.2.38: metric at level 0
        else:
            df.columns = df.columns.droplevel(0)   # older format: ticker at level 0
    return df.dropna(subset=["Close"])

# ── Batch download ──────────────────────────────────────────────────────────

def _bucket_start_date(dt: _date, bucket_days: int = STALE_BUCKET_DAYS) -> str:
    epoch = _date(1970, 1, 1)
    offset = ((dt - epoch).days // bucket_days) * bucket_days
    return (epoch + _td(days=offset)).isoformat()


def _batch_download(tickers: list[str], period: str | None = None,
                    start: str | None = None, end: str | None = None) -> dict[str, pd.DataFrame]:
    """
    Download multiple tickers in one yf.download() call.
    Returns {ticker: clean_DataFrame}. Missing/failed tickers are omitted.

    Use either `period` or `start`/`end` to control the range.
    ASX tickers (in _ASX_STOCKS) get the ".AX" suffix; all others get ".NS".
    """
    if not tickers:
        return {}
    results = {}

    def _yf_symbol(t: str) -> str:
        return t + ".AX" if t in _ASX_STOCKS else t + ".NS"

    args = {
        "tickers": [_yf_symbol(t) for t in tickers],
        "interval": "1d",
        "progress": False,
        "auto_adjust": True,
        "group_by": "ticker",
        "threads": True,
    }
    if start and end:
        args["start"] = start
        args["end"] = end
    elif period:
        args["period"] = period
    else:
        args["period"] = DATA_PERIOD

    import io as _io
    _old_stderr = sys.stderr
    _capture_stderr = _io.StringIO()
    sys.stderr = _capture_stderr   # silence yfinance "Failed downloads" noise
    try:
        raw = yf.download(**args)
    except Exception as e:
        sys.stderr = _old_stderr
        label = f"{period or f'{start}:{end}'}"
        print(f"\n  [!] Batch download error ({len(tickers)} tickers, {label}): {e}")
        return {}
    finally:
        _yf_noise = _capture_stderr.getvalue()
        sys.stderr = _old_stderr
        for _line in _yf_noise.splitlines():
            if _line.strip() and "YFTzMissing" not in _line and "Failed downloads" not in _line:
                print(f"  [yf] {_line}", file=sys.stderr)

    if raw.empty:
        return {}

    for ticker in tickers:
        yf_sym = _yf_symbol(ticker)
        try:
            if isinstance(raw.columns, pd.MultiIndex):
                if yf_sym in raw.columns.get_level_values(0):
                    df = raw[yf_sym].copy().dropna(subset=["Close"])
                elif ticker in raw.columns.get_level_values(0):
                    df = raw[ticker].copy().dropna(subset=["Close"])
                else:
                    continue
            else:
                # Only one ticker returned → flat columns
                df = raw.copy().dropna(subset=["Close"])
            if not df.empty:
                results[ticker] = df
        except Exception:
            continue

    return results

# ── Market cap batch fetch ──────────────────────────────────────────────────

def _fetch_mcap_batch(tickers: list[str]) -> dict[str, float]:
    """Fetch market cap for multiple tickers using threads. Returns {ticker: mcap}."""
    results = {}

    def _fetch_one(t):
        try:
            suffix = ".AX" if t in _ASX_STOCKS else ".NS"
            info = yf.Ticker(t + suffix).fast_info
            mc   = getattr(info, "market_cap", None)
            if mc and mc > 0:
                return t, float(mc)
        except Exception:
            pass
        return t, None

    with _cf.ThreadPoolExecutor(max_workers=DL_MAX_WORKERS) as exe:
        for ticker, mc in exe.map(_fetch_one, tickers):
            if mc:
                results[ticker] = mc
    return results

# ── Main pre-fetch entry point (called from main() before analyze loop) ────

def prefetch_all(tickers: list[str]) -> dict[str, int]:
    """
    Smart incremental download for the full universe.

    Returns a stats dict: {fresh, stale_updated, new_downloaded, failed}

    After this function:
      - Every reachable ticker has up-to-date data in _CACHE
      - _CACHE may be dirty; caller should call _maybe_save_cache() when done
    """
    global _CACHE_DIRTY

    today = _today_str()

    fresh_tickers   = []   # cache data already has last trading day → skip
    stale_tickers   = []   # in cache but older than last trading day
    new_tickers     = []   # not in cache at all → fetch max

    for t in tickers:
        entry = _CACHE.get(t)
        if isinstance(entry, dict) and isinstance(entry.get("df"), pd.DataFrame):
            if _is_fresh(entry):
                fresh_tickers.append(t)
            else:
                stale_tickers.append(t)
        else:
            new_tickers.append(t)

    total = len(tickers)
    print(f"\n  📦 Cache status:  "
          f"✅ fresh={len(fresh_tickers)}  "
          f"🔄 stale={len(stale_tickers)}  "
          f"🆕 new={len(new_tickers)}  "
          f"(total={total})")

    stats = {"fresh": len(fresh_tickers), "stale_updated": 0,
             "new_downloaded": 0, "failed": 0}

    # ── Update stale tickers (fetch last 60d and append) ────────────────────
    if stale_tickers:
        print(f"  🔄 Updating {len(stale_tickers)} stale tickers (download only missing dates)…")

        groups: dict[str, list[str]] = {}
        for ticker in stale_tickers:
            entry = _CACHE.get(ticker, {})
            last_date = entry.get("last_date")
            if not last_date:
                start_dt = _date.today() - _td(days=30)
            else:
                start_dt = _date.fromisoformat(last_date) + _td(days=1)
            if start_dt >= _date.today():
                groups.setdefault("SKIP", []).append(ticker)
                continue
            bucket_key = _bucket_start_date(start_dt)
            groups.setdefault(bucket_key, []).append(ticker)

        group_items = sorted((k, v) for k, v in groups.items() if k != "SKIP")
        processed = 0
        total_groups = len(group_items)
        for group_index, (bucket_start, group) in enumerate(group_items, start=1):
            start_date = bucket_start
            end_date = (_date.today() + _td(days=1)).isoformat()
            print(f"  • Bucket {group_index}/{total_groups}: start={start_date} tickers={len(group)}")

            for batch_start in range(0, len(group), DL_BATCH_SIZE):
                batch = group[batch_start: batch_start + DL_BATCH_SIZE]
                pct = min(100, (processed + len(batch)) / len(stale_tickers) * 100)
                sys.stdout.write(
                    f"\r    Stale [{pct:5.1f}%]  batch {processed//DL_BATCH_SIZE+1}"
                    f"/{-(-len(stale_tickers)//DL_BATCH_SIZE)}  "
                    f"({processed+len(batch)}/{len(stale_tickers)})"
                )
                sys.stdout.flush()

                new_data = _batch_download(batch, start=start_date, end=end_date)
                for ticker in batch:
                    if ticker not in new_data:
                        stats["failed"] += 1
                        continue
                    old_df = _get_df(ticker)
                    if old_df is None:
                        old_df = pd.DataFrame()
                    merged  = _merge_df(old_df, new_data[ticker])
                    if len(merged) >= MIN_CANDLES:  # MIN_CANDLES=1: accept all
                        try:
                            _set_df(ticker, merged)
                            stats["stale_updated"] += 1
                        except Exception as _e:
                            print(f"\n  [!] Cache write error for {ticker}: {_e}")
                            stats["failed"] += 1
                            stats.setdefault("failed_list", []).append(ticker)
                    else:
                        # df has 0 rows — truly empty, skip
                        stats["failed"] += 1
                        stats.setdefault("failed_list", []).append(ticker)

                processed += len(batch)
                if _CACHE_DIRTY and (processed) % CACHE_SAVE_INT == 0:
                    _save_cache_v2()
        if "SKIP" in groups:
            for ticker in groups["SKIP"]:
                print(f"    Stale [SKIP] {ticker} already up-to-date")
        print()

    # ── Download new tickers (full history) ─────────────────────────────────
    if new_tickers:
        print(f"  🆕 Downloading {len(new_tickers)} new tickers (full history)…")
        for batch_start in range(0, len(new_tickers), DL_BATCH_SIZE):
            batch = new_tickers[batch_start: batch_start + DL_BATCH_SIZE]
            pct   = min(100, (batch_start + len(batch)) / len(new_tickers) * 100)
            sys.stdout.write(f"\r    New   [{pct:5.1f}%]  batch {batch_start//DL_BATCH_SIZE+1}"
                             f"/{-(-len(new_tickers)//DL_BATCH_SIZE)}  "
                             f"({batch_start+len(batch)}/{len(new_tickers)})")
            sys.stdout.flush()

            new_data = _batch_download(batch, period=DATA_PERIOD)
            for ticker in batch:
                if ticker not in new_data:
                    stats["failed"] += 1
                    stats.setdefault("failed_list", []).append(ticker)
                    continue
                df = new_data[ticker]
                if len(df) >= MIN_CANDLES:  # MIN_CANDLES=1: accept all non-empty
                    try:
                        _set_df(ticker, df)
                        stats["new_downloaded"] += 1
                    except Exception as _e:
                        print(f"\n  [!] Cache write error for {ticker}: {_e}")
                        stats["failed"] += 1
                        stats.setdefault("failed_list", []).append(ticker)
                else:
                    # df has 0 rows — truly empty, skip
                    stats["failed"] += 1
                    stats.setdefault("failed_list", []).append(ticker)

            if _CACHE_DIRTY and (batch_start + DL_BATCH_SIZE) % CACHE_SAVE_INT == 0:
                _save_cache_v2()
        print()

    # ── Market cap refresh (batch, threaded) ────────────────────────────────
    mcap_stale = [
        t for t in tickers
        if t in _CACHE and isinstance(_CACHE[t], dict)
        and (not _CACHE[t].get("marketcap")
             or (time.time() - _CACHE[t].get("mcap_ts", 0)) / 86400 > MCAP_TTL_DAYS)
    ]
    if mcap_stale:
        print(f"  💰 Refreshing market cap for {len(mcap_stale)} tickers… ", end="", flush=True)
        mcap_map = _fetch_mcap_batch(mcap_stale)
        now = time.time()
        for t, mc in mcap_map.items():
            if t in _CACHE and isinstance(_CACHE[t], dict):
                _CACHE[t]["marketcap"] = mc
                _CACHE[t]["mcap_ts"]   = now
                _MARKETCAP_MAP[t]      = mc
        _CACHE_DIRTY = True
        print(f"updated {len(mcap_map)}/{len(mcap_stale)}")

    # ── Final save ───────────────────────────────────────────────────────────
    _maybe_save_cache()

    cached_total = sum(1 for v in _CACHE.values()
                       if isinstance(v, dict) and isinstance(v.get("df"), pd.DataFrame))
    failed_list = stats.get("failed_list", [])
    if failed_list:
        print(f"\n  ⚠️  {len(failed_list)} tickers unavailable (delisted/no data): "
              f"{', '.join(failed_list[:20])}{'...' if len(failed_list)>20 else ''}")
        log_info(f"Failed tickers ({len(failed_list)}): {failed_list}")
    print(f"\n  ✅ Cache ready: {cached_total} tickers on disk  |  "
          f"updated {stats['stale_updated']+stats['new_downloaded']}  |  "
          f"failed {stats['failed']}\n")
    return stats


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 3 — INDICATORS
# ═════════════════════════════════════════════════════════════════════════════

def calc_rsi(close, period=14):
    delta = close.diff()
    gain  = delta.clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    loss  = (-delta).clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    return 100 - (100 / (1 + gain / (loss + 1e-10)))

def calc_macd(close, fast=12, slow=26, sig=9):
    line   = close.ewm(span=fast, adjust=False).mean() - close.ewm(span=slow, adjust=False).mean()
    signal = line.ewm(span=sig, adjust=False).mean()
    return line, signal, line - signal

def calc_cci(high, low, close, period=20):
    tp  = (high + low + close) / 3
    sma = tp.rolling(period).mean()
    mad = tp.rolling(period).apply(lambda x: np.abs(x - x.mean()).mean(), raw=True)
    return (tp - sma) / (0.015 * mad + 1e-10)

def calc_atr(high, low, close, period=14):
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs(),
    ], axis=1).max(axis=1)
    return tr.ewm(alpha=1 / period, adjust=False).mean()

def calc_bb(close, period=20, std_mult=2.0):
    """Bollinger Bands: returns (upper, mid, lower) series."""
    mid   = close.rolling(period).mean()
    std   = close.rolling(period).std(ddof=0)
    upper = mid + std_mult * std
    lower = mid - std_mult * std
    return upper, mid, lower

def calc_mfi(high, low, close, volume, period=14):
    """Money Flow Index (0-100): institutional buying/selling pressure."""
    tp      = (high + low + close) / 3
    raw_mf  = tp * volume
    pos_mf  = raw_mf.where(tp > tp.shift(1), 0.0)
    neg_mf  = raw_mf.where(tp < tp.shift(1), 0.0)
    pos_sum = pos_mf.rolling(period).sum()
    neg_sum = neg_mf.rolling(period).sum()
    mfr     = pos_sum / (neg_sum + 1e-10)
    return 100 - (100 / (1 + mfr))

def resample_ohlcv(df, rule):
    return df.resample(rule).agg({
        "Open": "first", "High": "max", "Low": "min",
        "Close": "last", "Volume": "sum",
    }).dropna(subset=["Close"])


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 4 — SWING + FIBONACCI
# ═════════════════════════════════════════════════════════════════════════════

def find_swing_points(high_s, low_s, lookback=252, order=5):
    h = high_s.iloc[-lookback:] if len(high_s) >= lookback else high_s
    l = low_s.iloc[-lookback:]  if len(low_s)  >= lookback else low_s
    pivot_highs, pivot_lows = [], []
    for i in range(order, len(h) - order):
        if h.iloc[i] == h.iloc[i - order: i + order + 1].max():
            pivot_highs.append((h.index[i], float(h.iloc[i]), i))
        if l.iloc[i] == l.iloc[i - order: i + order + 1].min():
            pivot_lows.append((l.index[i], float(l.iloc[i]), i))
    if not pivot_highs or not pivot_lows:
        phi = h.idxmax(); plo = l.idxmin()
        return float(l.loc[plo]), plo, float(h.loc[phi]), phi
    sh_dt, sh_val, _ = max(pivot_highs, key=lambda x: x[2])
    sl_dt, sl_val, _ = max(pivot_lows,  key=lambda x: x[2])
    return sl_val, sl_dt, sh_val, sh_dt

def fib_extensions(swing_low, swing_high):
    rng = swing_high - swing_low
    return {
        "127.2%": round(swing_high + rng * 0.272, 2),
        "161.8%": round(swing_high + rng * 0.618, 2),
        "200.0%": round(swing_high + rng * 1.000, 2),
        "261.8%": round(swing_high + rng * 1.618, 2),
        "423.6%": round(swing_high + rng * 3.236, 2),
    }

def fib_retracements(swing_high, swing_low):
    rng = swing_high - swing_low
    return {
        "23.6%": round(swing_high - rng * 0.236, 2),
        "38.2%": round(swing_high - rng * 0.382, 2),
        "50.0%": round(swing_high - rng * 0.500, 2),
        "61.8%": round(swing_high - rng * 0.618, 2),
        "78.6%": round(swing_high - rng * 0.786, 2),
    }


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 5 — PHASE + SCORING
# ═════════════════════════════════════════════════════════════════════════════

def detect_phase(rsi_d, rsi_w, rsi_m, macd_line, macd_signal, score):
    bulls = sum([rsi_d>55, rsi_w>52, rsi_m>50, macd_line>macd_signal, macd_line>0])
    bears = sum([rsi_d<45, rsi_w<48, rsi_m<50, macd_line<macd_signal, macd_line<0])
    if bulls >= 4 or (score >= SCORE_BUY and rsi_d > 50):  return "UPTREND"
    if bears >= 4 or (score <= 5 and rsi_d < 45):           return "BEARISH"
    return "SIDEWAYS"

def compute_score(rsi_d, rsi_d_sma, rsi_w, rsi_w_sma, rsi_m, rsi_m_sma,
                  macd_line, macd_sig, cci, fresh_d, fresh_w):
    score, sigs = 0, []
    if rsi_m > rsi_m_sma: score += 4; sigs.append("M-RSI>SMA ✅")
    if rsi_w > rsi_w_sma: score += 3; sigs.append("W-RSI>SMA ✅")
    if rsi_d > rsi_d_sma: score += 2; sigs.append("D-RSI>SMA ✅")
    if fresh_d:            score += 3; sigs.append("FRESH Daily 🚀")
    if fresh_w:            score += 2; sigs.append("FRESH Weekly 🔥")
    if macd_line > macd_sig: score += 2; sigs.append("MACD>Sig ✅")
    if macd_line > 0:        score += 1; sigs.append("MACD>0")
    if cci > 100:   score += 2; sigs.append("CCI>100 💪")
    elif cci > 0:   score += 1; sigs.append("CCI>0")
    if rsi_d > 60:  score += 1; sigs.append("D-RSI>60 🔥")
    if rsi_w > 55:  score += 1; sigs.append("W-RSI>55 💪")
    return score, sigs


def compute_explosive_score(rsi_d, vol_ratio, close, bb_upper, bb_slope,
                             macd_hist, macd_hist_prev, mfi, cci_200, donchian_d):
    """
    Explosive Breakout Score (0-12): catches MTAR/HFCL/Adani-type 50-100% moves.
    Weighted combination of volume surge, BB breakout, MACD acceleration,
    institutional MFI, long-term CCI(200), and Donchian channel breakout.
    RSI > 40 is a prerequisite filter, not a scored component here.
    """
    escore, sigs = 0, []

    # ── RSI momentum headroom ─────────────────────────────────────────────
    if rsi_d > 70:   escore += 2; sigs.append("RSI>70 🔥🔥")
    elif rsi_d > 60: escore += 1; sigs.append("RSI>60 🔥")

    # ── Volume surge (primary explosive signal) ───────────────────────────
    if vol_ratio >= 5.0:   escore += 3; sigs.append(f"Vol {vol_ratio:.1f}x 🚀🚀🚀")
    elif vol_ratio >= 2.5: escore += 2; sigs.append(f"Vol {vol_ratio:.1f}x 🚀🚀")
    elif vol_ratio >= 1.5: escore += 1; sigs.append(f"Vol {vol_ratio:.1f}x 🚀")

    # ── Bollinger Band breakout + slope ───────────────────────────────────
    if close > bb_upper:    escore += 2; sigs.append("BB Breakout ⚡")
    elif bb_slope > 1.0:    escore += 1; sigs.append("BB Steep 📈")

    # ── MACD histogram acceleration ───────────────────────────────────────
    if macd_hist > 0 and macd_hist > macd_hist_prev:
        escore += 2; sigs.append("MACD Accel ✅")
    elif macd_hist > 0:
        escore += 1; sigs.append("MACD+ ✅")

    # ── MFI: institutional money flow ─────────────────────────────────────
    if mfi > 70:   escore += 2; sigs.append("MFI>70 💪💪")
    elif mfi > 50: escore += 1; sigs.append("MFI>50 💪")

    # ── CCI(200): long-term trend alignment ───────────────────────────────
    if cci_200 > 100:  escore += 2; sigs.append("CCI200>100 💥")
    elif cci_200 > 0:  escore += 1; sigs.append("CCI200>0 ✅")

    # ── Donchian 20-day channel breakout ──────────────────────────────────
    if donchian_d is not None and donchian_d >= -0.5:
        escore += 1; sigs.append("D20 Break 🎯")

    return min(escore, 12), sigs

def signal_label(score, phase, fresh_d, fresh_w, rsi_d, rsi_w, rsi_m,
                 rsi_d_sma, rsi_w_sma, rsi_m_sma):
    triple = rsi_d > rsi_d_sma and rsi_w > rsi_w_sma and rsi_m > rsi_m_sma
    if score >= SCORE_STRONG_BUY and triple and (fresh_d or fresh_w):
        return "STRONG BUY 🚀", "sig-strong-buy"
    if score >= SCORE_BUY and rsi_d > rsi_d_sma and rsi_w > rsi_w_sma:
        return "BUY ✅", "sig-buy"
    if score >= SCORE_WATCH:
        return "WATCH 👀", "sig-watch"
    if phase == "BEARISH":
        return "AVOID ❌", "sig-avoid"
    return "NEUTRAL", "sig-neutral"


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 6 — RANKING ENGINE
# ═════════════════════════════════════════════════════════════════════════════

def compute_rankings(results: list[dict]) -> list[dict]:
    """
    Add two ranking fields to every result dict:

    rank_nifty50    — percentile (0-100) of this stock's score vs the
                      subset of Nifty50 stocks that were successfully analysed.
                      100 = best among Nifty50, 0 = worst.

    rank_universe   — percentile (0-100) of this stock's score vs ALL
                      scanned stocks.  100 = top 1% of entire universe.

    Also adds:
    rank_nifty50_pos  — integer rank  (1 = best Nifty50 stock)
    rank_nifty50_of   — total Nifty50 stocks in scan
    rank_univ_pos     — integer rank in full universe
    rank_univ_of      — total stocks in universe
    """
    n50_set = set(NIFTY50)

    # ── All scores ─────────────────────────────────────────────
    all_scores  = [d["score"] for d in results]
    n50_results = [d for d in results if d["ticker"] in n50_set]
    n50_scores  = [d["score"] for d in n50_results]

    def pct_rank(score, score_list):
        """Percentile rank: what % of scores are ≤ this score."""
        if not score_list:
            return 0
        below = sum(1 for s in score_list if s <= score)
        return round(below / len(score_list) * 100, 1)

    # Sort for integer rank (1 = highest score)
    sorted_all = sorted(results, key=lambda d: d["score"], reverse=True)
    sorted_n50 = sorted(n50_results,  key=lambda d: d["score"], reverse=True)

    rank_all_map = {d["ticker"]: i + 1 for i, d in enumerate(sorted_all)}
    rank_n50_map = {d["ticker"]: i + 1 for i, d in enumerate(sorted_n50)}

    for d in results:
        d["rank_nifty50"]     = pct_rank(d["score"], n50_scores)
        d["rank_universe"]    = pct_rank(d["score"], all_scores)
        d["rank_nifty50_pos"] = rank_n50_map.get(d["ticker"], 0)
        d["rank_nifty50_of"]  = len(n50_results)
        d["rank_univ_pos"]    = rank_all_map.get(d["ticker"], 0)
        d["rank_univ_of"]     = len(results)
        d["is_nifty50"]       = d["ticker"] in n50_set

    return results


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 7 — HISTORICAL SIGNALS
# ═════════════════════════════════════════════════════════════════════════════

def historical_signals(close, rsi_series, rsi_sma_series, max_signals=12):
    results = []
    rsi_arr = rsi_series.values
    sma_arr = rsi_sma_series.values
    cls_arr = close.values
    dates   = close.index
    for i in range(1, len(rsi_arr)):
        if np.isnan(rsi_arr[i]) or np.isnan(sma_arr[i]):
            continue
        crossed_above = rsi_arr[i] > sma_arr[i] and rsi_arr[i-1] <= sma_arr[i-1]
        crossed_below = rsi_arr[i] < sma_arr[i] and rsi_arr[i-1] >= sma_arr[i-1]
        if not (crossed_above or crossed_below):
            continue
        sig_type = "BUY" if crossed_above else "SELL"
        sig_px   = float(cls_arr[i])
        def ret(fwd):
            j = i + fwd
            return round((cls_arr[j] / sig_px - 1) * 100, 1) if j < len(cls_arr) else None
        results.append({
            "date": dates[i].strftime("%d-%b-%y"), "type": sig_type,
            "price": round(sig_px, 2), "rsi": round(float(rsi_arr[i]), 1),
            "r5d": ret(5), "r10d": ret(10), "r20d": ret(20),
        })
    return results[-max_signals:]


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 8 — NSE UNIVERSE LOADER
# ═════════════════════════════════════════════════════════════════════════════

BUILTIN = [
    "RELIANCE","TCS","HDFCBANK","INFY","ICICIBANK","HINDUNILVR","ITC","SBIN",
    "BAJFINANCE","BHARTIARTL","KOTAKBANK","LT","AXISBANK","ASIANPAINT","MARUTI",
    "SUNPHARMA","TITAN","WIPRO","ULTRACEMCO","NTPC","POWERGRID","ONGC","JSWSTEEL",
    "TATASTEEL","COALINDIA","TECHM","HCLTECH","DRREDDY","CIPLA","DIVISLAB",
    "ADANIENT","ADANIGREEN","ADANIPORTS","ADANIPOWER","TATAPOWER","GAIL","IOC",
    "BPCL","HINDPETRO","HINDALCO","VEDL","HINDZINC","NATIONALUM","NMDC",
    "BANKBARODA","PNB","CANBK","FEDERALBNK","RBLBANK","BANDHANBNK","INDUSINDBK",
    "MUTHOOTFIN","BAJAJFINSV","CHOLAFIN","HDFCLIFE","ICICIGI","SBILIFE","LICI",
    "TATAMOTORS","M&M","BAJAJ-AUTO","HEROMOTOCO","EICHERMOT","SIEMENS","ABB",
    "BHEL","HAVELLS","VOLTAS","POLYCAB","RVNL","HAL","BEL","BEML",
    "AUROPHARMA","LUPIN","TORNTPHARM","ALKEM","APOLLOHOSP","FORTIS",
    "TATACONSUM","NESTLEIND","BRITANNIA","DABUR","GODREJCP","DMART","TRENT",
    "DEEPAKNTR","PIIND","SRF","TATACHEM","PERSISTENT","COFORGE","LTIM","NAUKRI",
    "DLF","GODREJPROP","PHOENIXLTD","ZOMATO","IRCTC","PIDILITIND","KALYANKJIL",
    "ATUL","NAVINFLUOR","VINATI","CLEAN","DIXON","AMBER","TATAELXSI",
    "IRFC","RECLTD","PFC","JSWENERGY","TORNTPOWER","CESC","NHPC","SJVN",
    "ANGELONE","BSE","CDSL","CAMS","MOFSL","JUBLFOOD","RADICO","MCDOWELL-N",
    "GRSE","COCHINSHIP","RAILTEL","TITAGARH","DATAPATTNS","KEC","KALPATPOWR",
    "JINDALSTEL","JSL","SAIL","APLAPOLLO","GRAPHITE","APARINDS","CUMMINSIND",
    "ELGIEQUIP","GRINDWELL","EXIDEIND","MOTHERSON","BOSCHLTD","MRF","APOLLOTYRE",
    "LALPATHLAB","METROPOLIS","MAXHEALTH","PAGEIND","COLPAL","EMAMILTD",
    "HDFCAMC","NIPPONLIFE","ABSLAMC","SBICARD","OBEROIRLTY","PRESTIGE","BRIGADE",
]

def _load_fo_list():
    """
    Load NSE F&O eligible securities into _FO_SET.

    Sources tried in order:
    1. LOCAL_FO_CSV  — cached CSV from a previous download or user-placed file
                        Columns expected: SYMBOL (+ optional INSTRUMENT, MARKET_TYPE)
    2. NSE API       — https://www.nseindia.com/api/foSecList
                        Auto-downloaded and cached to LOCAL_FO_CSV on success.

    The CSV can be generated once with:
        import requests, pandas as pd
        s = requests.Session()
        s.headers.update({"User-Agent": "Mozilla/5.0", "Referer": "https://www.nseindia.com/"})
        s.get("https://www.nseindia.com/")
        r = s.get("https://www.nseindia.com/api/foSecList")
        df = pd.DataFrame(r.json()["data"])
        df.to_csv("india/NSE/nse_fo_list.csv", index=False)
    """
    global _FO_SET

    # ── Try local CSV first ──────────────────────────────────────────────────
    if os.path.exists(LOCAL_FO_CSV):
        try:
            with open(LOCAL_FO_CSV, encoding="utf-8", errors="replace") as f:
                raw = f.read().lstrip("\ufeff")
            reader = csv.DictReader(io.StringIO(raw))
            loaded = set()
            for row in reader:
                # Normalise column names (strip spaces)
                clean = {k.strip().upper(): (v.strip() if v else "") for k, v in row.items() if k}
                sym = clean.get("SYMBOL", "")
                if sym:
                    loaded.add(sym.upper())
            _FO_SET = loaded
            print(f"  ✅ F&O list loaded : {len(_FO_SET)} symbols ← '{LOCAL_FO_CSV}'")
            return
        except Exception as e:
            print(f"  [!] Error reading '{LOCAL_FO_CSV}': {e}")

    # ── Try NSE API ───────────────────────────────────────────────────────────
    print(f"  ℹ️  '{LOCAL_FO_CSV}' not found — attempting NSE API download...")
    try:
        s = requests.Session()
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://www.nseindia.com/",
        })
        s.get("https://www.nseindia.com/", timeout=10)          # seed cookies
        r = s.get("https://www.nseindia.com/api/foSecList", timeout=15)
        r.raise_for_status()
        data = r.json()
        rows = data.get("data", data) if isinstance(data, dict) else data
        if not rows:
            raise ValueError("Empty response from foSecList API")

        symbols = {str(row.get("SYMBOL", "")).strip().upper()
                   for row in rows if row.get("SYMBOL")}
        symbols.discard("")
        _FO_SET = symbols
        print(f"  ✅ F&O list fetched : {len(_FO_SET)} symbols from NSE API")

        # Cache to CSV for future runs
        try:
            os.makedirs(os.path.dirname(LOCAL_FO_CSV), exist_ok=True)
            import pandas as pd
            df = pd.DataFrame(rows)
            df.to_csv(LOCAL_FO_CSV, index=False)
            print(f"  💾 F&O list cached  : '{LOCAL_FO_CSV}'")
        except Exception as ce:
            print(f"  [!] Could not cache F&O list: {ce}")

    except Exception as e:
        print(f"  [!] NSE API failed ({e}) — F&O filter unavailable")
        _FO_SET = set()


def _parse_index_symbols(rows: list[tuple]) -> tuple[int, set[str]]:
    """
    Auto-detect symbol column and return (header_row_idx, set_of_symbols).
    Handles NSE-style CSV: Symbol, Company Name, Series, ISIN Code
    """
    SKIP = {"SYMBOL","SYMBOLS","TICKER","COMPANY","COMPANY NAME","ISIN",
            "ISIN CODE","NAME","SERIES","SECURITY","INDEX NAME","SCRIP","N/A","NA","-",""}
    if not rows:
        return 0, set()
    header_row_idx = 0
    sym_col = None
    first = [str(c).strip().upper() if c else '' for c in rows[0]]
    for ci, val in enumerate(first):
        if val in ("SYMBOL","SYMBOLS","TICKER","SCRIP","NSE SYMBOL"):
            sym_col = ci
            header_row_idx = 1
            break
    if sym_col is None:
        if first[0] if first else '' in SKIP:
            header_row_idx = 1
        sym_col = 0
    symbols: set[str] = set()
    for row in rows[header_row_idx:]:
        if not row: continue
        cell = row[sym_col] if len(row) > sym_col else None
        if not cell: continue
        sym = str(cell).strip().upper()
        if sym and sym not in SKIP and not sym.isdigit() and len(sym) <= 20:
            symbols.add(sym)
    return header_row_idx, symbols


def _populate_indices():
    """
    Populate _INDEX_MAP and _SECTOR_MAP from NIFTY_Indices_Master.xlsx.

    _INDEX_MAP : { index_name -> set of symbols }  — all 52 indices
    _SECTOR_MAP: { symbol -> list of sector labels }
                 Every Sectoral / Thematic / Strategy index a symbol belongs to
                 becomes a separate sector label: "NIFTY CPSE - Thematic",
                 "NIFTY AUTO - Sectoral", "NIFTY ALPHA 50 - Strategy", etc.
                 Broad-Based indices (NIFTY 50, NIFTY 200 …) are NOT added here
                 because they are market-cap bands, not sector/industry groups.

    Sector dropdown shows all 38 Sectoral+Thematic+Strategy options.
    Selecting one shows every stock that belongs to that index.
    """
    _INDEX_MAP["NIFTY50"] = set(NIFTY50)   # hardcoded fallback always present

    CAT_SHORT = {
        "Sectoral Indices"   : "Sectoral",
        "Thematic Indices"   : "Thematic",
        "Strategy Indices"   : "Strategy",
        "Broad Based Indices": "",          # broad → index only, not sector
    }

    if not os.path.exists(LOCAL_INDICES_XLSX):
        print(f"  ⚠️  '{LOCAL_INDICES_XLSX}' not found")
        print(f"       Expected at: {os.path.abspath(LOCAL_INDICES_XLSX)}")
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(LOCAL_INDICES_XLSX, read_only=True, data_only=True)

            STOCK_SHEETS = [s for s in wb.sheetnames
                            if s.endswith('_Stocks') or s == 'All_Stocks_Combined']
            if not STOCK_SHEETS:
                STOCK_SHEETS = [s for s in wb.sheetnames if s != 'Summary']

            # sym → set of sector labels (one per qualifying index)
            _sym_sectors: dict[str, set[str]] = {}

            for sheet_name in STOCK_SHEETS:
                ws   = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 4:
                    continue

                # Locate header row that contains 'SYMBOL'
                header_idx = sym_col = idx_col = cat_col = None
                for ri, row in enumerate(rows[:6]):
                    if not row:
                        continue
                    norm = [str(c).strip().upper() if c else '' for c in row]
                    if 'SYMBOL' in norm:
                        header_idx = ri
                        sym_col    = norm.index('SYMBOL')
                        idx_col    = norm.index('INDEX NAME') if 'INDEX NAME' in norm else None
                        cat_col    = norm.index('CATEGORY')   if 'CATEGORY'   in norm else None
                        break

                if header_idx is None or sym_col is None:
                    continue

                for row in rows[header_idx + 1:]:
                    if not row or len(row) <= sym_col:
                        continue
                    sym = row[sym_col]
                    if not sym:
                        continue
                    sym = str(sym).strip().upper()
                    if not sym or sym.isdigit() or len(sym) > 20:
                        continue

                    idx_name = (str(row[idx_col]).strip()
                                if idx_col is not None and len(row) > idx_col and row[idx_col]
                                else sheet_name)
                    category = (str(row[cat_col]).strip()
                                if cat_col is not None and len(row) > cat_col and row[cat_col]
                                else '')

                    # Add to _INDEX_MAP (all categories)
                    if idx_name not in _INDEX_MAP:
                        _INDEX_MAP[idx_name] = set()
                    _INDEX_MAP[idx_name].add(sym)

                    # Add sector label for Sectoral / Thematic / Strategy only
                    short = CAT_SHORT.get(category, None)
                    if short is not None and short != '':
                        label = f"{idx_name} - {short}"
                        if sym not in _sym_sectors:
                            _sym_sectors[sym] = set()
                        _sym_sectors[sym].add(label)

            # Commit to _SECTOR_MAP as sorted list per symbol
            for sym, labels in _sym_sectors.items():
                _SECTOR_MAP[sym] = sorted(labels)

            # Remove sheet-name ghost keys
            for ghost in ['All_Stocks_Combined', 'Broad_Based_Stocks',
                          'Sectoral_Stocks', 'Thematic_Stocks', 'Strategy_Stocks']:
                _INDEX_MAP.pop(ghost, None)

            total_indices  = len(_INDEX_MAP)
            n_sec_labels   = len({l for ls in _SECTOR_MAP.values() for l in ls})
            print(f"  ✅ Indices loaded : {total_indices} indices, "
                  f"{sum(len(v) for v in _INDEX_MAP.values())} total memberships "
                  f"← '{LOCAL_INDICES_XLSX}'")
            print(f"  ✅ Sectors derived: {len(_SECTOR_MAP)} symbols → "
                  f"{n_sec_labels} unique sector labels (Sectoral+Thematic+Strategy)")

        except ImportError:
            print("  [!] openpyxl not installed — run: pip install openpyxl")
        except Exception as e:
            import traceback
            print(f"  [!] Error loading '{LOCAL_INDICES_XLSX}': {e}")
            traceback.print_exc()

    # ── Individual index CSVs from LOCAL_INDICES_DIR (optional) ──────────────
    import re   # needed for re.sub in idx_name extraction below
    # Supports any folder path; filenames like:
    #   ind_nifty50list.csv           → NIFTY50
    #   ind_nifty_adityabirlalist.csv → NIFTY ADITYABIRLA
    #   ind_nifty_Alpha_Index.csv     → NIFTY ALPHA INDEX
    #   ind_nifty_alpha_lowvol30list.csv → NIFTY ALPHA LOWVOL30
    # Strips:  leading "ind_" (case-insensitive), trailing "list" or "_list",
    #          then collapses underscores to spaces
    csv_loaded = 0
    if os.path.isdir(LOCAL_INDICES_DIR):
        for fname in sorted(os.listdir(LOCAL_INDICES_DIR)):
            if not fname.lower().endswith(".csv"):
                continue
            fpath = os.path.join(LOCAL_INDICES_DIR, fname)
            try:
                with open(fpath, encoding="utf-8", errors="replace") as f:
                    raw = f.read().lstrip('\ufeff')
                # ── Parse full CSV as DictReader to get Industry column ────
                reader_dict = csv.DictReader(io.StringIO(raw))
                sym_col_candidates = {"Symbol","SYMBOL","Symbols","SYMBOLS","Ticker","TICKER","NSE Symbol"}
                ind_col_candidates = {"Industry","INDUSTRY","Sector","SECTOR","industry","sector"}
                sym_col = None
                ind_col = None
                csv_rows_dict = []
                for row in reader_dict:
                    if sym_col is None:  # detect columns from first row
                        for c in row.keys():
                            cs = c.strip()
                            if sym_col is None and cs in sym_col_candidates:
                                sym_col = c
                            if ind_col is None and cs in ind_col_candidates:
                                ind_col = c
                    csv_rows_dict.append(row)

                # Collect symbols + industry from DictReader rows
                SKIP_SYMS = {"SYMBOL","SYMBOLS","TICKER","COMPANY","COMPANY NAME","ISIN",
                             "ISIN CODE","NAME","SERIES","SECURITY","INDEX NAME","SCRIP","N/A","NA","-",""}
                symbols: set[str] = set()
                for row in csv_rows_dict:
                    sym_val = ""
                    if sym_col:
                        sym_val = str(row.get(sym_col, "") or "").strip().upper()
                    else:
                        # Fallback: try "Symbol" key case-insensitively
                        for k, v in row.items():
                            if k.strip().upper() in {"SYMBOL","SYMBOLS","TICKER"}:
                                sym_val = str(v or "").strip().upper()
                                break
                    if not sym_val or sym_val in SKIP_SYMS or sym_val.isdigit() or len(sym_val) > 20:
                        continue
                    symbols.add(sym_val)
                    # Capture industry
                    if ind_col:
                        ind_val = str(row.get(ind_col, "") or "").strip()
                        if ind_val and sym_val not in _INDUSTRY_MAP:
                            _INDUSTRY_MAP[sym_val] = ind_val

                if not symbols:
                    # Fallback: use raw rows for symbol detection
                    reader2 = csv.reader(io.StringIO(raw))
                    rows_csv = [tuple(r) for r in reader2]
                    _, symbols = _parse_index_symbols(rows_csv)
                if not symbols:
                    continue

                # ── Derive clean index name from filename ──────────────────
                base = os.path.splitext(fname)[0]
                # Strip leading "ind_" or "Ind_" (case-insensitive)
                base = re.sub(r'^[Ii][Nn][Dd]_', '', base)
                # Strip trailing "_list" or "list" (case-insensitive)
                base = re.sub(r'[_]?[Ll][Ii][Ss][Tt]$', '', base)
                # Replace underscores with spaces; uppercase
                idx_name = re.sub(r'_', ' ', base).strip().upper()

                if idx_name not in _INDEX_MAP:
                    _INDEX_MAP[idx_name] = symbols
                    csv_loaded += 1
                else:
                    _INDEX_MAP[idx_name] |= symbols
            except Exception as e:
                print(f"  [!] Error reading '{fname}': {e}")
        if csv_loaded:
            print(f"  ✅ Indices CSVs  : +{csv_loaded} extra indices from '{LOCAL_INDICES_DIR}'")
            n_ind = len({v for v in _INDUSTRY_MAP.values() if v})
            print(f"  ✅ Industries    : {len(_INDUSTRY_MAP)} symbols → {n_ind} unique industries")

    total = len(_INDEX_MAP)
    if total <= 1:
        print(f"  ⚠️  Only built-in NIFTY50 loaded. "
              f"Place xlsx at: {os.path.abspath(LOCAL_INDICES_XLSX)}")
    else:
        print(f"  📊 _INDEX_MAP: {total} indices | _SECTOR_MAP: {len(_SECTOR_MAP)} symbols")
def _parse_nse_csv(text: str, series_filters: list[str], is_sme: bool = False) -> list[str]:
    import re
    
    # Fix malformed CSV with embedded newlines in quoted field names
    # The SME CSV has headers like: "SYMBOL \n","SERIES \n" which breaks CSV parsing
    text = text.lstrip('\ufeff')  # Remove BOM if present
    
    if is_sme:
        # FIX: apply the embedded-newline cleanup only to the header line,
        # not the full body (which would merge real data rows together).
        lines = text.splitlines(keepends=True)
        if lines:
            lines[0] = re.sub(r' \n', ' ', lines[0])
        text = "".join(lines)
    
    # NOW parse the (hopefully fixed) CSV
    _build_company_map(text)
    
    reader, tickers = csv.DictReader(io.StringIO(text)), []
    series_seen = set()   # track unique SERIES values found (for debug)

    for row in reader:
        if not row:  # Skip empty rows
            continue
        # Sanitize row keys by stripping whitespace
        clean_row = {}
        for k, v in row.items():
            if k:
                clean_row[k.strip()] = (v.strip() if isinstance(v, str) else v)
        series = clean_row.get("SERIES", "").strip()
        symbol = clean_row.get("SYMBOL", "").strip()

        if series:
            series_seen.add(series)

        if is_sme:
            # ── SME-dedicated CSV: the ENTIRE file is SME stocks ──────────
            # Accept every row that has a non-empty, non-numeric SYMBOL.
            # We deliberately ignore series_filters here because:
            #   • The file is already filtered to SME at the NSE-export level
            #   • Series codes vary (SM, ST, BE, …) and a mismatch silently
            #     drops all SME stocks, giving n_sme = 0 in the HTML.
            if symbol and not symbol.isdigit():
                tickers.append(symbol)
                _SME_STOCKS.add(symbol)
        else:
            if symbol and series in series_filters:
                tickers.append(symbol)

    if is_sme:
        print(f"  [SME] Series codes found in CSV : {sorted(series_seen) or '(none)'}")
        print(f"  [SME] Symbols loaded into _SME_STOCKS : {len(_SME_STOCKS)}")

    return tickers


def _parse_asx_csv(text: str) -> list[str]:
    """
    Parse nyse-listed.csv (ASX stock list) and return bare ASX codes (no .AX suffix).
    Expected columns: "ACT Symbol", "Company name", "GICS industry group"
    Tickers are appended with ".AX" at download time (e.g. "A2M" → "A2M.AX").
    Populates _COMPANY_MAP (code → company name) and _SECTOR_MAP (code → GICS industry group).
    """
    text = text.lstrip('\ufeff')  # Remove BOM if present
    reader = csv.DictReader(io.StringIO(text))
    tickers = []
    for row in reader:
        if not row:
            continue
        clean = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
        # Primary column name in nyse-listed.csv is "ACT Symbol"
        code    = (clean.get("ACT Symbol") or clean.get("ASX Code") or "").strip().upper()
        name    = clean.get("Company name", "").strip()
        sector  = clean.get("GICS industry group", "").strip()
        if not code:
            continue
        tickers.append(code)
        _ASX_STOCKS.add(code)
        _COMPANY_MAP[code]  = name or code
        if sector:
            _SECTOR_MAP[code] = sector
    return tickers


def load_universe() -> list[str]:
    global _CACHE
    _populate_indices()          # Populate index membership mapping
    _load_fo_list()              # Load F&O eligible symbols
    _CACHE = _load_cache_v2()   # Smart v2 cache (migrates old format automatically)

    all_tickers = []

    # Load NSE EQ stocks
    if os.path.exists(LOCAL_NSE_CSV) and os.path.getsize(LOCAL_NSE_CSV) > 512:
        try:
            with open(LOCAL_NSE_CSV, encoding="utf-8", errors="replace") as f:
                raw = f.read()
            t = _parse_nse_csv(raw, SERIES_FILTER, is_sme=False)
            if t:
                print(f"  ✅ Local '{LOCAL_NSE_CSV}': {len(t)} EQ stocks | "
                      f"{len(_COMPANY_MAP)} companies mapped")
                all_tickers.extend(t)
        except Exception as e:
            print(f"  [!] Local NSE CSV error: {e}")

    # Load NSE SME stocks
    if os.path.exists(LOCAL_SME_CSV):
        try:
            with open(LOCAL_SME_CSV, encoding="utf-8", errors="replace") as f:
                raw = f.read()
            t_sme = _parse_nse_csv(raw, SME_SERIES_FILTER, is_sme=True)
            # Safety net: ensure every ticker returned is in _SME_STOCKS
            for sym in t_sme:
                _SME_STOCKS.add(sym)
            if t_sme:
                print(f"  ✅ Local '{LOCAL_SME_CSV}': {len(t_sme)} SME stocks"
                      f"  |  {len(_SME_STOCKS)} total in SME set")
                all_tickers.extend(t_sme)
            else:
                print(f"  ⚠️  SME CSV found but 0 symbols parsed — check CSV format")
        except Exception as e:
            print(f"  [!] Local SME CSV error: {e}")

    # Load ASX stocks
    if os.path.exists(LOCAL_ASX_CSV) and os.path.getsize(LOCAL_ASX_CSV) > 10:
        try:
            with open(LOCAL_ASX_CSV, encoding="utf-8", errors="replace") as f:
                raw = f.read()
            t_asx = _parse_asx_csv(raw)
            if t_asx:
                print(f"  ✅ Local ASX CSV '{LOCAL_ASX_CSV}': {len(t_asx)} ASX stocks loaded")
                # ASX tickers are stored bare (e.g. "CBA"); yfinance suffix ".AX" added at download time
                all_tickers.extend(t_asx)
            else:
                print(f"  ⚠️  ASX CSV found but 0 symbols parsed — check column names "
                      f"('ACT Symbol', 'Company name', 'GICS industry group')")
        except Exception as e:
            print(f"  [!] Local ASX CSV error: {e}")
    else:
        if LOCAL_ASX_CSV:
            print(f"  ℹ️  ASX CSV not found at '{LOCAL_ASX_CSV}' — skipping ASX stocks")

    # Create ASX chart output directory if needed
    if _ASX_STOCKS:
        os.makedirs(ASX_CHART_OUTPUT_DIR, exist_ok=True)

    if all_tickers:
        return list(dict.fromkeys(all_tickers))  # Remove duplicates while preserving order

    # Fallback to live download if local files don't exist
    try:
        s = requests.Session()
        s.headers.update({"User-Agent": "Mozilla/5.0", "Accept": "*/*"})
        s.get("https://www.nseindia.com/", timeout=12)
        time.sleep(1.5)
        s.headers["Referer"] = "https://www.nseindia.com/"
        r = s.get(NSE_CSV_URL, timeout=20)
        r.raise_for_status()
        t = _parse_nse_csv(r.text, SERIES_FILTER, is_sme=False)
        if t:
            print(f"  ✅ Live NSE: {len(t)} EQ stocks | {len(_COMPANY_MAP)} companies")
            try:
                with open(LOCAL_NSE_CSV, "w", encoding="utf-8") as f:
                    f.write(r.text)
                print(f"  💾 Saved → '{LOCAL_NSE_CSV}'")
            except Exception:
                pass
            all_tickers.extend(t)
    except Exception as e:
        print(f"  [!] NSE download failed: {e}")

    print(f"  ⚠️  Using built-in list: {len(BUILTIN)} stocks")
    return list(dict.fromkeys(BUILTIN))


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 9 — PER-STOCK ANALYSIS  (with full error logging)
# ═════════════════════════════════════════════════════════════════════════════

def analyze_stock(ticker: str) -> dict | None:
    company = get_company_name(ticker)
    min_candles = get_min_candles_required(ticker)

    # ── Stage A: Data (always from cache after prefetch_all) ─────────────
    try:
        df = _get_df(ticker)
        if df is None:
            # Fallback: single download if prefetch missed this ticker
            _yf_suffix = ".AX" if ticker in _ASX_STOCKS else ".NS"
            raw = yf.download(ticker + _yf_suffix, period=DATA_PERIOD, interval="1d",
                              progress=False, auto_adjust=True)
            df  = _clean_df(raw)
            if len(df) >= 1:  # accept any non-empty data
                _set_df(ticker, df)
                _maybe_save_cache()
        else:
            df = df.dropna(subset=["Close"])

        if len(df) < min_candles:
            log_warn(ticker, company,
                     f"Insufficient data: {len(df)} bars < {min_candles} required")
            return None
    except Exception as exc:
        log_error(ticker, company, "DOWNLOAD", exc)
        return None

    # ── Stage B: Resample ─────────────────────────────────────────
    try:
        wk = resample_ohlcv(df, "W-FRI")
        mo = resample_ohlcv(df, "ME")
        if len(wk) < 1 or len(mo) < 1:
            log_warn(ticker, company,
                     f"No resampled bars: W={len(wk)} M={len(mo)}")
            return None
    except Exception as exc:
        log_error(ticker, company, "RESAMPLE", exc)
        return None

    # ── Stage C: Indicators ───────────────────────────────────────
    try:
        rsi_d  = calc_rsi(df["Close"], RSI_P)
        sma_d  = rsi_d.rolling(RSI_SMA_P).mean()
        ml_d, ms_d, mh_d = calc_macd(df["Close"], MACD_F, MACD_S, MACD_SIG_P)
        cci_d  = calc_cci(df["High"], df["Low"], df["Close"], CCI_P)
        atr_d  = calc_atr(df["High"], df["Low"], df["Close"], ATR_P)

        rsi_w  = calc_rsi(wk["Close"], RSI_P)
        sma_w  = rsi_w.rolling(RSI_SMA_P).mean()
        ml_w, ms_w, mh_w = calc_macd(wk["Close"], MACD_F, MACD_S, MACD_SIG_P)
        cci_w  = calc_cci(wk["High"], wk["Low"], wk["Close"], CCI_P)

        rsi_m  = calc_rsi(mo["Close"], RSI_P)
        sma_m  = rsi_m.rolling(RSI_SMA_P).mean()
        ml_m, ms_m, mh_m = calc_macd(mo["Close"], MACD_F, MACD_S, MACD_SIG_P)
        cci_m  = calc_cci(mo["High"], mo["Low"], mo["Close"], CCI_P)

        # ── Explosive breakout indicators (daily) ─────────────────────────
        bb_upper_s, bb_mid_s, bb_lower_s = calc_bb(df["Close"], 20, 2.0)
        mfi_d   = calc_mfi(df["High"], df["Low"], df["Close"], df["Volume"], 14)
        cci_200 = calc_cci(df["High"], df["Low"], df["Close"], 200)
    except Exception as exc:
        log_error(ticker, company, "INDICATORS", exc)
        return None

    # ── Stage D: Feature extraction ───────────────────────────────
    try:
        def f(s, i=-1):
            v = s.iloc[i]
            return float(v) if not (isinstance(v, float) and np.isnan(v)) else 0.0

        v_rsi_d = f(rsi_d); v_sma_d = f(sma_d)
        v_rsi_w = f(rsi_w); v_sma_w = f(sma_w)
        v_rsi_m = f(rsi_m); v_sma_m = f(sma_m)
        v_ml_d  = f(ml_d);  v_ms_d  = f(ms_d)
        v_ml_w  = f(ml_w);  v_ms_w  = f(ms_w)
        v_ml_m  = f(ml_m);  v_ms_m  = f(ms_m)
        v_cci   = f(cci_d); v_cci_w = f(cci_w); v_cci_m = f(cci_m)
        v_atr   = f(atr_d)
        v_close = f(df["Close"])
        v_h52   = float(df["Close"].rolling(252).max().iloc[-1])
        v_l52   = float(df["Close"].rolling(252).min().iloc[-1])
        v_d52   = round((v_close / v_h52 - 1) * 100, 1)

        # ── Explosive breakout feature extraction ─────────────────────────
        v_bb_upper = f(bb_upper_s)
        v_bb_mid   = f(bb_mid_s)
        v_bb_lower = f(bb_lower_s)
        v_mfi      = round(f(mfi_d), 1)
        v_cci_200  = round(f(cci_200), 1)
        # BB % position: 0=at lower, 100=at upper, >100=above upper
        v_bb_pct = round(
            (v_close - v_bb_lower) / (v_bb_upper - v_bb_lower) * 100, 1
        ) if v_bb_upper > v_bb_lower else 50.0
        # BB upper slope over last 5 days (% change): steep = explosive
        _bb5 = float(bb_upper_s.iloc[-5]) if len(bb_upper_s) >= 5 else v_bb_upper
        v_bb_slope = round((v_bb_upper / _bb5 - 1) * 100, 2) if _bb5 > 0 else 0.0
        # MACD histogram today vs yesterday
        v_mh_d      = f(mh_d)
        v_mh_d_prev = float(mh_d.iloc[-2]) if len(mh_d) >= 2 else 0.0
        # Volume ratio: today / 20-day avg
        _vol_today = float(df["Volume"].iloc[-1]) if "Volume" in df.columns else 0.0
        _vol_avg20 = float(df["Volume"].rolling(20).mean().iloc[-1]) if "Volume" in df.columns else 1.0
        v_vol_ratio = round(_vol_today / (_vol_avg20 + 1), 1)

        # ── All-Time High (ATH) — uses High column for true price extreme ────
        ath_price    = float(df["High"].max())
        ath_idx      = df["High"].idxmax()
        ath_date_str = ath_idx.strftime("%d %b %Y")
        last_date    = df.index[-1].date()
        first_date   = df.index[0].date()
        ath_date_dt  = ath_idx.date()
        # Within 1% of ATH high → treat as "at ATH"
        is_ath  = v_close >= ath_price * 0.99
        ath_pct = round((v_close / ath_price - 1) * 100, 1)   # <= 0

        def _time_str(days: int) -> str:
            """Convert a day-count into a compact 'Xy Xm' string."""
            days = max(0, days)
            years, rem = divmod(days, 365)
            months = rem // 30
            if years and months:
                return f"{years}y {months}m"
            if years:
                return f"{years}y"
            if months:
                return f"{months}m"
            return f"{days}d"

        if is_ath:
            days_to_ath  = (ath_date_dt - first_date).days
            ath_time_str = _time_str(days_to_ath) + " to reach"
        else:
            days_since   = (last_date - ath_date_dt).days
            ath_time_str = _time_str(days_since) + " ago"

        # Donchian breakout metrics (20-period prior high breakout on D/W/M)
        def prev_period_max(series, window):
            val = series.shift(1).rolling(window).max().iloc[-1]
            return float(val) if not np.isnan(val) else None
        def prev_period_min(series, window):
            val = series.shift(1).rolling(window).min().iloc[-1]
            return float(val) if not np.isnan(val) else None

        high20_d = prev_period_max(df["High"], 20)
        low20_d  = prev_period_min(df["Low"], 20)
        v_donch_d = round((v_close / high20_d - 1) * 100, 1) if high20_d else None

        v_close_w = f(wk["Close"])
        high20_w  = prev_period_max(wk["High"], 20)
        low20_w   = prev_period_min(wk["Low"], 20)
        v_donch_w = round((v_close_w / high20_w - 1) * 100, 1) if high20_w else None

        v_close_m = f(mo["Close"])
        high20_m  = prev_period_max(mo["High"], 20)
        low20_m   = prev_period_min(mo["Low"], 20)
        v_donch_m = round((v_close_m / high20_m - 1) * 100, 1) if high20_m else None

        def is_fresh(rsi_s, sma_s, window):
            for lag in range(1, window + 2):
                if len(rsi_s) <= lag: break
                if rsi_s.iloc[-lag] > sma_s.iloc[-lag] and rsi_s.iloc[-lag-1] <= sma_s.iloc[-lag-1]:
                    return True, lag
            return False, 0

        fresh_d, fresh_d_bars = is_fresh(rsi_d, sma_d, FRESH_DAYS_D)
        fresh_w, fresh_w_bars = is_fresh(rsi_w, sma_w, FRESH_WEEKS_W)

        score, sig_list = compute_score(
            v_rsi_d, v_sma_d, v_rsi_w, v_sma_w, v_rsi_m, v_sma_m,
            v_ml_d, v_ms_d, v_cci, fresh_d, fresh_w)
        phase  = detect_phase(v_rsi_d, v_rsi_w, v_rsi_m, v_ml_d, v_ms_d, score)

        # ── Explosive breakout score ───────────────────────────────────────
        explosive_score, explosive_signals = compute_explosive_score(
            v_rsi_d, v_vol_ratio, v_close, v_bb_upper, v_bb_slope,
            v_mh_d, v_mh_d_prev, v_mfi, v_cci_200, v_donch_d)
        signal, sig_cls = signal_label(score, phase, fresh_d, fresh_w,
                                        v_rsi_d, v_rsi_w, v_rsi_m,
                                        v_sma_d, v_sma_w, v_sma_m)

        sw_low, _, sw_high, _ = find_swing_points(df["High"], df["Low"])
        atr_sl   = round(v_close - 2.0 * v_atr, 2)
        swing_sl = round(sw_low * 0.99, 2)

        if phase == "UPTREND":
            fib_levels = {k: v for k, v in fib_extensions(sw_low, sw_high).items() if v > v_close}
            fib_type   = "EXTENSION"
            fib_base   = f"Swing Low ₹{sw_low:,.0f} → Swing High ₹{sw_high:,.0f}"
        else:
            fib_levels = {k: v for k, v in fib_retracements(sw_high, sw_low).items()
                          if sw_low < v < sw_high}
            fib_type   = "RETRACEMENT"
            fib_base   = f"Swing High ₹{sw_high:,.0f} → Swing Low ₹{sw_low:,.0f}"

        hist_sigs = historical_signals(df["Close"], rsi_d, sma_d)

        if v_rsi_d > 65:
            entry_note = f"Wait for pullback to RSI~55 zone (~₹{v_close * 0.96:,.0f})"
        elif v_rsi_d > 55 and v_rsi_d > v_sma_d:
            entry_note = f"Entry at current close ₹{v_close:,.0f} or next dip"
        elif fresh_d:
            entry_note = f"Fresh cross — confirm next candle above ₹{v_close:,.0f}"
        else:
            entry_note = "Wait for RSI(14) to cross above SMA(14) on daily"

        sell_conds = ["RSI(14) daily crosses BELOW SMA(14)",
                      "CCI(20) drops below −100",
                      "MACD(12,26) crosses below signal line"]
        if v_rsi_d > 75:
            sell_conds.insert(0, "⚠️ RSI >75 — consider partial profit booking")

        r_sl_pct = round((atr_sl   / v_close - 1) * 100, 1)
        s_sl_pct = round((swing_sl / v_close - 1) * 100, 1)

    except Exception as exc:
        log_error(ticker, company, "FEATURES", exc)
        return None

    return {
        "ticker": ticker, "company": company,
        "close":  v_close, "high52": v_h52, "low52": v_l52, "dist52": v_d52,
        "rsi_d":  round(v_rsi_d,1), "sma_d": round(v_sma_d,1),
        "rsi_w":  round(v_rsi_w,1), "sma_w": round(v_sma_w,1),
        "rsi_m":  round(v_rsi_m,1), "sma_m": round(v_sma_m,1),
        "macd_l": round(v_ml_d,3),  "macd_s": round(v_ms_d,3),
        "macd_l_w": round(v_ml_w,3),"macd_s_w": round(v_ms_w,3),
        "macd_l_m": round(v_ml_m,3),"macd_s_m": round(v_ms_m,3),
        "cci":    round(v_cci,1),   "cci_w": round(v_cci_w,1), "cci_m": round(v_cci_m,1),
        "atr":    round(v_atr,2),
        "donchian_d": v_donch_d, "donchian_w": v_donch_w, "donchian_m": v_donch_m,
        "fresh_d": fresh_d, "fresh_d_bars": fresh_d_bars,
        "fresh_w": fresh_w, "fresh_w_bars": fresh_w_bars,
        "score":   score,   "sig_list": sig_list,
        "phase":   phase,   "signal": signal, "sig_cls": sig_cls,
        "entry_note": entry_note, "sell_conds": sell_conds,
        "atr_sl":  atr_sl,  "swing_sl": swing_sl,
        "r_sl_pct": r_sl_pct, "s_sl_pct": s_sl_pct,
        "sw_low": sw_low, "sw_high": sw_high,
        "fib_type": fib_type, "fib_levels": fib_levels, "fib_base": fib_base,
        "hist_sigs": hist_sigs,
        "ath_price": round(ath_price, 2), "ath_date": ath_date_str,
        "ath_pct": ath_pct, "is_ath": is_ath, "ath_time_str": ath_time_str,
        # ── Explosive breakout fields ──────────────────────────────────────
        "explosive_score":   explosive_score,
        "explosive_signals": explosive_signals,
        "vol_ratio":  v_vol_ratio,
        "bb_upper":   round(v_bb_upper, 2),
        "bb_mid":     round(v_bb_mid, 2),
        "bb_lower":   round(v_bb_lower, 2),
        "bb_pct":     v_bb_pct,
        "bb_slope":   v_bb_slope,
        "mfi":        v_mfi,
        "cci_200":    v_cci_200,
        "macd_hist":  round(v_mh_d, 4),
        # raw series — used for chart only, stripped before HTML table
        "_df": df, "_rsi_d": rsi_d, "_sma_d": sma_d,
        "_rsi_w_daily": rsi_w.reindex(df.index, method="ffill"),
        "_rsi_m_daily": rsi_m.reindex(df.index, method="ffill"),
        "_sma_w_daily": sma_w.reindex(df.index, method="ffill"),
        "_macd_l": ml_d, "_macd_s": ms_d, "_macd_h": mh_d,
        "_cci": cci_d,
        # ranking (filled later by compute_rankings)
        "rank_nifty50": 0, "rank_universe": 0,
        "rank_nifty50_pos": 0, "rank_nifty50_of": 0,
        "rank_univ_pos": 0,    "rank_univ_of": 0,
        "is_nifty50": False,
        "is_sme": is_sme_stock(ticker),
        "is_fo":  is_fo_stock(ticker),
        "sectors": get_sector(ticker),             # list: all Sectoral/Thematic/Strategy memberships
        "sector": get_sector(ticker)[0] if get_sector(ticker) else "Unknown",  # primary (first) for display
        "indices": get_indices(ticker),
        "industry": get_industry(ticker),          # from index CSV "Industry" column
        "marketcap": get_marketcap(ticker),
    }


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 10 — CHART GENERATOR
# ═════════════════════════════════════════════════════════════════════════════

def generate_chart(data: dict) -> str:
    """Generate a chart PNG file and return its relative path, or '' on error."""
    import matplotlib
    matplotlib.use("Agg")   # non-interactive backend — no GUI thread overhead
    ticker  = data["ticker"]
    company = data["company"]
    try:
        df_all = data["_df"]
        n_bars = min(CHART_BARS, len(df_all))
        df     = df_all.iloc[-n_bars:].copy()
        idx    = np.arange(len(df))

        rsi_d  = data["_rsi_d"].iloc[-n_bars:].values
        sma_d  = data["_sma_d"].iloc[-n_bars:].values
        rsi_w  = data["_rsi_w_daily"].iloc[-n_bars:].values
        rsi_m  = data["_rsi_m_daily"].iloc[-n_bars:].values
        macd_l = data["_macd_l"].iloc[-n_bars:].values
        macd_s = data["_macd_s"].iloc[-n_bars:].values
        macd_h = data["_macd_h"].iloc[-n_bars:].values
        cci    = data["_cci"].iloc[-n_bars:].values

        BG="#0d1117"; PANEL="#161b22"; GREEN="#26d07c"; RED="#ff4d6d"
        GOLD="#ffd700"; CYAN="#00d4ff"; PURPLE="#b39ddb"; ORANGE="#ff9800"
        GREY="#30363d"; TXT="#c9d1d9"; FIB_EXT="#4caf50"; FIB_RET="#ff7043"

        fig = plt.figure(figsize=(14, 8), facecolor=BG)
        fig.suptitle(
            f"{ticker} — {data['company']}  |  ₹{data['close']:,.2f}  "
            f"|  {data['phase']}  |  {data['signal']}  |  Score {data['score']}/21  "
            f"|  Univ rank #{data['rank_univ_pos']}/{data['rank_univ_of']}",
            color=TXT, fontsize=14, fontweight="bold", y=0.998
        )
        gs   = gridspec.GridSpec(5, 1, figure=fig, hspace=0.04,
                                 height_ratios=[4, 1.2, 1.8, 1.4, 1.4])
        axes = [fig.add_subplot(gs[i]) for i in range(5)]
        for ax in axes:
            ax.set_facecolor(PANEL)
            ax.tick_params(colors=TXT, labelsize=9, width=1.2, length=5)
            ax.spines[:].set_color(GREY)
            for spine in ax.spines.values():
                spine.set_linewidth(1.2)
            ax.grid(True, color=GREY, linewidth=0.5, linestyle="--", alpha=0.6)
            ax.set_xlim(-1, len(idx))

        step = max(1, len(idx) // 10)
        tpos = idx[::step]
        tlbl = [df.index[i].strftime("%b'%y") for i in tpos]
        for ax in axes:
            ax.set_xticks(tpos)
            ax.set_xticklabels([] if ax != axes[-1] else tlbl,
                               rotation=30, ha="right", fontsize=8, fontweight="bold")

        # Panel 1: Candlestick (vectorised — no iterrows)
        ax1 = axes[0]
        _o = df["Open"].values.astype(float)
        _h = df["High"].values.astype(float)
        _l = df["Low"].values.astype(float)
        _c = df["Close"].values.astype(float)
        _up = _c >= _o
        _cols = [GREEN if u else RED for u in _up]
        ax1.vlines(idx[_up],  _l[_up],  _h[_up],  colors=GREEN, lw=1.2, zorder=2)
        ax1.vlines(idx[~_up], _l[~_up], _h[~_up], colors=RED,   lw=1.2, zorder=2)
        ax1.bar(idx[_up],   _c[_up] -_o[_up],  bottom=_o[_up],  color=GREEN, width=0.75, linewidth=0, zorder=3)
        ax1.bar(idx[~_up],  _o[~_up]-_c[~_up], bottom=_c[~_up], color=RED,   width=0.75, linewidth=0, zorder=3)
        ax1.axhline(data["close"], color=GOLD, lw=1.5, linestyle="--", alpha=0.8, label="Current")
        fib_col = FIB_EXT if data["fib_type"] == "EXTENSION" else FIB_RET
        for lbl, level in data["fib_levels"].items():
            ax1.axhline(level, color=fib_col, lw=1.2, linestyle=":", alpha=0.85)
            ax1.text(len(idx)-1, level, f" {lbl} ₹{level:,.0f}",
                     color=fib_col, fontsize=7, va="center", fontweight="bold", bbox=dict(boxstyle="round,pad=0.3", facecolor=PANEL, edgecolor=fib_col, linewidth=0.5))
        ax1.axhline(data["atr_sl"],   color=RED, lw=1.2, linestyle="-.", alpha=0.7, label="ATR SL")
        ax1.axhline(data["swing_sl"], color=RED, lw=1.0, linestyle="-.", alpha=0.5, label="Swing SL")
        sig_dates = {s["date"]: s["type"] for s in data["hist_sigs"][-8:]}
        for i, dt in enumerate(df.index):
            lbl = sig_dates.get(dt.strftime("%d-%b-%y"))
            if lbl == "BUY":
                ax1.plot(i, float(df["Low"].iloc[i]) * 0.993, "^", color=GREEN, markersize=10, zorder=5, markeredgecolor="white", markeredgewidth=1.5)
            elif lbl == "SELL":
                ax1.plot(i, float(df["High"].iloc[i]) * 1.007, "v", color=RED, markersize=10, zorder=5, markeredgecolor="white", markeredgewidth=1.5)
        ax1.set_ylabel("Price ₹", color=TXT, fontsize=10, fontweight="bold")
        ax1.legend(handles=[mpatches.Patch(color=fib_col, label=f"Fib {data['fib_type']}")],
                   loc="upper left", facecolor=BG, edgecolor=GREY, labelcolor=TXT, fontsize=8, framealpha=0.95)

        # Panel 2: Volume
        ax2 = axes[1]
        vol_avg = pd.Series(df["Volume"].values).rolling(20).mean().values
        _vol = df["Volume"].values.astype(float)
        ax2.bar(idx[_up],  _vol[_up],  color=GREEN, width=0.75, alpha=0.75, linewidth=0)
        ax2.bar(idx[~_up], _vol[~_up], color=RED,   width=0.75, alpha=0.75, linewidth=0)
        ax2.plot(idx, vol_avg, color=GOLD, lw=2.0, label="Vol MA(20)", zorder=5)
        ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x/1e6:.1f}M"))
        ax2.set_ylabel("Volume", color=TXT, fontsize=10, fontweight="bold")
        ax2.legend(loc="upper right", facecolor=BG, edgecolor=GREY, labelcolor=TXT, fontsize=8, framealpha=0.95)

        # Panel 3: RSI D/W/M
        ax3 = axes[2]
        ax3.fill_between(idx, 30, 70, alpha=0.15, color=CYAN, label="Overbought/Oversold")
        ax3.axhline(70, color=RED,   lw=1.5, linestyle="--", alpha=0.7)
        ax3.axhline(55, color=GREEN, lw=1.2, linestyle=":",  alpha=0.6)
        ax3.axhline(50, color=TXT,   lw=1.0, linestyle="--", alpha=0.4)
        ax3.axhline(30, color=GREEN, lw=1.5, linestyle="--", alpha=0.7)
        ax3.plot(idx, rsi_d, color=CYAN,   lw=2.0, label=f"RSI({RSI_P})-D {data['rsi_d']}", zorder=4)
        ax3.plot(idx, sma_d, color=ORANGE, lw=1.8, linestyle="--", label=f"SMA({RSI_SMA_P}) {data['sma_d']}", zorder=4)
        ax3.plot(idx, rsi_w, color=PURPLE, lw=1.5, linestyle="-.", label=f"RSI({RSI_P})-W {data['rsi_w']}", alpha=0.8, zorder=3)
        ax3.plot(idx, rsi_m, color=GOLD,   lw=1.5, linestyle=":",  label=f"RSI({RSI_P})-M {data['rsi_m']}", alpha=0.8, zorder=3)
        if data["fresh_d"] and data["fresh_d_bars"] <= n_bars:
            cx = len(idx) - data["fresh_d_bars"]
            ax3.axvline(cx, color=GREEN, lw=2.0, linestyle="--", alpha=0.8, label="Fresh Cross")
            ax3.text(cx, 74, "FRESH", color=GREEN, fontsize=9, ha="center", fontweight="bold", bbox=dict(boxstyle="round,pad=0.4", facecolor=PANEL, edgecolor=GREEN, linewidth=1.5))
        ax3.set_ylim(10, 90); ax3.set_ylabel("RSI", color=TXT, fontsize=10, fontweight="bold")
        ax3.legend(loc="upper left", facecolor=BG, edgecolor=GREY, labelcolor=TXT, fontsize=8, ncol=2, framealpha=0.95)

        # Panel 4: MACD
        ax4 = axes[3]
        ax4.axhline(0, color=GREY, lw=1.5, alpha=0.7)
        _mpos = np.array(macd_h) >= 0
        ax4.bar(idx[_mpos],  np.array(macd_h)[_mpos],  color=GREEN, width=0.75, alpha=0.75, linewidth=0)
        ax4.bar(idx[~_mpos], np.array(macd_h)[~_mpos], color=RED,   width=0.75, alpha=0.75, linewidth=0)
        ax4.plot(idx, macd_l, color=CYAN,   lw=2.0, label=f"MACD {data['macd_l']:.3f}", zorder=4)
        ax4.plot(idx, macd_s, color=ORANGE, lw=1.8, linestyle="--",
                 label=f"Signal {data['macd_s']:.3f}", zorder=4)
        ax4.set_ylabel("MACD(12,26)", color=TXT, fontsize=10, fontweight="bold")
        ax4.legend(loc="upper left", facecolor=BG, edgecolor=GREY, labelcolor=TXT, fontsize=8, framealpha=0.95)

        # Panel 5: CCI
        ax5 = axes[4]
        ax5.axhline(100,  color=RED,   lw=1.5, linestyle="--", alpha=0.7)
        ax5.axhline(0,    color=GREY,  lw=1.2, alpha=0.6)
        ax5.axhline(-100, color=GREEN, lw=1.5, linestyle="--", alpha=0.7)
        _cpos = cci >= 0
        ax5.bar(idx[_cpos],  cci[_cpos],  color=GREEN, width=0.75, alpha=0.75, linewidth=0)
        ax5.bar(idx[~_cpos], cci[~_cpos], color=RED,   width=0.75, alpha=0.75, linewidth=0)
        ax5.plot(idx, cci, color=CYAN, lw=2.0, label=f"CCI(20) {data['cci']:.1f}", zorder=4)
        ax5.set_ylabel("CCI(20)", color=TXT, fontsize=10, fontweight="bold")
        ax5.legend(loc="upper left", facecolor=BG, edgecolor=GREY, labelcolor=TXT, fontsize=8, framealpha=0.95)

        plt.tight_layout(rect=[0, 0, 1, 0.996])
        updated_at = datetime.now().strftime("%d %b %Y %H:%M")
        fig.text(0.995, 0.005, f"Updated: {updated_at}", ha="right", va="bottom", color=TXT, fontsize=9, style="italic")
        chart_dir = ASX_CHART_OUTPUT_DIR if ticker in _ASX_STOCKS else CHART_OUTPUT_DIR
        os.makedirs(chart_dir, exist_ok=True)
        chart_path = os.path.join(chart_dir, f"{ticker}.png")
        fig.savefig(chart_path, format="png", dpi=CHART_DPI, bbox_inches="tight", facecolor=BG)
        plt.close(fig)
        return chart_path.replace("\\", "/")

    except Exception as exc:
        log_error(ticker, company, "CHART", exc)
        plt.close("all")
        return ""


def _generate_chart_worker(data: dict) -> tuple[str, str]:
    ticker = data["ticker"]
    path = generate_chart(data)
    return ticker, path


def _load_chart_cache_meta() -> dict:
    try:
        with open(CHART_CACHE_META, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_chart_cache_meta(meta: dict) -> None:
    try:
        with open(CHART_CACHE_META, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        log_error("CACHE", "CACHE", "CHARTMETA", exc)


def _compute_chart_hash(data: dict) -> str:
    try:
        n_bars = min(CHART_BARS, len(data["_df"]))
        if n_bars <= 0:
            return ""
        df = data["_df"].iloc[-n_bars:][["Open", "High", "Low", "Close", "Volume"]].astype("float64")
        hasher = hashlib.sha256()
        hasher.update(df.to_numpy().tobytes())
        for arr_key in ["_rsi_d", "_sma_d", "_rsi_w_daily", "_rsi_m_daily", "_macd_l", "_macd_s", "_macd_h", "_cci"]:
            arr = np.asarray(data[arr_key].iloc[-n_bars:])
            hasher.update(arr.astype("float64").tobytes())
        hasher.update(json.dumps(data.get("fib_levels", {}), sort_keys=True, default=str).encode("utf-8"))
        hasher.update(str(data.get("atr_sl", "")).encode("utf-8"))
        hasher.update(str(data.get("swing_sl", "")).encode("utf-8"))
        return hasher.hexdigest()
    except Exception:
        return ""


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 11 — HTML REPORT
# ═════════════════════════════════════════════════════════════════════════════

_CSS = """
:root{
  --bg:#0d1117;--card:#161b22;--border:#30363d;--text:#c9d1d9;
  --sub:#8b949e;--green:#26d07c;--red:#ff4d6d;--gold:#ffd700;
  --cyan:#00d4ff;--purple:#b39ddb;--orange:#ff9800;
  --bs-body-bg:#0d1117;--bs-body-color:#c9d1d9;--bs-border-color:#30363d;
}
*{box-sizing:border-box}
body{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;font-size:13px}
a{color:var(--cyan)}
/* Bootstrap dark overrides */
.form-control,.form-select{background:var(--card)!important;border-color:var(--border)!important;color:var(--text)!important;font-size:12px}
.form-control::placeholder{color:var(--sub)}
.form-control:focus,.form-select:focus{background:var(--card)!important;border-color:var(--cyan)!important;color:var(--text)!important;box-shadow:0 0 0 .2rem rgba(0,212,255,.15)!important}
.form-select option{background:#161b22;color:var(--text)}
.btn-outline-secondary{color:var(--sub);border-color:var(--border);font-size:12px}
.btn-outline-secondary:hover{background:var(--border);color:var(--text);border-color:var(--border)}

/* ── Header ─────────────────────────────────────────── */
.app-header{background:#010409;border-bottom:2px solid #21262d;padding:18px 20px 14px}
.app-header h1{font-size:20px;font-weight:700;color:var(--cyan);letter-spacing:1px;margin:0}
.subtitle{color:var(--sub);font-size:11.5px;margin-top:4px}
.stats-row{display:flex;gap:10px;margin-top:12px;flex-wrap:wrap}
.stat-box{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:9px 16px;min-width:100px;flex:1;min-width:90px;max-width:160px}
.stat-box .val{font-size:22px;font-weight:700}
.stat-box .lbl{font-size:10px;color:var(--sub);margin-top:2px}
.stat-box.green .val{color:var(--green)}.stat-box.gold .val{color:var(--gold)}
.stat-box.red .val{color:var(--red)}.stat-box.cyan .val{color:var(--cyan)}

/* ── Filter section ────────────────────────────────── */
.filter-section{background:#010409;padding:10px 20px;border-bottom:1px solid var(--border);position:sticky;top:0;z-index:1000}
.filter-row1{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:8px}
.filter-row2{display:flex;gap:6px;flex-wrap:wrap;align-items:center;margin-bottom:6px}
.filter-row3{display:flex;gap:6px;flex-wrap:wrap;align-items:center;padding:6px 0;border-top:1px solid #21262d;padding-top:8px}
.filter-input{flex:1;min-width:180px;max-width:260px;background:var(--card);border:1px solid var(--border);color:var(--text);border-radius:20px;padding:5px 14px;font-size:12px;outline:none}
.filter-input:focus{border-color:var(--cyan);box-shadow:0 0 0 2px rgba(0,212,255,.12)}
.filter-input::placeholder{color:var(--sub)}
.filter-select{background:var(--card);border:1px solid var(--border);color:var(--text);border-radius:20px;padding:5px 12px;font-size:12px;cursor:pointer;outline:none;appearance:auto}
.filter-select:focus{border-color:var(--cyan)}
.filter-select option{background:#161b22}
.phase-btn{background:var(--card);border:1px solid var(--border);color:var(--sub);border-radius:20px;padding:4px 13px;cursor:pointer;font-size:12px;transition:all .15s;white-space:nowrap}
.phase-btn:hover,.phase-btn.active{background:var(--cyan);color:#000;border-color:var(--cyan);font-weight:600}
.idx-tab-btn{background:var(--card);border:1px solid var(--border);color:var(--sub);border-radius:20px;padding:5px 14px;cursor:pointer;font-size:12px;transition:all .15s;white-space:nowrap}
.idx-tab-btn:hover{background:#1e2a3a;border-color:var(--cyan);color:var(--cyan)}
.idx-tab-btn.active{background:#0a2440;border-color:var(--cyan);color:var(--cyan);font-weight:700}
.idx-panel tr:hover{background:#161b22}
.rank-sort-label{font-size:11px;color:var(--sub);font-weight:700;white-space:nowrap;letter-spacing:.5px;padding:0 4px}
.rank-sort-btn{background:var(--card);border:2px solid var(--border);color:var(--sub);border-radius:18px;padding:3px 11px;cursor:pointer;font-size:11px;transition:all .15s;white-space:nowrap;font-weight:600}
.rank-sort-btn:hover{background:var(--border);color:var(--text);border-color:var(--gold)}
.rank-sort-btn.active{background:var(--gold);color:#000;border-color:var(--gold);font-weight:700}
.clear-btn{background:transparent;border:1px solid #444;color:var(--sub);border-radius:20px;padding:4px 12px;font-size:12px;cursor:pointer;transition:all .15s}
.clear-btn:hover{border-color:var(--red);color:var(--red)}
.results-info{font-size:11px;color:var(--sub);margin-left:4px;white-space:nowrap}
.results-info b{color:var(--cyan)}
.active-chips{display:flex;gap:5px;flex-wrap:wrap;align-items:center}
.chip{display:inline-flex;align-items:center;gap:4px;background:#002d40;color:var(--cyan);border:1px solid #00d4ff33;border-radius:12px;padding:2px 10px;font-size:10.5px;font-weight:600}
.chip .x{cursor:pointer;opacity:.7;font-size:12px;line-height:1}
.chip .x:hover{opacity:1}

/* ── Summary table ─────────────────────────────────── */
.table-section{padding:0 20px 6px}
.sort-hint{padding:6px 0 4px;font-size:11px;color:var(--sub)}
.table-wrap{display:block;overflow-x:auto;overflow-y:hidden;-webkit-overflow-scrolling:touch;scrollbar-width:thin;scrollbar-color:var(--cyan) var(--card);}
.table-wrap::-webkit-scrollbar{height:10px}
.table-wrap::-webkit-scrollbar-track{background:var(--card)}
.table-wrap::-webkit-scrollbar-thumb{background:var(--cyan);border-radius:10px}
.sum-table{min-width:1380px;border-collapse:collapse;font-size:11.5px}
.sum-table th{background:#21262d;color:var(--sub);padding:7px 9px;text-align:left;
              font-weight:600;white-space:nowrap;position:sticky;top:0;z-index:5}
.sum-table th[data-col]{cursor:pointer;user-select:none}
.sum-table th[data-col]:hover{color:var(--cyan)}
.sort-ind{display:inline-block;min-width:12px;font-size:10px;margin-left:2px;opacity:.7}
.sum-table td{padding:6px 9px;border-bottom:1px solid #21262d;white-space:nowrap}
.sum-table tr:hover td{background:#1c2128}
.rsi-stack{display:flex;flex-direction:column;align-items:center;line-height:1.2}
.rsi-stack .rv{font-weight:600;font-size:12px}
.rsi-stack .sv{font-size:10px;color:var(--sub);margin-top:1px}

/* rank pill */
.rank-pill{display:inline-block;border-radius:10px;padding:1px 8px;font-size:10px;font-weight:700}
.rank-top{background:#0d3320;color:var(--green);border:1px solid #26d07c44}
.rank-mid{background:#2d2600;color:var(--gold); border:1px solid #ffd70044}
.rank-low{background:#2d0a0a;color:var(--red);  border:1px solid #ff4d6d44}

/* badges */
.badge{display:inline-block;border-radius:12px;padding:2px 9px;font-size:10px;font-weight:700;letter-spacing:.4px}
.badge-UPTREND {background:#0d3320;color:var(--green);border:1px solid #26d07c33}
.badge-SIDEWAYS{background:#2d2600;color:var(--gold); border:1px solid #ffd70033}
.badge-BEARISH {background:#2d0a0a;color:var(--red);  border:1px solid #ff4d6d33}
.sig-strong-buy{color:#00e676;font-weight:700}.sig-buy{color:var(--green);font-weight:600}
.sig-watch{color:var(--gold)}.sig-avoid{color:var(--red)}.sig-neutral{color:var(--sub)}
.fresh-tag{background:#002d40;color:var(--cyan);border-radius:8px;padding:1px 7px;
           font-size:10px;font-weight:700;border:1px solid #00d4ff44}
.n50-tag{background:#1a0d30;color:var(--purple);border-radius:8px;padding:1px 7px;
         font-size:10px;font-weight:700;border:1px solid #b39ddb44}
.sme-tag{background:#1a2d0d;color:#4caf50;border-radius:8px;padding:1px 7px;
         font-size:10px;font-weight:700;border:1px solid #4caf5044}
.fo-tag{background:#1a0d2e;color:#ce93d8;border-radius:8px;padding:1px 7px;font-size:10px;font-weight:700;margin-left:3px}
.index-tag{background:#0d2440;color:#03a9f4;border-radius:8px;padding:1px 7px;
           font-size:10px;font-weight:700;border:1px solid #03a9f444}
.sector-tag{background:#2d1a0d;color:#ff9800;border-radius:8px;padding:1px 7px;
            font-size:10px;font-weight:700;border:1px solid #ff980044}
.cap-large{background:#0d1a2d;color:#4caf50;border-radius:8px;padding:1px 7px;
           font-size:10px;font-weight:700;border:1px solid #4caf5044}
.cap-mid{background:#1a2d0d;color:#8bc34a;border-radius:8px;padding:1px 7px;
         font-size:10px;font-weight:700;border:1px solid #8bc34a44}
.cap-small{background:#2d2d0d;color:#fdd835;border-radius:8px;padding:1px 7px;
           font-size:10px;font-weight:700;border:1px solid #fdd83544}
.cap-micro{background:#2d1a1a;color:#ff6f00;border-radius:8px;padding:1px 7px;
           font-size:10px;font-weight:700;border:1px solid #ff6f0044}
.cap-unknown{background:#1a1a1a;color:#888;border-radius:8px;padding:1px 7px;
             font-size:10px;font-weight:700;border:1px solid #88888844}

/* ── Cards ──────────────────────────────────────────── */
.cards-section{padding:12px 20px 40px}
.cards-section>h2{font-size:13px;color:var(--sub);margin-bottom:10px;letter-spacing:1px}

details.stock-card{background:var(--card);border:1px solid var(--border);
                   border-radius:10px;margin-bottom:20px;overflow:hidden}
details.stock-card[open]{border-color:var(--cyan)}
details.stock-card>summary{
  list-style:none;display:flex;align-items:center;gap:12px;padding:13px 18px;
  background:#0d1117;cursor:pointer;flex-wrap:wrap;user-select:none;position:relative}
details.stock-card>summary::-webkit-details-marker,
details.stock-card>summary::-moz-list-bullet{display:none}
details.stock-card>summary .card-arrow{
  display:inline-flex;align-items:center;justify-content:center;
  width:18px;height:18px;color:var(--sub);font-size:12px;flex-shrink:0;
  transition:transform .2s,color .2s}
details.stock-card[open]>summary .card-arrow{transform:rotate(90deg);color:var(--cyan)}
.card-ticker{font-size:17px;font-weight:700;color:var(--cyan)}
.card-price {font-size:15px;font-weight:600}
.card-score {font-size:12px;background:#21262d;border-radius:7px;padding:2px 11px;
             color:var(--gold);font-weight:700}
.card-body{border-top:1px solid var(--border)}
.chart-wrap img{width:100%;display:block;image-rendering:-webkit-optimize-contrast;image-rendering:crisp-edges;image-rendering:high-quality;max-width:100%;height:auto}
.chart-wrap{overflow:hidden;border-radius:4px;min-height:280px}
.chart-placeholder{display:flex;align-items:center;justify-content:center;
                   height:80px;color:var(--sub);font-size:12px;
                   background:var(--bg);border-bottom:1px solid var(--border)}

/* detail panels inside <details> */
.card-details{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));
              gap:0;padding:0}
details.detail-panel{border-right:1px solid var(--border);border-top:1px solid var(--border)}
details.detail-panel:last-child{border-right:none}
details.detail-panel>summary{
  list-style:none;padding:9px 14px;cursor:pointer;font-size:11px;
  font-weight:700;text-transform:uppercase;letter-spacing:.8px;
  color:var(--sub);user-select:none;display:flex;align-items:center;gap:6px}
details.detail-panel>summary::-webkit-details-marker{display:none}
details.detail-panel>summary::after{content:"⌄";margin-left:auto;font-size:13px}
details.detail-panel[open]>summary{color:var(--cyan);border-bottom:1px solid var(--border)}
details.detail-panel[open]>summary::after{transform:rotate(180deg);display:inline-block}
.detail-content{padding:12px 14px}

/* inner tables */
.mini-table{width:100%;border-collapse:collapse;font-size:11.5px}
.mini-table th{color:var(--sub);text-align:left;padding:3px 5px;font-size:10px;font-weight:600}
.mini-table td{padding:4px 5px;border-bottom:1px solid #21262d}
.mini-table tr:last-child td{border-bottom:none}
.g{color:var(--green)}.r{color:var(--red)}

/* trade + sell */
.trade-row{display:flex;justify-content:space-between;padding:4px 0;
           border-bottom:1px solid #21262d;font-size:12px}
.trade-row:last-child{border-bottom:none}
.tl{color:var(--sub)}.tv{font-weight:600}
.tv.green{color:var(--green)}.tv.red{color:var(--red)}.tv.gold{color:var(--gold)}
.entry-box{background:#0d2218;border:1px solid #26d07c33;border-radius:5px;
           padding:7px 9px;margin-top:7px;font-size:11px;color:var(--green)}
.sell-cond{color:var(--red);font-size:11px;padding:3px 0;border-bottom:1px solid #21262d}
.sell-cond:last-child{border-bottom:none}

/* fib */
.fib-row{display:flex;justify-content:space-between;padding:4px 0;
         border-bottom:1px solid #21262d;font-size:11.5px}
.fib-row:last-child{border-bottom:none}
.fl{color:var(--sub);font-size:10.5px}.fv{font-weight:700}
.ext-val{color:#4caf50}.ret-val{color:#ff7043}

/* hist */
.hist-table{width:100%;border-collapse:collapse;font-size:11px}
.hist-table th{color:var(--sub);text-align:right;padding:3px 5px;font-size:10px;font-weight:600}
.hist-table th:first-child,.hist-table th:nth-child(2),.hist-table th:nth-child(3){text-align:left}
.hist-table td{padding:4px 5px;border-bottom:1px solid #21262d;text-align:right}
.hist-table td:first-child,.hist-table td:nth-child(2),.hist-table td:nth-child(3){text-align:left}
.hist-buy{color:var(--green);font-weight:700}.hist-sell{color:var(--red);font-weight:700}
.ret-pos{color:var(--green)}.ret-neg{color:var(--red)}

/* sig dots */
.sig-item{display:flex;align-items:center;gap:7px;padding:4px 0;
          border-bottom:1px solid #21262d;font-size:11.5px}
.sig-item:last-child{border-bottom:none}
.sig-dot{width:6px;height:6px;border-radius:50%;flex-shrink:0}

/* ── Footer ─────────────────────────────────────────── */
.footer{text-align:center;padding:18px;color:var(--sub);
        font-size:11px;border-top:1px solid var(--border)}
/* sort hint */
.sort-hint{padding:5px 0 3px;font-size:11px;color:var(--sub)}
/* ── Pagination ─────────────────────────────────────────── */
#tbl-pager{display:flex;align-items:center;gap:10px;padding:8px 0 4px;flex-wrap:wrap}
.page-btn{background:var(--card);border:1px solid var(--border);color:var(--text);
          border-radius:8px;padding:4px 14px;cursor:pointer;font-size:12px;transition:all .15s}
.page-btn:hover:not(:disabled){border-color:var(--cyan);color:var(--cyan)}
.page-btn:disabled{opacity:.35;cursor:default}
.page-info{font-size:11px;color:var(--sub)}
/* ── Load More ──────────────────────────────────────────── */
.load-more-btn{display:block;width:100%;max-width:420px;margin:14px auto 24px;
               background:var(--card);border:1px solid var(--border);color:var(--cyan);
               border-radius:10px;padding:10px 0;font-size:13px;font-weight:600;
               cursor:pointer;transition:all .2s;letter-spacing:.4px}
.load-more-btn:hover{background:#002d40;border-color:var(--cyan)}
/* ── All-Time High tags ─────────────────────────────────── */
.ath-tag{display:inline-flex;align-items:center;gap:3px;background:linear-gradient(135deg,#1a3a1a,#0d2d0d);
         color:#00e676;border:1px solid #00e67688;border-radius:8px;
         padding:2px 8px;font-size:10px;font-weight:700;letter-spacing:.3px;white-space:nowrap}
.ath-tag .ath-sub{font-weight:400;color:#69f0ae;font-size:9px;margin-left:2px}
.ath-away-tag{display:inline-flex;align-items:center;gap:3px;
              background:#1a1a0d;color:#ffd700;border:1px solid #ffd70055;
              border-radius:8px;padding:2px 8px;font-size:10px;font-weight:700;
              letter-spacing:.3px;white-space:nowrap}
.ath-away-tag .ath-sub{font-weight:400;color:#bdb76b;font-size:9px;margin-left:2px}
/* ── Column Tooltips ─────────────────────────────────── */
.th-wrap{position:relative;display:inline-flex;align-items:center;gap:4px;cursor:pointer}
.th-wrap .tip-icon{
  display:inline-flex;align-items:center;justify-content:center;
  width:13px;height:13px;border-radius:50%;
  background:#21262d;border:1px solid #444;
  color:#888;font-size:9px;font-weight:700;font-style:normal;
  flex-shrink:0;cursor:help;line-height:1}
.th-wrap:hover .tip-icon{background:#002d40;border-color:var(--cyan);color:var(--cyan)}
.col-tooltip{
  display:none;position:absolute;top:calc(100% + 6px);left:50%;
  transform:translateX(-50%);z-index:9999;
  width:240px;background:#1c2128;border:1px solid #30363d;
  border-radius:8px;padding:9px 12px;
  font-size:11px;line-height:1.55;color:#c9d1d9;
  box-shadow:0 8px 24px rgba(0,0,0,.6);pointer-events:none;
  white-space:normal;text-align:left;font-weight:400}
.col-tooltip b{color:var(--cyan)}
.col-tooltip .tip-action{margin-top:6px;padding-top:6px;border-top:1px solid #30363d;font-size:10.5px}
.col-tooltip .tip-buy{color:var(--green)}.col-tooltip .tip-sell{color:var(--red)}
.col-tooltip .tip-watch{color:var(--gold)}
.th-wrap:hover .col-tooltip{display:block}
/* keep tooltip visible when pointer moves into it */
.th-wrap .col-tooltip:hover{display:block}
/* flip tooltip for last few columns so it doesn't go off-screen */
.th-wrap.tip-left .col-tooltip{left:auto;right:0;transform:none}

/* mobile */
@media(max-width:600px){
  .app-header{padding:14px 14px 12px}
  .filter-section{padding:8px 14px}
  .cards-section{padding:10px 14px 30px}
  .table-section{padding:0 14px 4px}
  .card-ticker{font-size:14px}
  .stat-box{min-width:75px;padding:8px 10px}
  .stat-box .val{font-size:18px}
  .col-tooltip{width:190px;font-size:10px}
}
"""

# ─── All rendering is data-driven from STOCKS JSON (injected below) ───────────
# No stock HTML is pre-rendered in Python; JS renders table rows and cards on demand.
_JS = """
// ═══════════════════════════════════════════════════════════════════
//  VIRTUAL RENDER ENGINE  — operates on STOCKS array, not DOM nodes
//  Table: 200 rows/page  |  Cards: 50 at a time with Load More
//  Filter/Sort: pure array ops, then re-render
// ═══════════════════════════════════════════════════════════════════

// STOCKS, CHART_DIR, PAGE_TBL, PAGE_CARDS are injected by Python above this block

const F = { phase:'all', cap:'all', sector:'all', industry:'all', index:'all', fo:'all', signal:'all', ath:'all', search:'' };
let sortKeys   = [];
let filtered   = [];
let tblPage    = 0;
let cardCount  = 0;

// ── HTML escape ───────────────────────────────────────────────────
function esc(s){
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;')
                      .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

// ── Currency formatter ────────────────────────────────────────────
function fmtINR(n){
  if(n===null||n===undefined||n==='') return '—';
  return '₹'+Number(n).toLocaleString('en-IN',{minimumFractionDigits:2,maximumFractionDigits:2});
}
function fmtINR0(n){
  if(!n&&n!==0) return '—';
  return '₹'+Number(n).toLocaleString('en-IN',{maximumFractionDigits:0});
}

// ── Badge / pill helpers ──────────────────────────────────────────
function phaseBadge(ph){
  return `<span class="badge badge-${ph}">${ph}</span>`;
}
function rankPill(pct,pos,of){
  if(!of) return '—';
  const cls=pct>=70?'rank-top':pct>=40?'rank-mid':'rank-low';
  const indicator = pct>=90?'⭐⭐⭐':pct>=70?'⭐⭐':pct>=40?'⭐':'•';
  return `<span class="rank-pill ${cls}">${indicator} #${pos}/${of} (${Math.round(pct)}%ile)</span>`;
}

// ── ATH tag helper ───────────────────────────────────────────────
function athTag(s){
  if(s.is_ath)
    return `<span class="ath-tag">🏆 ATH<span class="ath-sub">${esc(s.ath_time_str)}</span></span>`;
  if(s.ath_pct!=null&&s.ath_pct<0){
    const cls=s.ath_pct>=-5?'var(--gold)':s.ath_pct>=-15?'var(--gold)':'var(--red)';
    return `<span class="ath-away-tag">${s.ath_pct}% ATH<span class="ath-sub">${esc(s.ath_time_str)}</span></span>`;
  }
  return '';
}

// ── Match stock against current filters ───────────────────────────
function matchStock(s){
  if(F.phase!=='all'){
    if(F.phase==='fresh'   &&!(s.fresh_d||s.fresh_w)) return false;
    if(F.phase==='nifty50' &&!s.is_nifty50)           return false;
    if(F.phase==='sme'     &&!s.is_sme)               return false;
    if(!['fresh','nifty50','sme'].includes(F.phase)&&s.phase!==F.phase) return false;
  }
  if(F.cap   !=='all'&&s.cap_cls !==F.cap)    return false;
  if(F.sector!=='all'&&!(s.sectors||[]).includes(F.sector)) return false;
  if(F.industry!=='all'&&(s.industry||'')!==F.industry) return false;
  if(F.index !=='all'&&!(s.indices||[]).includes(F.index)) return false;
  if(F.fo    !=='all'&&((F.fo==='fo'&&!s.is_fo)||(F.fo==='cash'&&s.is_fo))) return false;
  if(F.signal!=='all'&&s.sig_cls !==F.signal) return false;
  if(F.ath!=='all'){
    const p=s.ath_pct;
    if(F.ath==='at'  &&!s.is_ath)                        return false;
    if(F.ath==='w5'  &&(s.is_ath||p==null||p<-5))        return false;
    if(F.ath==='w10' &&(s.is_ath||p==null||p<-10))       return false;
    if(F.ath==='w20' &&(s.is_ath||p==null||p<-20))       return false;
    if(F.ath==='far' &&(p==null||p>=-20))                 return false;
  }
  if(F.search){
    const q=F.search.toLowerCase();
    if(!s.ticker.toLowerCase().includes(q)&&!s.company.toLowerCase().includes(q)) return false;
  }
  return true;
}

// ── Sort filtered array in-place ──────────────────────────────────
const COL_FIELD={ticker:'ticker',score:'score',rsid:'rsi_d',rsiw:'rsi_w',rsim:'rsi_m',
  cci:'cci',macd:'macd_l',close:'close',dist52:'dist52',
  donchd:'donchian_d',donchw:'donchian_w',donchm:'donchian_m',
  rn50:'rank_nifty50',runiv:'rank_universe'};

function applySort(){
  if(!sortKeys.length) return;
  filtered.sort((a,b)=>{
    for(const {col,dir} of sortKeys){
      const f=COL_FIELD[col]||col;
      const av=a[f], bv=b[f];
      if(av===bv) continue;
      if(typeof av==='string') return dir==='desc'?bv.localeCompare(av):av.localeCompare(bv);
      return dir==='desc'?bv-av:av-bv;
    }
    return 0;
  });
}

// ── Main filter entry-point ───────────────────────────────────────
function applyFilters(){
  try{
  filtered=STOCKS.filter(matchStock);
  applySort();
  tblPage=0;
  cardCount=0;
  renderChips();
  const rc=document.getElementById('rc');
  if(rc) rc.textContent=filtered.length;
  renderTable();
  requestAnimationFrame(()=>{ renderCards(true); });
  }catch(e){console.error('applyFilters error:',e);}
}

// ═══════════════════════════════════════════════════════════════════
//  TABLE RENDERING  (paginated, 200 rows/page)
// ═══════════════════════════════════════════════════════════════════
function rowHTML(s){
  const frTag =(s.fresh_d||s.fresh_w)?'<span class="fresh-tag">FRESH</span>':'';
  const n50Tag=s.is_nifty50?'<span class="n50-tag">N50</span>':'';
  const smeTag=s.is_sme?'<span class="sme-tag">SME</span>':'';
  const capTag=s.cap_cat!=='Unknown'?`<span class="${s.cap_cls} badge">${s.cap_cat}</span>`:'';
  // Sector & index tags for Ticker/Company column in table
  // Sector tags — show ALL sectoral/thematic/strategy memberships (from s.sectors[])
  const _secList=(s.sectors&&s.sectors.length)?s.sectors:((s.sector&&s.sector!=='Unknown')?[s.sector]:[]);
  const tblSecTags=_secList.slice(0,2).map(l=>`<span class="sector-tag">${esc(l)}</span>`).join(' ')
    +(_secList.length>2?` <span class="sector-tag" title="${esc(_secList.slice(2).join(' | '))}">+${_secList.length-2}</span>`:'');
  // Index tags — up to 3 then +N overflow
  const _idxList=(s.indices||[]);
  const tblIdxTags=_idxList.slice(0,3).map(i=>`<span class="index-tag">${esc(i)}</span>`).join(' ')
    +(_idxList.length>3?` <span class="index-tag" title="${esc(_idxList.slice(3).join(', '))}">+${_idxList.length-3}</span>`:'');
  // F&O tag
  const foTag=s.is_fo?'<span class="fo-tag">F&amp;O</span>':'';
  // D-RSI vs SMA
  const rsiCol =s.rsi_d>s.sma_d?'var(--green)':'var(--red)';
  const rsiArr =s.rsi_d>s.sma_d?'▲':'▼';
  // W-RSI vs SMA
  const rsiColW=s.rsi_w>s.sma_w?'var(--green)':'var(--red)';
  const rsiArrW=s.rsi_w>s.sma_w?'▲':'▼';
  // M-RSI vs SMA
  const rsiColM=s.rsi_m>s.sma_m?'var(--green)':'var(--red)';
  const rsiArrM=s.rsi_m>s.sma_m?'▲':'▼';
  // MACD color
  const mCol   =s.macd_l>0?'var(--green)':'var(--red)';
  // 52W% color
  const d52Col =s.dist52<-10?'var(--red)':s.dist52>-5?'var(--green)':'';
  // CCI color: >100 bullish, <-100 bearish
  const cciCol =s.cci>100?'var(--green)':s.cci<-100?'var(--red)':'';
  // Donchian color helper: >=−2% near/above 20d high = strong, <−10% = weak
  function donchCol(v){ return v==null?'':v>=-2?'var(--green)':v<-10?'var(--red)':'var(--gold)'; }
  function donchFmt(v){ return v!=null?`${v}%`:'—'; }
  // Market Cap display
  const mcap=s.marketcap!=null
    ?`<div style="font-size:10px;color:var(--sub)">₹${(s.marketcap/1e7).toLocaleString('en-IN',{maximumFractionDigits:0})} Cr</div>`
    :'<div style="font-size:10px;color:var(--sub)">—</div>';
  const _athTag = athTag(s);
  return `<tr class="sum-row">
  <td><b style="color:var(--cyan)">${esc(s.ticker)}</b> ${frTag}${n50Tag}${smeTag}
      <div style="font-size:10px;color:var(--sub)">${esc(s.company.substring(0,28))}</div>
      <div style="font-size:10px;color:var(--gold);font-weight:600">${fmtINR(s.close)}</div>
      ${foTag?`<span style="margin-left:2px">${foTag}</span>`:''}
      ${tblSecTags?`<div style="margin-top:2px">${tblSecTags}</div>`:''}
      ${tblIdxTags?`<div style="margin-top:2px">${tblIdxTags}</div>`:''}
      ${_athTag?`<div style="margin-top:3px">${_athTag}</div>`:''}</td>
  <td style="text-align:center">${phaseBadge(s.phase)}</td>
  <td style="text-align:center"><span class="${s.sig_cls}">${esc(s.signal)}</span></td>
  <td style="text-align:center"><b>${s.score}</b>/21</td>
  <td style="text-align:center">
    <div class="rsi-stack">
      <span class="rv" style="color:${rsiCol}">${s.rsi_d} ${rsiArr}</span>
      <span class="sv">SMA ${s.sma_d}</span>
    </div></td>
  <td style="text-align:center">
    <div class="rsi-stack">
      <span class="rv" style="color:${rsiColW}">${s.rsi_w} ${rsiArrW}</span>
      <span class="sv">SMA ${s.sma_w}</span>
    </div></td>
  <td style="text-align:center">
    <div class="rsi-stack">
      <span class="rv" style="color:${rsiColM}">${s.rsi_m} ${rsiArrM}</span>
      <span class="sv">SMA ${s.sma_m}</span>
    </div></td>
  <td style="text-align:center;color:${cciCol}">${s.cci}</td>
  <td style="text-align:center;color:${mCol}">${s.macd_l!=null?Number(s.macd_l).toFixed(3):'—'}</td>
  <td style="text-align:center">${fmtINR(s.close)}</td>
  <td style="text-align:center;color:${d52Col}">${s.dist52!=null?`${s.dist52}%`:'—'}</td>
  <td style="text-align:center">${capTag}${mcap}</td>
  <td style="text-align:center;color:${donchCol(s.donchian_d)}">${donchFmt(s.donchian_d)}</td>
  <td style="text-align:center;color:${donchCol(s.donchian_w)}">${donchFmt(s.donchian_w)}</td>
  <td style="text-align:center;color:${donchCol(s.donchian_m)}">${donchFmt(s.donchian_m)}</td>
  <td style="text-align:center">${rankPill(s.rank_nifty50,s.rank_nifty50_pos,s.rank_nifty50_of)}</td>
  <td style="text-align:center">${rankPill(s.rank_universe,s.rank_univ_pos,s.rank_univ_of)}</td>
</tr>`;
}

function renderTable(){
  const tbody=document.getElementById('tbl-body');
  if(!tbody) return;
  const start=tblPage*PAGE_TBL;
  const page=filtered.slice(start,start+PAGE_TBL);
  const rows=page.map(s=>{try{return rowHTML(s);}catch(e){return `<tr><td colspan="17" style="color:var(--red);font-size:11px">⚠ Render error: ${esc(String(e))}</td></tr>`;}});
  tbody.innerHTML=rows.join('');
  const pager=document.getElementById('tbl-pager');
  if(!pager) return;
  const total=filtered.length, pages=Math.ceil(total/PAGE_TBL)||1;
  if(pages<=1){ pager.innerHTML=''; return; }
  pager.innerHTML=
    `<button class="page-btn"${tblPage===0?' disabled':''} onclick="goPage(-1)">◀ Prev</button>`+
    `<span class="page-info">Page ${tblPage+1} / ${pages} &nbsp;·&nbsp; ${total} rows</span>`+
    `<button class="page-btn"${tblPage>=pages-1?' disabled':''} onclick="goPage(1)">Next ▶</button>`;
}

function goPage(delta){
  const pages=Math.ceil(filtered.length/PAGE_TBL)||1;
  tblPage=Math.max(0,Math.min(pages-1,tblPage+delta));
  renderTable();
  document.getElementById('sumtable')?.scrollIntoView({behavior:'smooth',block:'start'});
}

// ── Multi-column sort (click header) ─────────────────────────────
function sortTable(col,e){
  const shift=e&&e.shiftKey;
  if(!shift){
    const ex=sortKeys.find(k=>k.col===col);
    const nd=(ex&&sortKeys[0].col===col&&ex.dir==='desc')?'asc':'desc';
    sortKeys=[{col,dir:nd}];
  } else {
    const idx=sortKeys.findIndex(k=>k.col===col);
    if(idx===-1){ if(sortKeys.length<3) sortKeys.push({col,dir:'desc'}); }
    else if(sortKeys[idx].dir==='desc') sortKeys[idx].dir='asc';
    else sortKeys.splice(idx,1);
  }
  document.querySelectorAll('#sumtable th[data-col]').forEach(th=>{
    const ki=sortKeys.findIndex(k=>k.col===th.dataset.col);
    const si=th.querySelector('.sort-ind');
    if(ki===-1){si.textContent='↕';th.style.color='';}
    else{
      const arrow=sortKeys[ki].dir==='desc'?'▼':'▲';
      si.innerHTML=arrow+(sortKeys.length>1?`<sup style="font-size:8px">${ki+1}</sup>`:'');
      th.style.color='var(--cyan)';
    }
  });
  applySort();
  tblPage=0;
  renderTable();
}

// ═══════════════════════════════════════════════════════════════════
//  CARD RENDERING  (50 at a time, Load More)
// ═══════════════════════════════════════════════════════════════════
function cardSummaryHTML(s,idx){
  const frTags=(s.fresh_d?`<span class="fresh-tag">🚀 Daily (${s.fresh_d_bars}d)</span>`:'')
              +(s.fresh_w?`<span class="fresh-tag">📅 Weekly (${s.fresh_w_bars}w)</span>`:'');
  const n50Tag=s.is_nifty50?'<span class="n50-tag">NIFTY50</span>':'';
  const smeTag=s.is_sme?'<span class="sme-tag">SME</span>':'';
  const foTagC=s.is_fo?'<span class="fo-tag">F&amp;O</span>':'';
  const capTag=s.cap_cat!=='Unknown'?`<span class="${s.cap_cls}">${s.cap_cat}</span>`:'';
  const idxTags=(s.indices||[]).map(i=>`<span class="index-tag">${esc(i)}</span>`).join(' ');
  const _cSecList=(s.sectors&&s.sectors.length)?s.sectors:((s.sector&&s.sector!=='Unknown')?[s.sector]:[]);
  const secTag=_cSecList.map(l=>`<span class="sector-tag">${esc(l)}</span>`).join(' ');
  const _ath=athTag(s);
  return `<details class="stock-card" data-idx="${idx}">
  <summary>
    <span class="card-arrow">▶</span>
    <span class="card-ticker">${esc(s.ticker)}</span>
    <span class="card-price">${fmtINR(s.close)}</span>
    ${phaseBadge(s.phase)}
    <span class="${s.sig_cls}" style="font-weight:700">${esc(s.signal)}</span>
    <span class="card-score">Score ${s.score}/21</span>
    ${frTags}${n50Tag}${smeTag}${foTagC}${capTag}${idxTags}${secTag}${_ath?_ath:''}
    <span style="margin-left:auto;color:var(--sub);font-size:11px;text-align:right">
      D ${s.rsi_d} W ${s.rsi_w} M ${s.rsi_m} RSI
      &nbsp;|&nbsp; N50: ${rankPill(s.rank_nifty50,s.rank_nifty50_pos,s.rank_nifty50_of)}
      &nbsp;|&nbsp; All: ${rankPill(s.rank_universe,s.rank_univ_pos,s.rank_univ_of)}
      &nbsp;|&nbsp; D:${s.donchian_d!==null?`${s.donchian_d}%`:'—'} W:${s.donchian_w!==null?`${s.donchian_w}%`:'—'} M:${s.donchian_m!==null?`${s.donchian_m}%`:'—'}
    </span>
  </summary>
  <div class="card-body" data-rendered="0">
    <div class="chart-placeholder" style="height:36px;font-size:12px">▶ Click to expand details</div>
  </div>
</details>`;
}

function renderCards(reset){
  const container=document.getElementById('cards-container');
  if(!container) return;
  if(reset){ container.innerHTML=''; cardCount=0; }
  const batch=filtered.slice(cardCount,cardCount+PAGE_CARDS);
  if(!batch.length){
    if(!cardCount) container.innerHTML='<div style="padding:20px;color:var(--sub)">No stocks match the current filter.</div>';
    document.getElementById('load-more-btn')?.remove();
    return;
  }
  // Build DOM fragment — only summaries, no detail content yet
  const frag=document.createDocumentFragment();
  batch.forEach((s,i)=>{
    try{
      const wrap=document.createElement('div');
      wrap.innerHTML=cardSummaryHTML(s,cardCount+i);
      frag.appendChild(wrap.firstElementChild);
    }catch(e){
      const err=document.createElement('div');
      err.style.cssText='color:var(--red);font-size:11px;padding:4px 8px';
      err.textContent=`⚠ Card render error for ${s&&s.ticker||'?'}: ${e}`;
      frag.appendChild(err);
    }
  });
  container.appendChild(frag);
  cardCount+=batch.length;
  // Load More button
  let btn=document.getElementById('load-more-btn');
  if(cardCount<filtered.length){
    if(!btn){
      btn=document.createElement('button');
      btn.id='load-more-btn';
      btn.className='load-more-btn';
      btn.onclick=()=>renderCards(false);
      container.after(btn);
    }
    btn.textContent=`⬇ Load ${Math.min(PAGE_CARDS,filtered.length-cardCount)} more  (${filtered.length-cardCount} remaining)`;
  } else {
    btn?.remove();
  }
}

// ═══════════════════════════════════════════════════════════════════
//  CARD DETAIL RENDERING  (lazy — only when expanded)
// ═══════════════════════════════════════════════════════════════════
function renderCardDetail(card){
  const body=card.querySelector('.card-body');
  if(!body||body.dataset.rendered==='1') return;
  body.dataset.rendered='1';
  const s=filtered[parseInt(card.dataset.idx)];
  if(!s){ body.innerHTML='<div style="padding:12px;color:var(--sub)">Data not found.</div>'; return; }
  body.innerHTML=buildCardBody(s);
  // Lazy-load chart image
  const img=body.querySelector('img.lazy-chart');
  if(img&&img.dataset.src) img.src=img.dataset.src;
}

function buildCardBody(s){
  // Chart
  const chartHtml=s.has_chart
    ?`<div class="chart-wrap"><img class="lazy-chart"
        src="data:image/gif;base64,R0lGODlhAQABAIAAAAAAAP///ywAAAAAAQABAAACAUwAOw=="
        data-src="${s.chart_path||CHART_DIR+'/'+esc(s.ticker)+'.png'}"
        alt="${esc(s.ticker)} chart" loading="lazy"
        style="opacity:0;transition:opacity .4s;width:100%;display:block"
        onload="this.style.opacity=1"></div>`
    :'<div class="chart-placeholder">📊 Chart not included</div>';

  // RSI table
  function rsiRow(tf,rv,sv,isFr){
    const cls=rv>sv?'g':'r', arr=rv>sv?'▲':'▼';
    const fr=isFr?'<span class="fresh-tag">FRESH</span>':'';
    return `<tr><td>${tf}</td><td class="${cls}"><b>${rv}</b></td>
      <td style="font-size:10px;color:var(--sub)">${sv}</td>
      <td class="${cls}">${arr}${rv>sv?'ABOVE':'BELOW'} ${fr}</td></tr>`;
  }
  const rsiHtml=`<table class="mini-table">
    <tr><th>TF</th><th>RSI(14)</th><th>SMA(14)</th><th>Status</th></tr>
    ${rsiRow('Daily',s.rsi_d,s.sma_d,s.fresh_d)}
    ${rsiRow('Weekly',s.rsi_w,s.sma_w,s.fresh_w)}
    ${rsiRow('Monthly',s.rsi_m,s.sma_m,false)}</table>`;

  // CCI table
  function cciRow(tf,v){
    const cls=v>0?'g':'r';
    const lbl=v>100?'🚀 STRONG':v>0?'✅ Positive':v<-100?'⚠️ EXTREME':'❌ Negative';
    return `<tr><td>${tf}</td><td class="${cls}"><b>${v}</b></td><td class="${cls}">${lbl}</td></tr>`;
  }
  const cciHtml=`<table class="mini-table">
    <tr><th>TF</th><th>CCI(20)</th><th>Signal</th></tr>
    ${cciRow('Daily',s.cci)}${cciRow('Weekly',s.cci_w)}${cciRow('Monthly',s.cci_m)}</table>`;

  // MACD table
  function macdRow(tf,ml,ms){
    const cls=ml>ms?'g':'r', lbl=ml>ms?'▲ BULLISH':'▼ BEARISH';
    return `<tr><td>${tf}</td><td class="${cls}"><b>${ml.toFixed(3)}</b></td><td class="${cls}">${lbl}</td></tr>`;
  }
  const macdHtml=`<table class="mini-table">
    <tr><th>TF</th><th>MACD(12,26)</th><th>Status</th></tr>
    ${macdRow('Daily',s.macd_l,s.macd_s)}
    ${macdRow('Weekly',s.macd_l_w,s.macd_s_w)}
    ${macdRow('Monthly',s.macd_l_m,s.macd_s_m)}</table>`;

  // Trade panel
  const sellHtml=(s.sell_conds||[]).map(c=>`<div class="sell-cond">⚠ ${esc(c)}</div>`).join('');
  const athRowHtml=s.is_ath
    ?`<div class="trade-row"><span class="tl">All-Time High</span>
        <span class="tv" style="color:#00e676">🏆 ${fmtINR(s.ath_price)} <span style="font-size:10px;color:#69f0ae">${esc(s.ath_time_str)}</span></span></div>`
    :`<div class="trade-row"><span class="tl">All-Time High</span>
        <span class="tv">${fmtINR(s.ath_price)}
          <span style="color:var(--red);font-size:11px">&nbsp;${s.ath_pct}%</span>
          <span style="color:var(--sub);font-size:10px">&nbsp;(${esc(s.ath_time_str)})</span>
        </span></div>`;
  const tradeHtml=`
    <div class="trade-row"><span class="tl">Close</span><span class="tv gold">${fmtINR(s.close)}</span></div>
    ${athRowHtml}
    <div class="trade-row"><span class="tl">ATR(14) SL</span>
      <span class="tv red">${fmtINR(s.atr_sl)} (${s.r_sl_pct>0?'+':''}${s.r_sl_pct}%)</span></div>
    <div class="trade-row"><span class="tl">Swing Low SL</span>
      <span class="tv red">${fmtINR(s.swing_sl)} (${s.s_sl_pct>0?'+':''}${s.s_sl_pct}%)</span></div>
    <div class="trade-row"><span class="tl">52W High</span><span class="tv">${fmtINR(s.high52)}</span></div>
    <div class="trade-row"><span class="tl">52W Low</span><span class="tv">${fmtINR(s.low52)}</span></div>
    <div class="entry-box">💡 ${esc(s.entry_note)}</div>
    <div style="margin-top:9px;font-size:10px;color:var(--sub);font-weight:700;
                text-transform:uppercase;letter-spacing:.8px;margin-bottom:4px">EXIT when:</div>
    ${sellHtml}`;

  // Rankings panel
  const rankHtml=`
    <div class="trade-row"><span class="tl">vs Nifty50</span>
      <span class="tv">${rankPill(s.rank_nifty50,s.rank_nifty50_pos,s.rank_nifty50_of)}</span></div>
    <div class="trade-row"><span class="tl">vs All NSE</span>
      <span class="tv">${rankPill(s.rank_universe,s.rank_univ_pos,s.rank_univ_of)}</span></div>
    <div class="trade-row"><span class="tl">Score</span><span class="tv gold">${s.score}/21</span></div>
    <div class="trade-row"><span class="tl">Phase</span><span class="tv">${phaseBadge(s.phase)}</span></div>
    <div class="trade-row"><span class="tl">Signal</span><span class="tv ${s.sig_cls}">${esc(s.signal)}</span></div>`;

  // Market Cap panel
  let capHtml='';
  if(s.marketcap){
    const cr=s.marketcap/1e7;
    const disp=cr>=1?`₹${cr.toLocaleString('en-IN',{maximumFractionDigits:0})} Cr`:`₹${(s.marketcap/1e6).toFixed(1)}M`;
    capHtml+=`<div class="trade-row"><span class="tl">Market Cap</span><span class="tv gold">${disp}</span></div>`;
    capHtml+=`<div class="trade-row"><span class="tl">Category</span><span class="tv"><span class="${s.cap_cls} badge">${s.cap_cat}</span></span></div>`;
  } else {
    capHtml+='<div class="trade-row"><span class="tl">Market Cap</span><span class="tv" style="color:var(--sub)">Not available</span></div>';
  }
  if(s.sector&&s.sector!=='Unknown')
    capHtml+=`<div class="trade-row"><span class="tl">Sector</span><span class="tv"><span class="sector-tag badge">${esc(s.sector)}</span></span></div>`;
  if((s.indices||[]).length)
    capHtml+=`<div class="trade-row"><span class="tl">Indices</span><span class="tv">${s.indices.map(i=>`<span class="index-tag badge">${esc(i)}</span>`).join(' ')}</span></div>`;

  // Fib panel
  const fibCol=s.fib_type==='EXTENSION'?'ext-val':'ret-val';
  const fibLbl=s.fib_type==='EXTENSION'?'🎯 Fib Extension — Upside Targets':'🛡️ Fib Retracement — Support';
  const fibBody=Object.entries(s.fib_levels||{}).map(([lvl,price])=>{
    const pct=((price/s.close-1)*100).toFixed(1);
    return `<div class="fib-row"><span class="fl">${lvl}</span>
      <span class="fv ${fibCol}">${fmtINR(price)}
        <span style="color:var(--sub);font-size:10px">${pct>=0?'+':''}${pct}%</span>
      </span></div>`;
  }).join('')||'<span style="color:var(--sub)">No levels near price</span>';
  const fibHtml=`<div style="font-size:10.5px;color:var(--sub);margin-bottom:6px">${esc(s.fib_base)}</div>${fibBody}`;

  // Active signals
  const dotMap={'✅':'#26d07c','🚀':'#00d4ff','🔥':'#ff9800','💪':'#b39ddb','💰':'#ffd700'};
  const sigsHtml=(s.sig_list||[]).map(sig=>{
    const hit=Object.entries(dotMap).find(([e])=>sig.includes(e));
    return `<div class="sig-item"><div class="sig-dot" style="background:${hit?hit[1]:'#26d07c'}"></div><span>${esc(sig)}</span></div>`;
  }).join('')||'<span style="color:var(--sub)">No active signals</span>';

  // Historical signals
  function retSpan(v){
    if(v===null||v===undefined) return '<span style="color:#555">—</span>';
    return `<span class="${v>=0?'ret-pos':'ret-neg'}">${v>=0?'+':''}${v}%</span>`;
  }
  const histRows=[...(s.hist_sigs||[])].reverse().map(h=>
    `<tr><td>${esc(h.date)}</td>
     <td class="${h.type==='BUY'?'hist-buy':'hist-sell'}">${h.type}</td>
     <td>${fmtINR(h.price)}</td><td>RSI ${h.rsi}</td>
     <td>${retSpan(h.r5d)}</td><td>${retSpan(h.r10d)}</td><td>${retSpan(h.r20d)}</td></tr>`
  ).join('')||'<tr><td colspan="7" style="color:var(--sub)">No signals in history</td></tr>';
  const histHtml=`<table class="hist-table">
    <tr><th>Date</th><th>Type</th><th>Price</th><th>RSI</th>
        <th>5D</th><th>10D</th><th>20D</th></tr>
    ${histRows}</table>`;

  function dp(title,content,open=false){
    return `<details class="detail-panel"${open?' open':''}><summary>${title}</summary><div class="detail-content">${content}</div></details>`;
  }
  return `${chartHtml}<div class="card-details">
    ${dp('📊 RSI · Daily · Weekly · Monthly',rsiHtml,true)}
    ${dp('🎯 CCI(20) · D · W · M',cciHtml)}
    ${dp('📈 MACD(12,26) · D · W · M',macdHtml)}
    ${dp('💼 Entry / Stop Loss / Exit',tradeHtml,true)}
    ${dp('🏆 Rankings',rankHtml,true)}
    ${dp('💰 Market Cap',capHtml)}
    ${dp(fibLbl,fibHtml)}
    ${dp('⚡ Active Signals',sigsHtml)}
    <details class="detail-panel" style="grid-column:1/-1">
      <summary>📅 Historical RSI Crossover Signals — recent first</summary>
      <div class="detail-content">${histHtml}</div>
    </details>
  </div>`;
}

// ─── Expand card lazily ───────────────────────────────────────────
document.addEventListener('toggle', e=>{
  const card=e.target;
  if(!card.classList?.contains('stock-card')||!card.open) return;
  renderCardDetail(card);
}, true);

// ─── Filter controls ──────────────────────────────────────────────
function filterPhase(phase,btn){
  document.querySelectorAll('.phase-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  F.phase=phase;
  applyFilters();
}

// ─── Quick ranking sort ────────────────────────────────────────────
function sortByRanking(col,dir,evt){
  sortKeys=[{col,dir}];
  applySort();
  tblPage=0;
  cardCount=0;
  renderTable();
  requestAnimationFrame(()=>{ renderCards(true); });
  // Visual feedback
  document.querySelectorAll('.rank-sort-btn').forEach(b=>b.classList.remove('active'));
  if(evt&&evt.target) evt.target.classList.add('active');
}
function onDropChange(){
  F.cap   =document.getElementById('capSel').value;
  F.sector=document.getElementById('secSel').value;
  F.industry=document.getElementById('indSel')?.value||'all';
  F.index =document.getElementById('idxSel').value;
  F.fo    =document.getElementById('foSel')?.value||'all';
  F.signal=document.getElementById('sigSel').value;
  F.ath   =document.getElementById('athSel').value;
  applyFilters();
}
let _st=null;
function onSearch(v){
  clearTimeout(_st);
  _st=setTimeout(()=>{ F.search=v.trim(); applyFilters(); },220);
}
function clearAll(){
  Object.assign(F,{phase:'all',cap:'all',sector:'all',industry:'all',index:'all',fo:'all',signal:'all',ath:'all',search:''});
  ['capSel','secSel','indSel','idxSel','foSel','sigSel','athSel'].forEach(id=>{const el=document.getElementById(id);if(el)el.value='all';});
  const si=document.getElementById('searchInp');if(si)si.value='';
  document.querySelectorAll('.phase-btn').forEach(b=>b.classList.remove('active'));
  document.querySelector('.phase-btn[data-phase="all"]')?.classList.add('active');
  applyFilters();
}

// ─── Filter chips ─────────────────────────────────────────────────
const CAP_LABELS={'cap-large':'Large Cap','cap-mid':'Mid Cap','cap-small':'Small Cap','cap-micro':'Micro Cap'};
const ATH_LABELS={at:'🏆 At ATH',w5:'✅ Within 5% ATH',w10:'🟡 Within 10% ATH',w20:'🟠 Within 20% ATH',far:'📉 >20% below ATH'};
const SIG_LABELS={'sig-strong-buy':'🚀 Strong Buy','sig-buy':'✅ Buy','sig-watch':'👀 Watch','sig-avoid':'❌ Avoid','sig-neutral':'⚪ Neutral'};
function renderChips(){
  const c=document.getElementById('chips');
  if(!c) return;
  const chips=[];
  if(F.phase!=='all')  chips.push([`Phase: ${F.phase}`,()=>{F.phase='all';document.querySelectorAll('.phase-btn').forEach(b=>b.classList.remove('active'));document.querySelector('.phase-btn[data-phase="all"]')?.classList.add('active');}]);
  if(F.cap!=='all')    chips.push([CAP_LABELS[F.cap]||F.cap,()=>{F.cap='all';document.getElementById('capSel').value='all';}]);
  if(F.sector!=='all') chips.push([`🏭 ${F.sector}`,()=>{F.sector='all';document.getElementById('secSel').value='all';}]);
  if(F.industry!=='all') chips.push([`🏢 ${F.industry}`,()=>{F.industry='all';const el=document.getElementById('indSel');if(el)el.value='all';}]);
  if(F.index!=='all')  chips.push([`📊 ${F.index}`,()=>{F.index='all';document.getElementById('idxSel').value='all';}]);
  if(F.fo!=='all')     chips.push([`🔮 ${F.fo==='fo'?'F&O Only':'Cash Only'}`,()=>{F.fo='all';const el=document.getElementById('foSel');if(el)el.value='all';}]);
  if(F.signal!=='all') chips.push([SIG_LABELS[F.signal]||F.signal,()=>{F.signal='all';document.getElementById('sigSel').value='all';}]);
  if(F.ath!=='all')    chips.push([ATH_LABELS[F.ath]||F.ath,()=>{F.ath='all';document.getElementById('athSel').value='all';}]);
  if(F.search)         chips.push([`"${F.search}"`,()=>{F.search='';document.getElementById('searchInp').value='';}]);
  c.innerHTML=chips.map((ch,i)=>
    `<span class="chip">${esc(ch[0])} <span class="x" onclick="(${chips[i][1].toString()})();applyFilters()">✕</span></span>`
  ).join('');
}

// ─── Init ─────────────────────────────────────────────────────────
document.addEventListener('DOMContentLoaded',()=>{
  try{
  filtered = STOCKS.slice();
  // Update count display
  const _rc = document.getElementById('rc');
  if(_rc) _rc.textContent = filtered.length;
  // Mark "All" phase button active
  const _pb = document.querySelector('.phase-btn[data-phase="all"]');
  if(_pb) _pb.classList.add('active');
  // Render table synchronously (fast — 100 rows direct innerHTML)
  renderTable();
  // Render cards in next frame so table paints first
  requestAnimationFrame(()=>{ renderCards(true); });
  }catch(e){
    console.error('DOMContentLoaded render error:',e);
    document.body.insertAdjacentHTML('afterbegin',
      `<div style="background:#3d0000;color:#ff6b6b;padding:12px 18px;font-family:monospace;white-space:pre-wrap;z-index:9999;position:fixed;top:0;left:0;right:0">
⚠ Report render error (open browser DevTools → Console for details):\n${e}</div>`);
  }
});
"""


def _ret_span(val):
    if val is None: return '<span style="color:#555">—</span>'
    cls = "ret-pos" if val >= 0 else "ret-neg"
    return f'<span class="{cls}">{val:+.1f}%</span>'

def _phase_badge(phase):
    return f'<span class="badge badge-{phase}">{phase}</span>'

def _sig_span(signal, cls):
    return f'<span class="{cls}">{signal}</span>'

def _rank_pill(pct, pos, of):
    if of == 0: return "—"
    cls = "rank-top" if pct >= 70 else "rank-mid" if pct >= 40 else "rank-low"
    # Add visual indicator based on ranking tier
    indicator = "⭐⭐⭐" if pct >= 90 else "⭐⭐" if pct >= 70 else "⭐" if pct >= 40 else "•"
    return f'<span class="rank-pill {cls}">{indicator} #{pos}/{of} ({pct:.0f}%ile)</span>'


def _build_detail_panels(d: dict) -> str:
    close = d["close"]

    # RSI table
    def rsi_row(tf, rv, sv, is_fresh_tf):
        cls = "g" if rv > sv else "r"
        arr = "▲" if rv > sv else "▼"
        fr  = ' <span class="fresh-tag">FRESH</span>' if is_fresh_tf else ""
        return (f'<tr><td>{tf}</td>'
                f'<td class="{cls}"><b>{rv}</b></td>'
                f'<td style="font-size:10px;color:var(--sub)">{sv}</td>'
                f'<td class="{cls}">{arr}{"ABOVE" if rv>sv else "BELOW"}{fr}</td></tr>')

    def cci_row(tf, v):
        cls = "g" if v > 0 else "r"
        lbl = ("🚀 STRONG" if v > 100 else "✅ Positive" if v > 0
               else "⚠️ EXTREME" if v < -100 else "❌ Negative")
        return f'<tr><td>{tf}</td><td class="{cls}"><b>{v}</b></td><td class="{cls}">{lbl}</td></tr>'

    def macd_row(tf, ml, ms):
        cls = "g" if ml > ms else "r"
        lbl = "▲ BULLISH" if ml > ms else "▼ BEARISH"
        return f'<tr><td>{tf}</td><td class="{cls}"><b>{ml:.3f}</b></td><td class="{cls}">{lbl}</td></tr>'

    rsi_html = f"""<table class="mini-table">
      <tr><th>TF</th><th>RSI(14)</th><th>SMA(14)</th><th>Status</th></tr>
      {rsi_row("Daily",   d["rsi_d"], d["sma_d"], d["fresh_d"])}
      {rsi_row("Weekly",  d["rsi_w"], d["sma_w"], d["fresh_w"])}
      {rsi_row("Monthly", d["rsi_m"], d["sma_m"], False)}
    </table>"""

    cci_html = f"""<table class="mini-table">
      <tr><th>TF</th><th>CCI(20)</th><th>Signal</th></tr>
      {cci_row("Daily",   d["cci"])}
      {cci_row("Weekly",  d["cci_w"])}
      {cci_row("Monthly", d["cci_m"])}
    </table>"""

    macd_html = f"""<table class="mini-table">
      <tr><th>TF</th><th>MACD(12,26)</th><th>Status</th></tr>
      {macd_row("Daily",   d["macd_l"],   d["macd_s"])}
      {macd_row("Weekly",  d["macd_l_w"], d["macd_s_w"])}
      {macd_row("Monthly", d["macd_l_m"], d["macd_s_m"])}
    </table>"""

    trade_html = f"""
    <div class="trade-row"><span class="tl">Close</span><span class="tv gold">₹{close:,.2f}</span></div>
    <div class="trade-row"><span class="tl">ATR(14) SL</span>
      <span class="tv red">₹{d['atr_sl']:,.2f} ({d['r_sl_pct']:+.1f}%)</span></div>
    <div class="trade-row"><span class="tl">Swing Low SL</span>
      <span class="tv red">₹{d['swing_sl']:,.2f} ({d['s_sl_pct']:+.1f}%)</span></div>
    <div class="trade-row"><span class="tl">52W High</span><span class="tv">₹{d['high52']:,.2f}</span></div>
    <div class="trade-row"><span class="tl">52W Low</span> <span class="tv">₹{d['low52']:,.2f}</span></div>
    <div class="entry-box">💡 {d['entry_note']}</div>
    <div style="margin-top:9px;font-size:10px;color:var(--sub);font-weight:700;
                text-transform:uppercase;letter-spacing:.8px;margin-bottom:4px">EXIT when:</div>
    {"".join(f'<div class="sell-cond">⚠ {c}</div>' for c in d['sell_conds'])}"""

    fib_col = "ext-val" if d["fib_type"] == "EXTENSION" else "ret-val"
    fib_lbl = "🎯 Fib Extension — Upside Targets" if d["fib_type"] == "EXTENSION" else "🛡️ Fib Retracement — Support"
    fib_body = "".join(
        f'<div class="fib-row"><span class="fl">{lvl}</span>'
        f'<span class="fv {fib_col}">₹{price:,.2f} '
        f'<span style="color:var(--sub);font-size:10px">{round((price/close-1)*100,1):+.1f}%</span>'
        f'</span></div>'
        for lvl, price in d["fib_levels"].items()
    ) or '<span style="color:var(--sub)">No levels near price</span>'
    fib_html = f'<div style="font-size:10.5px;color:var(--sub);margin-bottom:6px">{d["fib_base"]}</div>{fib_body}'

    dot_map = {"✅":"#26d07c","🚀":"#00d4ff","🔥":"#ff9800","💪":"#b39ddb","💰":"#ffd700"}
    sigs_html = "".join(
        f'<div class="sig-item"><div class="sig-dot" style="background:'
        f'{next((c for e,c in dot_map.items() if e in s),"#26d07c")}"></div>'
        f'<span>{s}</span></div>'
        for s in d["sig_list"]
    ) or '<span style="color:var(--sub)">No active signals</span>'

    hist_rows = "".join(
        f'<tr><td>{s["date"]}</td>'
        f'<td class="{"hist-buy" if s["type"]=="BUY" else "hist-sell"}">{s["type"]}</td>'
        f'<td>₹{s["price"]:,.2f}</td><td>RSI {s["rsi"]}</td>'
        f'<td>{_ret_span(s["r5d"])}</td><td>{_ret_span(s["r10d"])}</td>'
        f'<td>{_ret_span(s["r20d"])}</td></tr>'
        for s in reversed(d["hist_sigs"])
    ) or '<tr><td colspan="7" style="color:var(--sub)">No signals in history</td></tr>'
    hist_html = f"""<table class="hist-table">
      <tr><th>Date</th><th>Type</th><th>Price</th><th>RSI</th>
          <th>5D</th><th>10D</th><th>20D</th></tr>
      {hist_rows}</table>"""

    # Rankings
    rank_html = f"""
    <div class="trade-row"><span class="tl">vs Nifty50</span>
      <span class="tv">{_rank_pill(d['rank_nifty50'], d['rank_nifty50_pos'], d['rank_nifty50_of'])}</span></div>
    <div class="trade-row"><span class="tl">vs All NSE</span>
      <span class="tv">{_rank_pill(d['rank_universe'], d['rank_univ_pos'], d['rank_univ_of'])}</span></div>
    <div class="trade-row"><span class="tl">Score</span>
      <span class="tv gold">{d['score']}/21</span></div>
    <div class="trade-row"><span class="tl">Phase</span>
      <span class="tv">{_phase_badge(d['phase'])}</span></div>
    <div class="trade-row"><span class="tl">Signal</span>
      <span class="tv {d['sig_cls']}">{d['signal']}</span></div>"""

    def dp(title, content, open_default=False):
        op = " open" if open_default else ""
        return (f'<details class="detail-panel"{op}>'
                f'<summary>{title}</summary>'
                f'<div class="detail-content">{content}</div>'
                f'</details>')

    # Market Cap Panel
    cap_cat, cap_cls = categorize_marketcap(d['marketcap'])
    if d['marketcap']:
        cap_crore = d['marketcap'] / 1e7
        cap_display = f"₹{cap_crore:,.0f} Cr" if cap_crore >= 1 else f"₹{d['marketcap']/1e6:,.1f}M"
        cap_html = f"""<div class="trade-row"><span class="tl">Market Cap</span>
          <span class="tv gold">{cap_display}</span></div>
        <div class="trade-row"><span class="tl">Category</span>
          <span class="tv"><span class="{cap_cls} badge">{cap_cat}</span></span></div>"""
    else:
        cap_html = '<div class="trade-row"><span class="tl">Market Cap</span><span class="tv" style="color:var(--sub)">Not available</span></div>'
    # Add sector & indices to cap panel
    if d.get('sector') and d['sector'] != 'Unknown':
        cap_html += f'<div class="trade-row"><span class="tl">Sector</span><span class="tv"><span class="sector-tag badge">{d["sector"]}</span></span></div>'
    if d.get('indices'):
        idx_html = ' '.join(f'<span class="index-tag badge">{i}</span>' for i in d['indices'])
        cap_html += f'<div class="trade-row"><span class="tl">Indices</span><span class="tv">{idx_html}</span></div>'

    return f"""<div class="card-details">
      {dp("📊 RSI · Daily · Weekly · Monthly", rsi_html, True)}
      {dp("🎯 CCI(20) · D · W · M",            cci_html)}
      {dp("📈 MACD(12,26) · D · W · M",         macd_html)}
      {dp("💼 Entry / Stop Loss / Exit",         trade_html, True)}
      {dp("🏆 Rankings",                         rank_html, True)}
      {dp("💰 Market Cap",                       cap_html)}
      {dp(fib_lbl,                               fib_html)}
      {dp("⚡ Active Signals",                   sigs_html)}
      <details class="detail-panel" style="grid-column:1/-1">
        <summary>📅 Historical RSI Crossover Signals — recent first</summary>
        <div class="detail-content">{hist_html}</div>
      </details>
    </div>"""


def build_summary_table() -> str:
    """Return the table skeleton only — tbody is populated by JS from STOCKS JSON."""

    def th(lbl, col, tooltip_html, align="center", tip_left=False):
        """Sortable header with hover tooltip."""
        wrap_cls = "th-wrap tip-left" if tip_left else "th-wrap"
        return (
            f'<th data-col="{col}" onclick="sortTable(\'{col}\',event)" style="text-align:{align}">'
            f'<span class="{wrap_cls}">'
            f'{lbl} <span class="sort-ind">↕</span>'
            f'<i class="tip-icon">?</i>'
            f'<div class="col-tooltip">{tooltip_html}</div>'
            f'</span></th>'
        )

    def th_plain(lbl, tooltip_html, tip_left=False):
        """Non-sortable header with hover tooltip."""
        wrap_cls = "th-wrap tip-left" if tip_left else "th-wrap"
        return (
            f'<th style="text-align:center">'
            f'<span class="{wrap_cls}">'
            f'{lbl}'
            f'<i class="tip-icon">?</i>'
            f'<div class="col-tooltip">{tooltip_html}</div>'
            f'</span></th>'
        )

    # ── Tooltip content for each column ──────────────────────────────────────
    TIP_TICKER = (
        "<b>Ticker / Company</b><br>NSE stock symbol and company name. "
        "Tags show FRESH breakout, NIFTY50 membership, SME, market-cap category, "
        "sector/industry (orange), and index memberships (blue)."
        "<div class='tip-action'>"
        "<span class='tip-buy'>▲ Sort A→Z</span> to scan alphabetically, or use the search box to jump to any stock.<br>"
        "Use the <b>🏭 Sector</b> and <b>📊 Index</b> filters above to narrow by industry or index membership."
        "</div>"
    )

    TIP_PHASE = (
        "<b>Market Phase</b><br>Overall trend based on price, RSI, and SMA alignment across timeframes."
        "<div class='tip-action'>"
        "<span class='tip-buy'>📈 UPTREND</span> — All systems aligned bullish. Best phase to enter longs.<br>"
        "<span class='tip-watch'>➡ SIDEWAYS</span> — Consolidating. Wait for a breakout before entering.<br>"
        "<span class='tip-sell'>📉 BEARISH</span> — Downtrend. Avoid new buys; protect open positions."
        "</div>"
    )

    TIP_SIGNAL = (
        "<b>Composite Signal</b><br>Derived from RSI crossovers on Daily + Weekly + Monthly timeframes combined with CCI and MACD."
        "<div class='tip-action'>"
        "<span class='tip-buy'>STRONG BUY</span> — All 3 TFs bullish + CCI/MACD confirm. Highest-conviction entry.<br>"
        "<span class='tip-buy'>BUY</span> — Majority TFs bullish. Good entry with tight SL.<br>"
        "<span class='tip-watch'>WATCH</span> — Mixed signals. Monitor for confirmation.<br>"
        "<span class='tip-sell'>HOLD / AVOID</span> — Bearish or weakening. No new entries."
        "</div>"
    )

    TIP_SCORE = (
        "<b>Momentum Score (out of 21)</b><br>"
        "Points awarded across RSI (3 TFs), CCI (3 TFs), MACD (3 TFs), Donchian breakouts (3 TFs), "
        "Fibonacci position, fresh crossover bonus, and phase bonus."
        "<div class='tip-action'>"
        "<span class='tip-buy'>≥ 16 — STRONG BUY.</span> Maximum conviction. Prioritise these.<br>"
        "<span class='tip-buy'>12–15 — BUY.</span> Strong setup. Enter with defined SL.<br>"
        "<span class='tip-watch'>8–11 — WATCH.</span> Potential developing. Wait for score to rise.<br>"
        "<span class='tip-sell'>&lt; 8 — AVOID.</span> Too many bearish signals. Stand aside.<br><br>"
        "💡 Sort by Score descending to rank the strongest setups across the entire universe."
        "</div>"
    )

    TIP_DRSI = (
        "<b>Daily RSI(14) vs SMA(34)</b><br>"
        "RSI(14) on the daily chart compared to its 34-period smoothing average. "
        "The crossover of RSI above its SMA is the core entry trigger."
        "<div class='tip-action'>"
        "<span class='tip-buy'>▲ RSI &gt; SMA</span> — Daily momentum is bullish. Valid entry zone.<br>"
        "<span class='tip-sell'>▼ RSI &lt; SMA</span> — Daily momentum is bearish. Avoid or exit.<br>"
        "<b>FRESH</b> tag = crossover happened within last 3 bars — highest-quality entry signal."
        "</div>"
    )

    TIP_WRSI = (
        "<b>Weekly RSI(14) vs SMA(34)</b><br>"
        "Same RSI/SMA crossover logic on the weekly chart. Weekly alignment gives medium-term trend direction."
        "<div class='tip-action'>"
        "<span class='tip-buy'>▲ Weekly RSI &gt; SMA</span> — Medium-term trend is up. Adds conviction to daily buy signals.<br>"
        "<span class='tip-sell'>▼ Weekly RSI &lt; SMA</span> — Medium-term weak. Daily buy signals are lower quality.<br>"
        "Best entries: Daily <b>AND</b> Weekly both show ▲."
        "</div>"
    )

    TIP_MRSI = (
        "<b>Monthly RSI(14) vs SMA(34)</b><br>"
        "RSI/SMA crossover on the monthly chart. Monthly alignment confirms the macro bull trend."
        "<div class='tip-action'>"
        "<span class='tip-buy'>▲ All 3 TFs (D+W+M) above SMA</span> — Full multi-timeframe alignment. Strongest possible setup.<br>"
        "<span class='tip-sell'>▼ Monthly RSI &lt; SMA</span> — Long-term trend is down. Even strong daily signals are counter-trend trades."
        "</div>"
    )

    TIP_CCI = (
        "<b>Daily CCI(20) — Commodity Channel Index</b><br>"
        "Measures how far price is from its statistical average. Confirms momentum direction."
        "<div class='tip-action'>"
        "<span class='tip-buy'>&gt; 100</span> — 🚀 Strong bullish momentum. Trend is accelerating.<br>"
        "<span class='tip-buy'>0 to 100</span> — ✅ Mild positive bias. Entry is valid but momentum moderate.<br>"
        "<span class='tip-sell'>0 to −100</span> — ❌ Negative bias. Caution on entries.<br>"
        "<span class='tip-sell'>&lt; −100</span> — ⚠ Extreme oversold / bearish momentum. Avoid."
        "</div>"
    )

    TIP_MACD = (
        "<b>Daily MACD(12,26,9) — MACD Line vs Signal Line</b><br>"
        "Momentum oscillator. Shown value is the MACD line. Positive = MACD above Signal."
        "<div class='tip-action'>"
        "<span class='tip-buy'>▲ MACD &gt; Signal</span> — Bullish crossover. Trend strength increasing. Add confidence to buy.<br>"
        "<span class='tip-sell'>▼ MACD &lt; Signal</span> — Bearish crossover. Momentum fading. Consider reducing position.<br>"
        "Combines with RSI and CCI for a 3-indicator confirmation system."
        "</div>"
    )

    TIP_CLOSE = (
        "<b>Last Closing Price (₹)</b><br>"
        "Most recent daily closing price in Indian Rupees. Used as the reference for all SL and target calculations."
        "<div class='tip-action'>"
        "Compare with 52W High/Low to gauge where the stock is in its range.<br>"
        "ATR Stop-Loss and Swing SL are both calculated as a % below this price."
        "</div>"
    )

    TIP_52W = (
        "<b>Distance from 52-Week High (%)</b><br>"
        "How far the current price is below its 52-week high. Momentum stocks trade close to their highs."
        "<div class='tip-action'>"
        "<span class='tip-buy'>−5% to 0%</span> — Near 52W high. Strong momentum; potential breakout zone.<br>"
        "<span class='tip-watch'>−10% to −5%</span> — Modest pullback. Acceptable if other signals are bullish.<br>"
        "<span class='tip-sell'>&lt; −10%</span> — Significantly below highs. Needs strong catalyst to recover.<br>"
        "Sort ascending (least negative first) to find stocks at or near 52W highs."
        "</div>"
    )

    TIP_MCAP = (
        "<b>Market Capitalisation</b><br>"
        "Total market value of all outstanding shares in INR Crore."
        "<div class='tip-action'>"
        "<span class='tip-buy'>Large Cap (&gt;₹20K Cr)</span> — Lower risk, high liquidity. Suitable for larger positions.<br>"
        "<span class='tip-buy'>Mid Cap (₹5K–20K Cr)</span> — Balanced risk/reward. Core momentum plays.<br>"
        "<span class='tip-watch'>Small Cap (₹500–5K Cr)</span> — Higher volatility. Use smaller position sizes.<br>"
        "<span class='tip-sell'>Micro Cap (&lt;₹500 Cr)</span> — High risk. Verify liquidity before trading."
        "</div>"
    )

    TIP_DONCH_D = (
        "<b>Daily Donchian Channel Position (%)</b><br>"
        "How close the price is to the 20-day Donchian (highest high) channel. "
        "Near 0% = trading at or above the 20-day high — a breakout signal."
        "<div class='tip-action'>"
        "<span class='tip-buy'>≥ −2%</span> — At or near 20-day high. 🚀 Breakout territory. Strong buy zone.<br>"
        "<span class='tip-watch'>−2% to −10%</span> — Slight pullback from highs. Acceptable.<br>"
        "<span class='tip-sell'>&lt; −10%</span> — Well below highs. Momentum has stalled."
        "</div>"
    )

    TIP_DONCH_W = (
        "<b>Weekly Donchian Channel Position (%)</b><br>"
        "Same as D-Donch but on the weekly timeframe (20-week high). "
        "Weekly breakouts signal multi-month momentum."
        "<div class='tip-action'>"
        "<span class='tip-buy'>≥ −2%</span> — At multi-week highs. Strong weekly breakout — high-quality momentum.<br>"
        "<span class='tip-watch'>−2% to −10%</span> — Mild pullback. Consolidation phase.<br>"
        "<span class='tip-sell'>&lt; −10%</span> — Weekly momentum has cooled significantly."
        "</div>"
    )

    TIP_DONCH_M = (
        "<b>Monthly Donchian Channel Position (%)</b><br>"
        "Distance from the 20-month Donchian high. The strongest momentum stocks break out on all 3 Donchian timeframes simultaneously."
        "<div class='tip-action'>"
        "<span class='tip-buy'>All 3 Donchian ≥ −2%</span> — Multi-timeframe Donchian breakout. "
        "Rarest and most powerful momentum signal. Mark these as high-priority.<br>"
        "<span class='tip-sell'>&lt; −10%</span> — Monthly highs not yet challenged."
        "</div>"
    )

    TIP_N50 = (
        "<b>Rank vs Nifty 50</b><br>"
        "Percentile rank of this stock's momentum score compared to all 50 Nifty50 stocks in the scan. "
        "Shows relative strength against India's benchmark index."
        "<div class='tip-action'>"
        "<span class='tip-buy'>Top 20%</span> — Outperforming most Nifty50 blue-chips. Institutional-grade momentum.<br>"
        "<span class='tip-watch'>20–60%</span> — In-line with the index.<br>"
        "<span class='tip-sell'>Bottom 40%</span> — Underperforming the index. Avoid unless a specific catalyst exists."
        "</div>"
    )

    TIP_UNIV = (
        "<b>Rank vs All NSE Stocks</b><br>"
        "Percentile rank against every stock in the scan universe. "
        "This is the ultimate momentum filter — it shows where this stock stands in the entire NSE market."
        "<div class='tip-action'>"
        "<span class='tip-buy'>Top 10%</span> — 🏆 Market leader. These are the best momentum stocks in NSE right now.<br>"
        "<span class='tip-buy'>Top 25%</span> — Strong relative performer. High priority watchlist.<br>"
        "<span class='tip-watch'>25–50%</span> — Average momentum. Only trade if other signals are very strong.<br>"
        "<span class='tip-sell'>Bottom 50%</span> — Laggard. Capital is better deployed elsewhere.<br><br>"
        "💡 <b>Sort by this column descending</b> for an instant ranked list of today's strongest momentum stocks."
        "</div>"
    )

    return f"""
    <div class="sort-hint">
      💡 Click header to sort &nbsp;|&nbsp; <b>Shift+click</b> = add 2nd/3rd sort key &nbsp;|&nbsp;
      Click again to toggle ▲▼ &nbsp;|&nbsp; Shows {PAGE_TBL} rows per page &nbsp;|&nbsp;
      Hover <b>?</b> on any column for guidance
    </div>
    <div class="table-wrap">
      <table class="sum-table" id="sumtable">
        <thead><tr>
          {th('Ticker / Company', 'ticker', TIP_TICKER, 'left')}
          {th_plain('Phase',     TIP_PHASE)}
          {th_plain('Signal',    TIP_SIGNAL)}
          {th('Score',   'score',  TIP_SCORE)}
          {th('D-RSI/SMA','rsid', TIP_DRSI)}
          {th('W-RSI',    'rsiw', TIP_WRSI)}
          {th('M-RSI',    'rsim', TIP_MRSI)}
          {th('D-CCI',    'cci',  TIP_CCI)}
          {th('D-MACD',   'macd', TIP_MACD)}
          {th('Close',    'close',TIP_CLOSE)}
          {th('52W%',   'dist52', TIP_52W)}
          {th_plain('Market Cap', TIP_MCAP)}
          {th('D-Donch', 'donchd', TIP_DONCH_D)}
          {th('W-Donch', 'donchw', TIP_DONCH_W)}
          {th('M-Donch', 'donchm', TIP_DONCH_M, tip_left=True)}
          {th('vs N50',   'rn50',  TIP_N50,     tip_left=True)}
          {th('vs All',   'runiv', TIP_UNIV,    tip_left=True)}
        </tr></thead>
        <tbody id="tbl-body"></tbody>
      </table>
    </div>
    <div id="tbl-pager"></div>"""


def build_stock_card(d: dict, has_chart: bool) -> str:
    """
    Build a <details> card. Chart img is always present but src="" (lazy).
    JS fills it from data-src on first open.
    If has_chart=False, show placeholder instead.
    """
    fr_tags = ""
    if d["fresh_d"]: fr_tags += f' <span class="fresh-tag">🚀 Daily ({d["fresh_d_bars"]}d)</span>'
    if d["fresh_w"]: fr_tags += f' <span class="fresh-tag">📅 Weekly ({d["fresh_w_bars"]}w)</span>'
    n50_tag = ' <span class="n50-tag">NIFTY50</span>' if d["is_nifty50"] else ""
    sme_tag = ' <span class="sme-tag">SME</span>' if d["is_sme"] else ""
    idx_tags = ' ' + ' '.join(f'<span class="index-tag">{idx}</span>' for idx in d['indices']) if d['indices'] else ""
    sector_tag = f' <span class="sector-tag">{d["sector"]}</span>' if d['sector'] and d['sector'] != 'Unknown' else ""
    cap_cat, cap_cls = categorize_marketcap(d["marketcap"])
    cap_tag = f' <span class="{cap_cls}">{cap_cat}</span>' if cap_cat != "Unknown" else ""

    n50_rank  = _rank_pill(d["rank_nifty50"], d["rank_nifty50_pos"], d["rank_nifty50_of"])
    univ_rank = _rank_pill(d["rank_universe"], d["rank_univ_pos"],   d["rank_univ_of"])

    if has_chart:
        chart_html = (f'<div class="chart-wrap">'
                      f'<img class="lazy-chart" data-src="{CHART_OUTPUT_DIR}/{d["ticker"]}.png" '
                      f'src="data:image/gif;base64,R0lGODlhAQABAIAAAAAAAP///ywAAAAAAQABAAACAUwAOw==" '
                      f'alt="{d["ticker"]} chart" loading="lazy">'
                      f'</div>')
    else:
        chart_html = '<div class="chart-placeholder">📊 Chart not included (outside top limit)</div>'

    panels = _build_detail_panels(d)

    return f"""
<details class="stock-card" data-phase="{d['phase']}"
         data-fresh="{'1' if (d['fresh_d'] or d['fresh_w']) else '0'}"
         data-nifty="{'1' if d['is_nifty50'] else '0'}"
         data-sme="{'1' if d['is_sme'] else '0'}"
         data-cap="{cap_cls}"
         data-sector="{d.get('sector','')}"
         data-indices="{','.join(d.get('indices', []))}"
         data-ticker="{d['ticker']}"
         data-company="{d['company'][:50]}">
  <summary>
    <span class="card-arrow">▶</span>
    <span class="card-ticker">{d['ticker']}</span>
    <span class="card-price">₹{d['close']:,.2f}</span>
    {_phase_badge(d['phase'])}
    <span class="{d['sig_cls']}" style="font-weight:700">{d['signal']}</span>
    <span class="card-score">Score {d['score']}/21</span>
    {fr_tags}{n50_tag}{sme_tag}{cap_tag}{idx_tags}{sector_tag}
    <span style="margin-left:auto;color:var(--sub);font-size:11px;text-align:right">
      D {d['rsi_d']} W {d['rsi_w']} M {d['rsi_m']} RSI
      &nbsp;|&nbsp; N50: {n50_rank}
      &nbsp;|&nbsp; All: {univ_rank}
    </span>
  </summary>
  <div class="card-body">
    {chart_html}
    {panels}
  </div>
</details>"""


def _build_filter_options(all_results: list[dict]) -> tuple[str, str, str, str]:
    """
    Build sector, index, industry, and F&O <option> lists for filter dropdowns.

    Sectors  : every Sectoral/Thematic/Strategy index label a stock belongs to
               (s.sectors[] list) — so "NIFTY CPSE - Thematic" appears as its
               own filterable option even if a stock also has a Sectoral label.
    Indices  : all indices from _INDEX_MAP with analysed-stock counts.
    Industries: from _INDUSTRY_MAP (CSV "Industry" column) per stock.
    F&O      : simple yes/no — foSel dropdown has just two options beyond "All".
    """
    import html as _html

    analyzed_tickers = {d["ticker"] for d in all_results}

    # ── Fix sectors / indices / industry in-place (safety net for late xlsx load) ───────
    for d in all_results:
        ticker = d.get("ticker", "")
        if not ticker:
            continue
        # Re-derive sectors list from _SECTOR_MAP if missing/empty
        if not d.get("sectors"):
            raw = _SECTOR_MAP.get(ticker, [])
            d["sectors"] = raw if isinstance(raw, list) else ([raw] if raw else [])
        # Keep primary sector for display (first in sorted list)
        if not d.get("sector") or d["sector"] == "Unknown":
            d["sector"] = d["sectors"][0] if d["sectors"] else "Unknown"
        # Re-derive indices from _INDEX_MAP
        full_indices = [idx for idx, syms in _INDEX_MAP.items() if ticker in syms]
        if full_indices:
            d["indices"] = sorted(full_indices)
        # Re-derive is_fo
        if not d.get("is_fo"):
            d["is_fo"] = ticker in _FO_SET
        # Re-derive industry from _INDUSTRY_MAP if missing
        if not d.get("industry"):
            d["industry"] = _INDUSTRY_MAP.get(ticker, "")

    # ── Sectors (from s.sectors[] list — each label is a separate option) ────
    sec_counts: dict[str, int] = {}
    for d in all_results:
        for lbl in (d.get("sectors") or []):
            if lbl:
                sec_counts[lbl] = sec_counts.get(lbl, 0) + 1
    sec_opts = "".join(
        f'<option value="{_html.escape(s)}">{_html.escape(s)} ({c})</option>'
        for s, c in sorted(sec_counts.items()) if c > 0
    )

    # ── Indices (from _INDEX_MAP — all indices, with analysed counts) ─────────────
    idx_counts: dict[str, int] = {}
    for idx_name, syms in _INDEX_MAP.items():
        cnt = sum(1 for t in syms if t in analyzed_tickers)
        if cnt > 0:
            idx_counts[idx_name] = cnt
    for d in all_results:                          # belt-and-suspenders
        for i in (d.get("indices") or []):
            if i not in idx_counts:
                idx_counts[i] = 1
    idx_opts = "".join(
        f'<option value="{_html.escape(i)}">{_html.escape(i)} ({c})</option>'
        for i, c in sorted(idx_counts.items())
    )

    # ── Industries (from _INDUSTRY_MAP — "Industry" column in index CSVs) ────
    ind_counts: dict[str, int] = {}
    for d in all_results:
        ind = d.get("industry", "")
        if ind:
            ind_counts[ind] = ind_counts.get(ind, 0) + 1
    industry_opts = "".join(
        f'<option value="{_html.escape(i)}">{_html.escape(i)} ({c})</option>'
        for i, c in sorted(ind_counts.items()) if c > 0
    )

    # ── F&O count (for badge on "All F&O" option) ────────────────────────────
    # fo_opts is just a count — the actual options are hardcoded in HTML template
    n_fo = sum(1 for d in all_results if d.get("is_fo"))
    fo_opts = str(n_fo)   # passed as {fo_opts} into the template

    return sec_opts, idx_opts, industry_opts, fo_opts


_HTML_FIELDS = {
    'ticker','company','close','high52','low52','dist52','rsi_d','sma_d','rsi_w','sma_w',
    'rsi_m','sma_m','macd_l','macd_s','macd_l_w','macd_s_w','macd_l_m','macd_s_m',
    'cci','cci_w','cci_m','atr_sl','swing_sl','r_sl_pct','s_sl_pct','entry_note',
    'sell_conds','score','phase','signal','sig_cls','fresh_d','fresh_d_bars','fresh_w',
    'fresh_w_bars','donchian_d','donchian_w','donchian_m','is_nifty50','is_sme','is_fo',
    'sector','sectors','indices','industry','marketcap','cap_cat','cap_cls',
    'rank_nifty50','rank_nifty50_pos','rank_nifty50_of','rank_universe','rank_univ_pos',
    'rank_univ_of','fib_type','fib_levels','fib_base','sig_list','hist_sigs','has_chart',
    'ath_price','ath_date','ath_pct','is_ath','ath_time_str',
    'explosive_score','explosive_signals','vol_ratio',
    'bb_upper','bb_mid','bb_lower','bb_pct','bb_slope',
    'mfi','cci_200','macd_hist'
}


def _html_safe_stock(d: dict) -> dict:
    rec = {k: d[k] for k in _HTML_FIELDS if k in d}
    # Keep chart file reference (don't embed base64 — charts loaded on demand from file)
    if d.get('has_chart'):
        _cdir = ASX_CHART_OUTPUT_DIR if d['ticker'] in _ASX_STOCKS else CHART_OUTPUT_DIR
        chart_path = os.path.join(_cdir, f"{d['ticker']}.png").replace("\\", "/")
        rec['chart_path'] = chart_path
    return rec


def build_index_rsi_html(all_results: list[dict]) -> str:
    """
    Build an Index RSI Analysis tab HTML section.

    For each index in _INDEX_MAP (loaded from NseIndice CSV files):
      1. Show a table of constituent stocks with their Daily/Weekly/Monthly RSI
      2. Calculate combined (equal-weight average) index RSI and show it as a summary row

    Returns an HTML string to be embedded in the report.
    """
    import html as _html

    # Build lookup: ticker → stock data record
    ticker_map: dict[str, dict] = {d["ticker"]: d for d in all_results}
    analyzed = set(ticker_map.keys())

    def rsi_color(v) -> str:
        if v is None or v == "—":
            return "#888"
        try:
            v = float(v)
        except Exception:
            return "#888"
        if v >= 70:
            return "#ff4d6d"   # overbought red
        if v >= 55:
            return "#26d07c"   # bullish green
        if v >= 45:
            return "#ffd700"   # neutral gold
        if v >= 30:
            return "#ff9800"   # weak orange
        return "#ff4d6d"       # oversold red

    def rsi_badge(v, label="") -> str:
        if v is None:
            return '<span style="color:#555">—</span>'
        try:
            fv = float(v)
        except Exception:
            return '<span style="color:#555">—</span>'
        col = rsi_color(fv)
        lbl = f'<span style="font-size:10px;color:#888;margin-left:3px">{label}</span>' if label else ''
        return f'<span style="color:{col};font-weight:700">{fv:.1f}</span>{lbl}'

    # Only include indices that have at least 1 analysed stock
    indices_to_show = []
    for idx_name in sorted(_INDEX_MAP.keys()):
        syms = _INDEX_MAP[idx_name]
        members = [s for s in syms if s in analyzed]
        if members:
            indices_to_show.append((idx_name, sorted(members)))

    if not indices_to_show:
        return '<div style="padding:20px;color:#888">No index CSV constituents found in analysed stocks. Place CSVs in <code>india/NSE/NseIndice/</code>.</div>'

    parts = []

    # ── Tab nav pills ──────────────────────────────────────────────────────────
    nav_items = "".join(
        f'<button class="idx-tab-btn" onclick="showIdxTab({i},this)" id="idxbtn{i}">'
        f'{_html.escape(name)} <span style="font-size:10px;color:#888">({len(mems)})</span></button>'
        for i, (name, mems) in enumerate(indices_to_show)
    )
    parts.append(f'''
<div id="index-rsi-section" style="padding:16px 0">
  <h2 style="color:var(--cyan);margin-bottom:12px">📊 Index RSI Analysis
    <small style="font-size:12px;color:var(--sub);font-weight:400"> — Daily / Weekly / Monthly RSI per index</small>
  </h2>
  <div id="idx-tab-nav" style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:16px">
    {nav_items}
  </div>
''')

    # ── Per-index panels ────────────────────────────────────────────────────────
    for i, (idx_name, members) in enumerate(indices_to_show):
        display = "block" if i == 0 else "none"

        # Gather constituent RSIs
        rows_html = []
        rsi_d_vals, rsi_w_vals, rsi_m_vals = [], [], []

        for sym in members:
            d = ticker_map[sym]
            rd = d.get("rsi_d")
            rw = d.get("rsi_w")
            rm = d.get("rsi_m")
            phase = d.get("phase", "")
            signal = d.get("signal", "")
            close = d.get("close")
            industry = d.get("industry", "")
            sector_list = d.get("sectors", [])
            sector_str = sector_list[0] if sector_list else d.get("sector", "")

            if rd is not None:
                try:
                    rsi_d_vals.append(float(rd))
                except Exception:
                    pass
            if rw is not None:
                try:
                    rsi_w_vals.append(float(rw))
                except Exception:
                    pass
            if rm is not None:
                try:
                    rsi_m_vals.append(float(rm))
                except Exception:
                    pass

            phase_badge = f'<span class="badge badge-{phase}" style="font-size:10px">{phase}</span>' if phase else ""
            close_str = f"₹{float(close):,.2f}" if close else "—"
            ind_str = f'<span style="font-size:10px;color:#ff9800">{_html.escape(industry)}</span>' if industry else ""
            sec_str = f'<span style="font-size:10px;color:#ff9800">{_html.escape(sector_str[:30])}</span>' if sector_str and sector_str != "Unknown" else ""

            rows_html.append(f'''<tr>
  <td><b style="color:var(--cyan)">{_html.escape(sym)}</b></td>
  <td style="color:#c9d1d9;font-size:11px">{_html.escape(d.get("company","")[:28])}</td>
  <td>{ind_str or sec_str}</td>
  <td style="text-align:right">{close_str}</td>
  <td style="text-align:center">{rsi_badge(rd)}</td>
  <td style="text-align:center">{rsi_badge(rw)}</td>
  <td style="text-align:center">{rsi_badge(rm)}</td>
  <td style="text-align:center">{phase_badge}</td>
  <td style="text-align:center;font-size:10px;color:#888">{_html.escape(signal[:20]) if signal else "—"}</td>
</tr>''')

        # Combined index RSI (equal-weight average)
        avg_d = round(sum(rsi_d_vals) / len(rsi_d_vals), 1) if rsi_d_vals else None
        avg_w = round(sum(rsi_w_vals) / len(rsi_w_vals), 1) if rsi_w_vals else None
        avg_m = round(sum(rsi_m_vals) / len(rsi_m_vals), 1) if rsi_m_vals else None

        summary_bar = f'''
<div style="background:#161b22;border:1px solid #30363d;border-radius:10px;padding:14px 20px;margin-bottom:12px;display:flex;gap:24px;align-items:center;flex-wrap:wrap">
  <div style="color:#ffd700;font-weight:700;font-size:14px">📊 {_html.escape(idx_name)}</div>
  <div style="color:#888;font-size:11px">{len(members)} constituents analysed</div>
  <div style="display:flex;gap:16px;margin-left:auto;align-items:center">
    <div style="text-align:center">
      <div style="font-size:10px;color:#888;margin-bottom:2px">Daily RSI</div>
      <div style="font-size:22px;font-weight:700;color:{rsi_color(avg_d)}">{f"{avg_d:.1f}" if avg_d else "—"}</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:10px;color:#888;margin-bottom:2px">Weekly RSI</div>
      <div style="font-size:22px;font-weight:700;color:{rsi_color(avg_w)}">{f"{avg_w:.1f}" if avg_w else "—"}</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:10px;color:#888;margin-bottom:2px">Monthly RSI</div>
      <div style="font-size:22px;font-weight:700;color:{rsi_color(avg_m)}">{f"{avg_m:.1f}" if avg_m else "—"}</div>
    </div>
    <div style="text-align:center;padding-left:16px;border-left:1px solid #30363d">
      <div style="font-size:10px;color:#888;margin-bottom:2px">Index Strength</div>
      <div style="font-size:13px;font-weight:700;color:{rsi_color(avg_d)}">
        {"🔥 Strong Bull" if avg_d and avg_d>=65 else "📈 Bullish" if avg_d and avg_d>=55 else "➡️ Neutral" if avg_d and avg_d>=45 else "📉 Weak" if avg_d else "—"}
      </div>
    </div>
  </div>
</div>'''

        constituent_table = f'''
<div style="overflow-x:auto;max-height:480px;overflow-y:auto">
<table style="width:100%;border-collapse:collapse;font-size:12px">
  <thead style="position:sticky;top:0;background:#0d1117;z-index:2">
    <tr style="color:#888;border-bottom:1px solid #30363d">
      <th style="padding:6px 8px;text-align:left">Symbol</th>
      <th style="padding:6px 8px;text-align:left">Company</th>
      <th style="padding:6px 8px;text-align:left">Industry</th>
      <th style="padding:6px 8px;text-align:right">Close</th>
      <th style="padding:6px 8px;text-align:center">RSI-D</th>
      <th style="padding:6px 8px;text-align:center">RSI-W</th>
      <th style="padding:6px 8px;text-align:center">RSI-M</th>
      <th style="padding:6px 8px;text-align:center">Phase</th>
      <th style="padding:6px 8px;text-align:center">Signal</th>
    </tr>
  </thead>
  <tbody>
    {"".join(rows_html)}
  </tbody>
</table>
</div>'''

        parts.append(f'''
<div id="idx-panel-{i}" class="idx-panel" style="display:{display}">
  {summary_bar}
  {constituent_table}
</div>''')

    parts.append('''
</div>
<script>
function showIdxTab(idx, btn) {
  document.querySelectorAll('.idx-panel').forEach(p => p.style.display='none');
  document.querySelectorAll('.idx-tab-btn').forEach(b => b.classList.remove('active'));
  const panel = document.getElementById('idx-panel-'+idx);
  if(panel) panel.style.display='block';
  if(btn) btn.classList.add('active');
}
// Activate first tab on load
document.addEventListener('DOMContentLoaded', function() {
  const firstBtn = document.getElementById('idxbtn0');
  if(firstBtn) firstBtn.classList.add('active');
});
</script>''')

    return "\n".join(parts)


def build_html_report(all_results: list[dict], chart_data: dict[str, str],
                      run_ts: str, scanned: int) -> str:
    n_up  = sum(1 for d in all_results if d["phase"] == "UPTREND")
    n_sw  = sum(1 for d in all_results if d["phase"] == "SIDEWAYS")
    n_be  = sum(1 for d in all_results if d["phase"] == "BEARISH")
    n_fr  = sum(1 for d in all_results if d["fresh_d"] or d["fresh_w"])
    n_n50 = sum(1 for d in all_results if d["is_nifty50"])
    n_sme = sum(1 for d in all_results if d["is_sme"])
    total = len(all_results)

    # Signal counts for dropdown badges
    n_sig_sb  = sum(1 for d in all_results if d.get("sig_cls") == "sig-strong-buy")
    n_sig_buy = sum(1 for d in all_results if d.get("sig_cls") == "sig-buy")
    n_sig_wat = sum(1 for d in all_results if d.get("sig_cls") == "sig-watch")
    n_sig_av  = sum(1 for d in all_results if d.get("sig_cls") == "sig-avoid")
    n_sig_neu = sum(1 for d in all_results if d.get("sig_cls") == "sig-neutral")

    # ATH proximity counts for dropdown badges
    n_ath_at   = sum(1 for d in all_results if d.get("is_ath"))
    n_ath_5    = sum(1 for d in all_results if not d.get("is_ath") and d.get("ath_pct") is not None and d["ath_pct"] >= -5)
    n_ath_10   = sum(1 for d in all_results if not d.get("is_ath") and d.get("ath_pct") is not None and d["ath_pct"] >= -10)
    n_ath_20   = sum(1 for d in all_results if not d.get("is_ath") and d.get("ath_pct") is not None and d["ath_pct"] >= -20)
    n_ath_far  = sum(1 for d in all_results if d.get("ath_pct") is not None and d["ath_pct"] < -20)

    sec_opts, idx_opts, industry_opts, fo_opts = _build_filter_options(all_results)
    n_sectors = sec_opts.count('<option')
    n_indices = idx_opts.count('<option')
    n_industries = industry_opts.count('<option')
    n_fo      = int(fo_opts)

    # ── Build Index RSI Analysis section ─────────────────────────────────────
    index_rsi_html = build_index_rsi_html(all_results)

    # ── Add computed display fields + has_chart to each record ────────────────
    chart_tickers = set(chart_data.keys())
    for d in all_results:
        cap_cat, cap_cls = categorize_marketcap(d.get("marketcap"))
        d["cap_cat"]   = cap_cat
        d["cap_cls"]   = cap_cls
        d["has_chart"] = d["ticker"] in chart_tickers

    # ── Serialize to compact JSON (only the fields needed by HTML/JS) ───────
    safe_results = [_html_safe_stock(d) for d in all_results]
    stocks_json = json.dumps(safe_results, ensure_ascii=False, default=str, separators=(",", ":"))

    stat_boxes = f"""<div class="stats-row">
      <div class="stat-box cyan"><div class="val">{scanned}</div><div class="lbl">Scanned</div></div>
      <div class="stat-box green"><div class="val">{n_up}</div><div class="lbl">📈 Uptrend</div></div>
      <div class="stat-box gold"><div class="val">{n_sw}</div><div class="lbl">➡️ Sideways</div></div>
      <div class="stat-box red"><div class="val">{n_be}</div><div class="lbl">📉 Bearish</div></div>
      <div class="stat-box cyan"><div class="val">{n_fr}</div><div class="lbl">🚀 Fresh</div></div>
      <div class="stat-box gold"><div class="val">{n_n50}</div><div class="lbl">🏆 Nifty50</div></div>
      <div class="stat-box green"><div class="val">{n_sme}</div><div class="lbl">📊 SME</div></div>
    </div>"""

    filter_bar = f"""<div class="filter-section">
      <div class="filter-row1">
        <input id="searchInp" class="filter-input" type="text"
               placeholder="🔍 Search ticker / company…" oninput="onSearch(this.value)">
        <select id="capSel" class="filter-select" onchange="onDropChange()">
          <option value="all">💰 All Cap Sizes</option>
          <option value="cap-large">🟢 Large Cap (&gt;₹2L Cr)</option>
          <option value="cap-mid">🔵 Mid Cap (₹50K–2L Cr)</option>
          <option value="cap-small">🟡 Small Cap (₹5K–50K Cr)</option>
          <option value="cap-micro">🟠 Micro Cap (&lt;₹500 Cr)</option>
        </select>
        <select id="secSel" class="filter-select" onchange="onDropChange()">
          <option value="all">🏭 All Sectors / Industries ({n_sectors})</option>
          {sec_opts}
        </select>
        <select id="indSel" class="filter-select" onchange="onDropChange()">
          <option value="all">🏢 All Industries ({n_industries})</option>
          {industry_opts}
        </select>
        <select id="idxSel" class="filter-select" onchange="onDropChange()">
          <option value="all">📊 All Indices ({n_indices})</option>
          {idx_opts}
        </select>
        <select id="foSel" class="filter-select" onchange="onDropChange()">
          <option value="all">🔮 All F&amp;O / Cash ({n_fo})</option>
          <option value="fo">🔮 F&amp;O Stocks ({n_fo})</option>
          <option value="cash">💵 Cash Only</option>
        </select>
        <select id="sigSel" class="filter-select" onchange="onDropChange()">
          <option value="all">📶 All Signals</option>
          <option value="sig-strong-buy">🚀 Strong Buy ({n_sig_sb})</option>
          <option value="sig-buy">✅ Buy ({n_sig_buy})</option>
          <option value="sig-watch">👀 Watch ({n_sig_wat})</option>
          <option value="sig-avoid">❌ Avoid ({n_sig_av})</option>
          <option value="sig-neutral">⚪ Neutral ({n_sig_neu})</option>
        </select>
        <select id="athSel" class="filter-select" onchange="onDropChange()">
          <option value="all">🏔 All ATH Distances</option>
          <option value="at">🏆 At ATH ({n_ath_at})</option>
          <option value="w5">✅ Within 5% of ATH ({n_ath_5})</option>
          <option value="w10">🟡 Within 10% of ATH ({n_ath_10})</option>
          <option value="w20">🟠 Within 20% of ATH ({n_ath_20})</option>
          <option value="far">📉 More than 20% below ATH ({n_ath_far})</option>
        </select>
        <button class="clear-btn" onclick="clearAll()">✖ Clear</button>
        <span class="results-info">Showing <b id="rc">{total}</b> of {total} stocks</span>
      </div>
      <div class="filter-row2">
        <button class="phase-btn" data-phase="all"      onclick="filterPhase('all',this)">All ({total})</button>
        <button class="phase-btn" data-phase="fresh"    onclick="filterPhase('fresh',this)">🚀 Fresh ({n_fr})</button>
        <button class="phase-btn" data-phase="UPTREND"  onclick="filterPhase('UPTREND',this)">📈 Uptrend ({n_up})</button>
        <button class="phase-btn" data-phase="SIDEWAYS" onclick="filterPhase('SIDEWAYS',this)">➡️ Sideways ({n_sw})</button>
        <button class="phase-btn" data-phase="BEARISH"  onclick="filterPhase('BEARISH',this)">📉 Bearish ({n_be})</button>
        <button class="phase-btn" data-phase="nifty50"  onclick="filterPhase('nifty50',this)">🏆 Nifty50 ({n_n50})</button>
        <button class="phase-btn" data-phase="sme"      onclick="filterPhase('sme',this)">📊 SME ({n_sme})</button>
      </div>
      <div class="filter-row3">
        <span class="rank-sort-label">🏆 RANKING SORT:</span>
        <button class="rank-sort-btn" onclick="sortByRanking('rn50','desc',event)">⭐ Top vs N50</button>
        <button class="rank-sort-btn" onclick="sortByRanking('runiv','desc',event)">⭐ Top vs All Stocks</button>
        <button class="rank-sort-btn" onclick="sortByRanking('score','desc',event)">🎯 Highest Score</button>
        <button class="rank-sort-btn" onclick="sortByRanking('donchd','desc',event)">📈 Best D-Donch</button>
        <button class="rank-sort-btn" onclick="sortByRanking('donchw','desc',event)">📈 Best W-Donch</button>
        <button class="rank-sort-btn" onclick="sortByRanking('donchm','desc',event)">📈 Best M-Donch</button>
        <button class="rank-sort-btn" onclick="sortByRanking('rn50','asc',event)">📉 Lowest vs N50</button>
        <button class="rank-sort-btn" onclick="sortByRanking('runiv','asc',event)">📉 Lowest vs All</button>
        <div id="chips" class="active-chips"></div>
      </div>
    </div>"""

    sum_table = build_summary_table()   # skeleton only — JS fills tbody

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>RSI MTF Breakout Report — {run_ts}</title>
  <style>{_CSS}</style>
</head>
<body>
  <div class="app-header">
    <h1>📈 RSI Multi-Timeframe Breakout Report <small style="font-size:13px;color:var(--sub);font-weight:400">v4.0</small></h1>
    <div class="subtitle">
      NSE EQ Universe &nbsp;|&nbsp; {run_ts} IST &nbsp;|&nbsp;
      RSI(14) D/W/M + MACD(12,26) + CCI(20) &nbsp;|&nbsp;
      Ranked vs Nifty50 &amp; All NSE &nbsp;|&nbsp;
      {len(chart_data)} charts · virtual render (no browser hang)
    </div>
    {stat_boxes}
  </div>

  {filter_bar}

  <div class="table-section">
    {sum_table}
  </div>

  <div class="cards-section">
    <h2>🔍 DETAILED ANALYSIS — {total} stocks &nbsp;
      <span style="font-weight:400;font-size:11px">(click ▶ to expand · {PAGE_CARDS} cards loaded at a time)</span>
    </h2>
    <div id="cards-container"></div>
  </div>

  <div class="section-divider" style="margin:24px 0;border-top:1px solid #30363d"></div>
  {index_rsi_html}

  <div class="footer">
    RSI MTF Report v4.0 &nbsp;|&nbsp; {run_ts} &nbsp;|&nbsp;
    <b>Not financial advice.</b><br>
    Entry: RSI D+W+M &gt; SMA + CCI&gt;0 + MACD&gt;Signal &nbsp;|&nbsp;
    SL: 2×ATR or swing low &nbsp;|&nbsp;
    Exit: RSI crosses below SMA or CCI &lt; −100
  </div>

  <script>
const STOCKS={stocks_json};
const CHART_DIR={json.dumps(CHART_OUTPUT_DIR)};
const PAGE_TBL={PAGE_TBL};
const PAGE_CARDS={PAGE_CARDS};
{_JS}
  </script>
</body>
</html>"""


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main(force_charts: bool = False):
    os.system("cls" if os.name == "nt" else "clear")
    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║  📈  RSI MTF BREAKOUT REPORT  v4.0                               ║")
    print("║      Smart Cache · Batch Download · Virtual HTML · No Hang       ║")
    print("╚══════════════════════════════════════════════════════════════════╝")
    print(f"   {RUN_TS} IST  |  Errors → {ERROR_LOG}\n")
    print(f"   Start time : {START_TS} IST\n")
    if force_charts:
        print("   Force chart regeneration enabled — cache will be bypassed for all charts.\n")
    log_info(f"=== Scan started {START_TS} ===")

    # ── Step 1: Universe ──────────────────────────────────────────
    print("▶  STEP 1/4  Build universe")
    tickers = load_universe()
    print(f"   {len(tickers)} stocks loaded  |  {len(_COMPANY_MAP)} company names\n")
    log_info(f"Universe: {len(tickers)} stocks, {len(_COMPANY_MAP)} company names")

    # ── Step 2: Smart incremental cache update ────────────────────
    print("▶  STEP 2/4  Smart cache update (incremental download)")
    print("   ─────────────────────────────────────────────────────────")
    dl_stats = prefetch_all(tickers)
    log_info(f"Download: fresh={dl_stats['fresh']} updated={dl_stats['stale_updated']} "
             f"new={dl_stats['new_downloaded']} failed={dl_stats['failed']}")

    # ── Step 3: Analyse (all data comes from cache — no downloads) ─
    print("▶  STEP 3/4  Analyse (Daily · Weekly · Monthly indicators)")
    print("   ─────────────────────────────────────────────────────────")
    results, errors = [], 0
    t0, total = time.time(), len(tickers)

    for i, ticker in enumerate(tickers, 1):
        pct  = i / total * 100
        fill = int(pct / 2)
        elapsed = time.time() - t0
        current_time = datetime.now().strftime("%H:%M:%S")
        sys.stdout.write(
            f"\r  [{'█'*fill}{'░'*(50-fill)}] {pct:5.1f}%  {i:>4}/{total}  "
            f"{ticker:<14}  ok={len(results)}  err={errors}  "
            f"{current_time}  {format_timespan(elapsed)}"
        )
        sys.stdout.flush()

        try:
            res = analyze_stock(ticker)
        except Exception as _ae:
            log_info(f"analyze_stock error {ticker}: {_ae}")
            res = None
        if res:
            results.append(res)
        else:
            errors += 1

    elapsed = time.time() - t0
    print(f"\n\n   ✓ {len(results)} ok  |  {errors} failed  |  {elapsed:.0f}s")
    log_info(f"Analysis done: {len(results)} ok, {errors} failed, {elapsed:.0f}s")

    if not results:
        print("  ⚠️  No results produced — generating diagnostic HTML report.")
        log_info("No results produced during scan; building empty HTML report.")

    # ── Rankings ──────────────────────────────────────────────────
    print("   Computing rankings vs Nifty50 and full universe...")
    results = compute_rankings(results)
    results.sort(key=lambda d: (d["rank_universe"], d["score"]), reverse=True)

    n50_in_scan = sum(1 for d in results if d["is_nifty50"])
    sme_in_scan = sum(1 for d in results if d["is_sme"])
    print(f"   Nifty50 stocks in scan: {n50_in_scan}/{len(NIFTY50)}")
    print(f"   SME stocks in scan    : {sme_in_scan}  "
          f"(of {len(_SME_STOCKS)} in SME universe)\n")

    # ── Step 4: HTML ──────────────────────────────────────────────
    print("▶  STEP 4/4  Generate charts + build HTML")
    print("   ─────────────────────────────────────────────────────────")

    chart_candidates = results if MAX_CHART_STOCKS <= 0 else results[:MAX_CHART_STOCKS]
    chart_tickers = [d["ticker"] for d in chart_candidates]
    chart_data    = {}
    os.makedirs(CHART_OUTPUT_DIR, exist_ok=True)
    print(f"   Generating charts for {'all' if MAX_CHART_STOCKS <= 0 else 'top'} {len(chart_tickers)} stocks...")

    meta = _load_chart_cache_meta()
    stale = []
    cached = 0
    for d in chart_candidates:
        ticker = d["ticker"]
        _chart_dir = ASX_CHART_OUTPUT_DIR if ticker in _ASX_STOCKS else CHART_OUTPUT_DIR
        chart_path = os.path.join(_chart_dir, f"{ticker}.png").replace("\\", "/")
        chart_hash = _compute_chart_hash(d)
        if not force_charts and chart_hash and meta.get(ticker, {}).get("hash") == chart_hash and os.path.exists(chart_path):
            chart_data[ticker] = chart_path
            cached += 1
            continue
        stale.append((d, chart_hash, chart_path))

    print(f"   Reusing {cached}/{len(chart_tickers)} cached charts")

    if stale:
        workers = min(CHART_WORKERS, len(stale))
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(_generate_chart_worker, d): (d["ticker"], hash_val, path)
                       for d, hash_val, path in stale}
            completed = 0
            for future in as_completed(futures):
                ticker, chart_hash, chart_path = futures[future]
                completed += 1
                try:
                    _, generated_path = future.result()
                except Exception as exc:
                    generated_path = ""
                    log_error(ticker, get_company_name(ticker), "CHART-PARALLEL", exc)
                if generated_path:
                    chart_data[ticker] = generated_path
                    meta[ticker] = {"hash": chart_hash, "updated_at": datetime.now().strftime("%d %b %Y %H:%M")}
                elif os.path.exists(chart_path):
                    chart_data[ticker] = chart_path
                else:
                    print(f"\n   ⚠️ Chart failed for {ticker} — see {ERROR_LOG}")
                sys.stdout.write(f"\r   Charts completed: {completed}/{len(stale)}")
                sys.stdout.flush()
        print()
        _save_chart_cache_meta(meta)
    else:
        print("   No charts to generate.")

    print(f"   {len(chart_data)}/{len(chart_tickers)} charts generated")

    # Strip raw series before building HTML
    all_light = [{k: v for k, v in d.items() if not k.startswith("_")} for d in results]

    # Terminal summary
    print(f"\n   Phase summary:")
    print(f"     Uptrend : {sum(1 for d in results if d['phase']=='UPTREND')}")
    print(f"     Sideways: {sum(1 for d in results if d['phase']=='SIDEWAYS')}")
    print(f"     Bearish : {sum(1 for d in results if d['phase']=='BEARISH')}")
    print(f"   Fresh breakouts (Daily): {sum(1 for d in results if d['fresh_d'])}")
    print(f"   Fresh breakouts (Weekly): {sum(1 for d in results if d['fresh_w'])}\n")

    html = build_html_report(all_light, chart_data, RUN_TS, len(results))

    # Remove any old timestamped reports from previous runs
    import glob as _glob
    for _old in _glob.glob("rsi_mtf_report_*[0-9].html"):
        try: os.remove(_old)
        except: pass

    _TMP_HTML = OUTPUT_HTML + ".tmp"
    with open(_TMP_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    os.replace(_TMP_HTML, OUTPUT_HTML)   # atomic rename — no partial HTML on crash

    size_mb = os.path.getsize(OUTPUT_HTML) / 1024 / 1024
    total_elapsed = time.time() - START_TIME
    abs_path = os.path.abspath(OUTPUT_HTML)
    log_info(f"HTML saved: {abs_path} ({size_mb:.1f} MB)")
    print(f"\n  ✅ HTML saved : {abs_path}")
    print(f"     Size       : {size_mb:.1f} MB")
    print(f"     Stocks     : {len(results)}")
    print(f"  📋 Error log  : {ERROR_LOG}  ({errors} entries)")
    print(f"  Open HTML in any browser — cards start collapsed, charts lazy-load on expand\n")
    print(f"  Total runtime: {format_timespan(total_elapsed)}\n")

    print("  ─────────────────────────────────────────────────────────")
    print("  STRATEGY  ENTRY : RSI(14) D+W+M > SMA + CCI>0 + MACD>Signal")
    print("            SL    : 2×ATR(14) OR 1% below last swing low")
    print("            TARGET: Fib Extension from swing low→high (uptrend)")
    print("            EXIT  : RSI(14) daily crosses below SMA(14)")
    print("  ─────────────────────────────────────────────────────────\n")
    log_info(f"=== Run complete ===")


def parse_args():
    parser = argparse.ArgumentParser(description='RSI MTF report generator')
    parser.add_argument('-help', action='help',
                        help='Show this help message and exit.')
    parser.add_argument('--force-charts', action='store_true',
                        help='Always regenerate chart images, ignoring cached chart hashes.')
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    main(force_charts=args.force_charts)