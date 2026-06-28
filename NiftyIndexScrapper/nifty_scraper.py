"""
NIFTY Indices Scraper + Excel Builder
======================================
Run on YOUR machine (not in Claude sandbox — niftyindices.com is blocked there).

SETUP (one-time):
    pip install playwright openpyxl pandas requests
    playwright install chromium

RUN:
    python nifty_scraper.py

OUTPUT:
    NIFTY_Indices_Master.xlsx  (in the same folder)
"""

from __future__ import annotations
import io, time, sys
from pathlib import Path
from datetime import datetime
from typing import Optional

import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# INDEX REGISTRY
# (Display Name, Category, URL slug, CSV filename)
# Page URL : https://www.niftyindices.com/indices/equity/<slug>
# CSV  URL : https://www.niftyindices.com/IndexConstituent/<csv_file>
# ─────────────────────────────────────────────────────────────────────────────

BASE_CSV  = "https://www.niftyindices.com/IndexConstituent/"
BASE_PAGE = "https://www.niftyindices.com/indices/equity/"

INDICES = [
    # ── BROAD BASED ──────────────────────────────────────────────────────────
    ("NIFTY 50",                        "Broad Based", "broad-based-indices/NIFTY--50",                          "ind_nifty50list.csv"),
    ("NIFTY NEXT 50",                   "Broad Based", "broad-based-indices/NIFTY-NEXT-50",                      "ind_niftynext50list.csv"),
    ("NIFTY 100",                       "Broad Based", "broad-based-indices/NIFTY-100",                          "ind_nifty100list.csv"),
    ("NIFTY 200",                       "Broad Based", "broad-based-indices/NIFTY-200",                          "ind_nifty200list.csv"),
    ("NIFTY 500",                       "Broad Based", "broad-based-indices/NIFTY-500",                          "ind_nifty500list.csv"),
    ("NIFTY TOTAL MARKET",              "Broad Based", "broad-based-indices/NIFTY-TOTAL-MARKET",                 "ind_niftytotalmarket_list.csv"),
    ("NIFTY MIDCAP 50",                 "Broad Based", "broad-based-indices/NIFTY-MIDCAP-50",                    "ind_niftymidcap50list.csv"),
    ("NIFTY MIDCAP 100",                "Broad Based", "broad-based-indices/NIFTY-MIDCAP-100",                   "ind_niftymidcap100list.csv"),
    ("NIFTY MIDCAP 150",                "Broad Based", "broad-based-indices/NIFTY-MIDCAP-150",                   "ind_niftymidcap150list.csv"),
    ("NIFTY SMALLCAP 50",               "Broad Based", "broad-based-indices/NIFTY-SMALLCAP-50",                  "ind_niftysmallcap50list.csv"),
    ("NIFTY SMALLCAP 100",              "Broad Based", "broad-based-indices/NIFTY-SMALLCAP-100",                 "ind_niftysmallcap100list.csv"),
    ("NIFTY SMALLCAP 250",              "Broad Based", "broad-based-indices/NIFTY-SMALLCAP-250",                 "ind_niftysmallcap250list.csv"),
    ("NIFTY LARGEMIDCAP 250",           "Broad Based", "broad-based-indices/NIFTY-LARGEMIDCAP-250",              "ind_niftylargemidcap250list.csv"),
    ("NIFTY MIDSMALLCAP 400",           "Broad Based", "broad-based-indices/NIFTY-MIDSMALLCAP-400",              "ind_niftymidsmallcap400list.csv"),
    ("NIFTY MICROCAP 250",              "Broad Based", "broad-based-indices/NIFTY-MICROCAP-250",                 "ind_niftymicrocap250_list.csv"),
    # ── SECTORAL ─────────────────────────────────────────────────────────────
    ("NIFTY AUTO",                      "Sectoral",    "sectoral-indices/NIFTY-AUTO",                            "ind_niftyautolist.csv"),
    ("NIFTY BANK",                      "Sectoral",    "sectoral-indices/NIFTY-BANK",                            "ind_niftybanklist.csv"),
    ("NIFTY FINANCIAL SERVICES",        "Sectoral",    "sectoral-indices/NIFTY-FINANCIAL-SERVICES",              "ind_niftyfinancelist.csv"),
    ("NIFTY FINANCIAL SERVICES EX-BANK","Sectoral",    "sectoral-indices/NIFTY-FINANCIAL-SERVICES-Ex-Bank",      "ind_niftyfinancialservicesexbank_list.csv"),
    ("NIFTY FMCG",                      "Sectoral",    "sectoral-indices/NIFTY-FMCG",                            "ind_niftyfmcglist.csv"),
    ("NIFTY IT",                        "Sectoral",    "sectoral-indices/NIFTY-IT",                              "ind_niftyitlist.csv"),
    ("NIFTY MEDIA",                     "Sectoral",    "sectoral-indices/NIFTY-MEDIA",                           "ind_niftymedialist.csv"),
    ("NIFTY METAL",                     "Sectoral",    "sectoral-indices/NIFTY-METAL",                           "ind_niftymetallist.csv"),
    ("NIFTY PHARMA",                    "Sectoral",    "sectoral-indices/NIFTY-PHARMA",                          "ind_niftypharmalist.csv"),
    ("NIFTY PSU BANK",                  "Sectoral",    "sectoral-indices/NIFTY-PSU-BANK",                        "ind_niftypsubanklist.csv"),
    ("NIFTY PRIVATE BANK",              "Sectoral",    "sectoral-indices/NIFTY-PRIVATE-BANK",                    "ind_niftypvtbanklist.csv"),
    ("NIFTY REALTY",                    "Sectoral",    "sectoral-indices/NIFTY-REALTY",                          "ind_niftyrealtylist.csv"),
    ("NIFTY HEALTHCARE",                "Sectoral",    "sectoral-indices/NIFTY-HEALTHCARE-INDEX",                "ind_niftyhealthcarelist.csv"),
    ("NIFTY OIL & GAS",                 "Sectoral",    "sectoral-indices/NIFTY-OIL-GAS",                         "ind_niftyoilgaslist.csv"),
    ("NIFTY CONSUMER DURABLES",         "Sectoral",    "sectoral-indices/NIFTY-CONSUMER-DURABLES",               "ind_niftyconsumerdurableslist.csv"),
    ("NIFTY CAPITAL MARKETS",           "Sectoral",    "sectoral-indices/NIFTY-CAPITAL-MARKETS",                 "ind_niftycapitalmarketslist.csv"),
    # ── THEMATIC ─────────────────────────────────────────────────────────────
    ("NIFTY COMMODITIES",               "Thematic",    "thematic-indices/NIFTY-COMMODITIES",                     "ind_niftycommoditieslist.csv"),
    ("NIFTY INDIA CONSUMPTION",         "Thematic",    "thematic-indices/NIFTY-INDIA-CONSUMPTION",               "ind_niftyconsumptionlist.csv"),
    ("NIFTY CPSE",                      "Thematic",    "thematic-indices/NIFTY-CPSE",                            "ind_niftycpselist.csv"),
    ("NIFTY ENERGY",                    "Thematic",    "thematic-indices/NIFTY-ENERGY",                          "ind_niftyenergylist.csv"),
    ("NIFTY INFRASTRUCTURE",            "Thematic",    "thematic-indices/NIFTY-INFRASTRUCTURE",                  "ind_niftyinfralist.csv"),
    ("NIFTY MNC",                       "Thematic",    "thematic-indices/NIFTY-MNC",                             "ind_niftymnclist.csv"),
    ("NIFTY PSE",                       "Thematic",    "thematic-indices/NIFTY-PSE",                             "ind_niftypselist.csv"),
    ("NIFTY SERVICES SECTOR",           "Thematic",    "thematic-indices/NIFTY-SERVICES-SECTOR",                 "ind_niftyservicesectorlist.csv"),
    ("NIFTY INDIA DIGITAL",             "Thematic",    "thematic-indices/NIFTY-INDIA-DIGITAL",                   "ind_niftyindiadigital_list.csv"),
    ("NIFTY INDIA MANUFACTURING",       "Thematic",    "thematic-indices/NIFTY-INDIA-MANUFACTURING",             "ind_niftyindiamanufacturing_list.csv"),
    ("NIFTY INDIA DEFENCE",             "Thematic",    "thematic-indices/NIFTY-INDIA-DEFENCE",                   "ind_niftyindiadefence_list.csv"),
    ("NIFTY MOBILITY",                  "Thematic",    "thematic-indices/NIFTY-MOBILITY",                        "ind_niftymobility_list.csv"),
    ("NIFTY INDIA TOURISM",             "Thematic",    "thematic-indices/NIFTY-INDIA-TOURISM",                   "ind_niftytourism_list.csv"),
    # ── STRATEGY ─────────────────────────────────────────────────────────────
    ("NIFTY DIVIDEND OPPORTUNITIES 50",     "Strategy","strategy-indices/NIFTY-DIVIDEND-OPPORTUNITIES-50",        "ind_niftydividendopportunities50list.csv"),
    ("NIFTY GROWTH SECTORS 15",             "Strategy","strategy-indices/NIFTY-GROWTH-SECTORS-15",                "ind_niftygrowthsectors15list.csv"),
    ("NIFTY50 EQUAL WEIGHT",                "Strategy","strategy-indices/NIFTY50-EQUAL-WEIGHT",                   "ind_nifty50EWlist.csv"),
    ("NIFTY100 EQUAL WEIGHT",               "Strategy","strategy-indices/NIFTY100-EQUAL-WEIGHT",                  "ind_nifty100EWlist.csv"),
    ("NIFTY100 LOW VOLATILITY 30",          "Strategy","strategy-indices/NIFTY100-LOW-VOLATILITY-30",             "ind_nifty100lowvol30list.csv"),
    ("NIFTY ALPHA 50",                      "Strategy","strategy-indices/NIFTY-ALPHA-50",                         "ind_niftyalpha50list.csv"),
    ("NIFTY50 VALUE 20",                    "Strategy","strategy-indices/NIFTY50-VALUE-20",                       "ind_nifty50value20list.csv"),
    ("NIFTY MIDCAP150 QUALITY 50",          "Strategy","strategy-indices/NIFTY-MIDCAP150-QUALITY-50",             "ind_niftymidcap150quality50list.csv"),
    ("NIFTY QUALITY LOW-VOLATILITY 30",     "Strategy","strategy-indices/NIFTY-QUALITY-LOW-VOLATILITY-30",        "ind_niftyqlvlist.csv"),
    ("NIFTY ALPHA LOW-VOLATILITY 30",       "Strategy","strategy-indices/NIFTY-ALPHA-LOW-VOLATILITY-30",          "ind_niftyalpha_lowvol30list.csv"),
    ("NIFTY200 MOMENTUM 30",                "Strategy","strategy-indices/NIFTY200-MOMENTUM-30",                   "ind_nifty200momentum30list.csv"),
    ("NIFTY500 MOMENTUM 50",                "Strategy","strategy-indices/NIFTY500-MOMENTUM-50",                   "ind_nifty500momentum50list.csv"),
    ("NIFTY200 ALPHA 30",                   "Strategy","strategy-indices/NIFTY200-ALPHA-30",                      "ind_nifty200alpha30list.csv"),
    ("NIFTY MIDSMALL MOMENTUM QUALITY 100", "Strategy","strategy-indices/NIFTY-MIDSMALL-MOMENTUM-QUALITY-100",    "ind_niftymidsmall_momentum_quality100_list.csv"),
]

# ─────────────────────────────────────────────────────────────────────────────
# DOWNLOAD
# Strategy 1: direct CSV via requests (fast, works most of the time)
# Strategy 2: Playwright browser — visits the index page, clicks the
#             "Index Constituent" download button, captures the file
# ─────────────────────────────────────────────────────────────────────────────

def _parse_csv(text: str) -> Optional[pd.DataFrame]:
    text = text.lstrip("\ufeff")          # strip BOM
    if "Company Name" not in text and "Symbol" not in text:
        return None
    df = pd.read_csv(io.StringIO(text))
    df.columns = [c.strip() for c in df.columns]
    df = df[df["Symbol"].notna() & (df["Symbol"].str.strip() != "")]
    return df if len(df) else None


def fetch_all_csvs() -> dict[str, pd.DataFrame]:
    session = requests.Session()
    session.headers.update({
        "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                           "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })

    print("  Warming up session …")
    try:
        session.get("https://www.niftyindices.com/", timeout=20)
    except Exception:
        pass

    results: dict[str, pd.DataFrame] = {}
    failed_list = []

    for idx_name, category, slug, csv_file in INDICES:
        csv_url  = BASE_CSV  + csv_file
        page_url = BASE_PAGE + slug
        session.headers["Referer"] = page_url
        print(f"  [{category:<12}] {idx_name:<45} … ", end="", flush=True)
        try:
            r = session.get(csv_url, timeout=20)
            df = _parse_csv(r.text) if r.status_code == 200 else None
            if df is not None:
                results[idx_name] = df
                print(f"✓ {len(df)} stocks")
            else:
                print(f"✗ HTTP {r.status_code} → Playwright queue")
                failed_list.append((idx_name, category, slug, csv_file, csv_url, page_url))
        except Exception as e:
            print(f"✗ {e} → Playwright queue")
            failed_list.append((idx_name, category, slug, csv_file, csv_url, page_url))
        time.sleep(0.35)

    # ── Playwright fallback ───────────────────────────────────────────────────
    if not failed_list:
        return results

    print(f"\n  {len(failed_list)} indices need Playwright …")
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  ✗ Playwright not installed: pip install playwright && playwright install chromium")
        return results

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            accept_downloads=True,
        )
        page = ctx.new_page()

        # Download-button selectors (in priority order)
        DOWNLOAD_SELECTORS = [
            "a[href*='IndexConstituent']",
            "a[href*='indexconstituent']",
            "a:has-text('Index Constituent')",
            "button:has-text('Index Constituent')",
            "a:has-text('Download')",
            ".download-btn",
            "[data-target='#downloadModal']",
        ]

        for idx_name, category, slug, csv_file, csv_url, page_url in failed_list:
            print(f"  [{category:<12}] {idx_name:<45} … ", end="", flush=True)
            csv_text = None

            try:
                # ── method A: intercept network response ──────────────────
                def on_resp(resp):
                    nonlocal csv_text
                    if csv_file.lower() in resp.url.lower() and resp.status == 200:
                        try:
                            csv_text = resp.text()
                        except Exception:
                            pass

                page.on("response", on_resp)
                page.goto(page_url, wait_until="networkidle", timeout=30_000)
                time.sleep(2)

                if csv_text:
                    df = _parse_csv(csv_text)
                    if df is not None:
                        results[idx_name] = df
                        print(f"✓ {len(df)} stocks (intercepted)")
                        continue

                # ── method B: click download button ──────────────────────
                for sel in DOWNLOAD_SELECTORS:
                    try:
                        loc = page.locator(sel).first
                        if not loc.is_visible(timeout=1500):
                            continue
                        href = loc.get_attribute("href") or ""
                        if href.startswith("http"):
                            # direct link — grab via requests using page cookies
                            cookies = {c["name"]: c["value"] for c in ctx.cookies()}
                            r2 = requests.get(href, cookies=cookies, timeout=20,
                                              headers={"Referer": page_url,
                                                       "User-Agent": "Mozilla/5.0"})
                            df = _parse_csv(r2.text) if r2.status_code == 200 else None
                        else:
                            # trigger download
                            with page.expect_download(timeout=12_000) as dl_info:
                                loc.click()
                            dl_path = dl_info.value.path()
                            csv_text = Path(dl_path).read_text(encoding="utf-8-sig",
                                                               errors="replace")
                            df = _parse_csv(csv_text)

                        if df is not None:
                            results[idx_name] = df
                            print(f"✓ {len(df)} stocks (click)")
                            break
                    except Exception:
                        continue
                else:
                    # ── method C: cookie-reuse on direct CSV URL ──────────
                    cookies = {c["name"]: c["value"] for c in ctx.cookies()}
                    r3 = requests.get(csv_url, cookies=cookies, timeout=20,
                                      headers={"Referer": page_url,
                                               "User-Agent": "Mozilla/5.0"})
                    df = _parse_csv(r3.text) if r3.status_code == 200 else None
                    if df is not None:
                        results[idx_name] = df
                        print(f"✓ {len(df)} stocks (cookie-reuse)")
                    else:
                        print(f"✗ gave up")

            except Exception as e:
                print(f"✗ {e}")

        browser.close()

    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

CLR = {
    "broad_dk": "1F3864", "broad_lt": "D6E4F0",
    "sec_dk":   "7B2C00", "sec_lt":   "FCE4D6",
    "the_dk":   "1A5231", "the_lt":   "E2EFDA",
    "str_dk":   "3B006F", "str_lt":   "E8D5F5",
    "all_dk":   "2C3E50",
    "white":    "FFFFFF", "gray":     "F2F2F2",
    "txt":      "000000", "url":      "0563C1",
}

CAT_CLR = {
    "Broad Based": (CLR["broad_dk"], CLR["broad_lt"]),
    "Sectoral":    (CLR["sec_dk"],   CLR["sec_lt"]),
    "Thematic":    (CLR["the_dk"],   CLR["the_lt"]),
    "Strategy":    (CLR["str_dk"],   CLR["str_lt"]),
}

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _w(ws, r, c, v, bold=False, clr="000000", bg="FFFFFF",
       center=False, wrap=False, sz=10, italic=False, underline=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(name="Arial", bold=bold, color=clr, size=sz,
                          italic=italic, underline=underline)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=wrap)
    cell.border    = _border()
    return cell

def _title(ws, text, ncols, bg, sz=13, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, size=sz, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32

def _sub(ws, ncols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1,
                value=f"Source: www.niftyindices.com  |  Generated: {datetime.today():%d %b %Y}")
    c.font      = Font(name="Arial", italic=True, size=9, color="595959")
    c.fill      = PatternFill("solid", fgColor="EDEDED")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 16

def _hdr(ws, row, col, text, bg):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _border()
    ws.row_dimensions[row].height = 22

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Summary sheet ─────────────────────────────────────────────────────────────
def _build_summary(wb, data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = CLR["broad_dk"]

    NC = 8
    _title(ws, "NSE INDIA — NIFTY EQUITY INDICES  |  MASTER DIRECTORY", NC, CLR["all_dk"], sz=14)
    _sub(ws, NC)
    ws.row_dimensions[3].height = 5

    hdrs = ["#", "Category", "Index Name", "Page URL", "CSV Download URL",
            "Stocks", "Symbols", "Industries"]
    _widths(ws, [4, 15, 36, 55, 55, 8, 65, 38])
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, 4, c, h, CLR["all_dk"])

    row = 4
    for i, (idx_name, cat, slug, csv_file) in enumerate(INDICES, 1):
        row += 1
        dk, lt = CAT_CLR[cat]
        bg  = lt if i % 2 else CLR["white"]
        df  = data.get(idx_name)
        n   = len(df) if df is not None else "—"
        sym = ", ".join(df["Symbol"].str.strip().tolist()) if df is not None else "not fetched"
        ind = ", ".join(sorted(df["Industry"].dropna().unique())) \
              if df is not None and "Industry" in df.columns else "—"
        pg  = BASE_PAGE + slug
        csv = BASE_CSV  + csv_file

        for c, (v, b, fc, ctr) in enumerate(zip(
            [i, cat, idx_name, pg, csv, n, sym, ind],
            [False, True, True, False, False, True, False, False],
            [CLR["txt"], dk, dk, CLR["url"], CLR["url"],
             CLR["txt"], CLR["txt"], CLR["txt"]],
            [True, False, False, False, False, True, False, False],
        ), 1):
            ul = "single" if c in (4, 5) else None
            _w(ws, row, c, v, bold=b, clr=fc, bg=bg, center=ctr, wrap=True,
               sz=9, underline=ul)
        ws.row_dimensions[row].height = 30

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(NC)}{row}"


# ── Index-list sheet per category ─────────────────────────────────────────────
def _build_cat_list(wb, cat, cat_entries, data):
    dk, lt = CAT_CLR[cat]
    ws = wb.create_sheet(cat.replace(" ", "_"))
    ws.sheet_properties.tabColor = dk

    _title(ws, f"NSE INDIA — {cat.upper()} INDICES", 6, dk)
    _sub(ws, 6)
    hdrs = ["#", "Index Name", "Page URL", "CSV Download URL", "Stocks", "Industries"]
    _widths(ws, [4, 36, 54, 54, 8, 48])
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, 4, c, h, dk)

    row = 4
    for i, (idx_name, cat_, slug, csv_file) in enumerate(cat_entries, 1):
        row += 1
        bg  = lt if i % 2 else CLR["white"]
        df  = data.get(idx_name)
        n   = len(df) if df is not None else "—"
        ind = ", ".join(sorted(df["Industry"].dropna().unique())) \
              if df is not None and "Industry" in df.columns else "—"
        pg  = BASE_PAGE + slug
        csv = BASE_CSV  + csv_file

        for c, (v, b, fc, ctr) in enumerate(zip(
            [i, idx_name, pg, csv, n, ind],
            [False, True, False, False, True, False],
            [CLR["txt"], dk, CLR["url"], CLR["url"], CLR["txt"], CLR["txt"]],
            [True, False, False, False, True, False],
        ), 1):
            ul = "single" if c in (3, 4) else None
            _w(ws, row, c, v, bold=b, clr=fc, bg=bg, center=ctr, wrap=True,
               sz=9, underline=ul)
        ws.row_dimensions[row].height = 28

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(6)}{row}"


# ── Stocks sheet per category ─────────────────────────────────────────────────
def _build_cat_stocks(wb, cat, cat_entries, data) -> int:
    dk, lt = CAT_CLR[cat]
    ws = wb.create_sheet(f"{cat.replace(' ','_')}_Stocks")
    ws.sheet_properties.tabColor = dk

    _title(ws, f"NSE INDIA — {cat.upper()} | ALL CONSTITUENT STOCKS", 7, dk)
    _sub(ws, 7)

    # discover extra columns beyond the standard five
    std = {"Symbol", "Company Name", "Industry", "Series", "ISIN Code"}
    extra_cols = []
    for idx_name, *_ in cat_entries:
        df = data.get(idx_name)
        if df is not None:
            for col in df.columns:
                if col not in std and col not in extra_cols:
                    extra_cols.append(col)

    hdrs   = ["#", "Index Name", "Symbol", "Company Name",
              "Industry", "Series", "ISIN Code"] + extra_cols
    widths = [4, 34, 16, 42, 28, 8, 18] + [14] * len(extra_cols)
    _widths(ws, widths)
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, 4, c, h, dk)

    row = serial = 0
    toggle = 0
    last_idx = None
    row = 4
    for idx_name, *_ in cat_entries:
        df = data.get(idx_name)
        if df is None:
            continue
        if idx_name != last_idx:
            toggle ^= 1
            last_idx = idx_name
        bg = lt if toggle else CLR["white"]

        for _, sr in df.iterrows():
            row    += 1
            serial += 1
            sym  = str(sr.get("Symbol",       "")).strip()
            name = str(sr.get("Company Name", "")).strip()
            ind  = str(sr.get("Industry",     "")).strip()
            ser  = str(sr.get("Series",       "")).strip()
            isin = str(sr.get("ISIN Code",    "")).strip()
            vals = [serial, idx_name, sym, name, ind, ser, isin] + \
                   [str(sr.get(ec, "")).strip() for ec in extra_cols]

            for c, v in enumerate(vals, 1):
                _w(ws, row, c, v,
                   bold=(c == 3), clr=dk if c == 3 else CLR["txt"],
                   bg=bg, center=(c in (1, 3, 6, 7)), sz=9)
            ws.row_dimensions[row].height = 15

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(hdrs))}{row}"
    return serial


# ── Combined all-stocks sheet ─────────────────────────────────────────────────
def _build_all_stocks(wb, data):
    ws = wb.create_sheet("All_Stocks")
    ws.sheet_properties.tabColor = "404040"

    NC = 8
    _title(ws, "ALL NIFTY INDEX CONSTITUENTS — COMBINED (ALL CATEGORIES)", NC, CLR["all_dk"])
    _sub(ws, NC)
    hdrs = ["#", "Symbol", "Company Name", "Industry",
            "Series", "ISIN Code", "Index Name", "Category"]
    _widths(ws, [4, 16, 42, 28, 8, 18, 36, 15])
    for c, h in enumerate(hdrs, 1):
        _hdr(ws, 4, c, h, CLR["all_dk"])

    row = 4
    serial = 0
    toggle = 0
    last_idx = None

    for idx_name, cat, *_ in INDICES:
        df = data.get(idx_name)
        if df is None:
            continue
        dk, lt = CAT_CLR[cat]
        if idx_name != last_idx:
            toggle ^= 1
            last_idx = idx_name
        bg = lt if toggle else CLR["white"]

        for _, sr in df.iterrows():
            row    += 1
            serial += 1
            vals = [
                serial,
                str(sr.get("Symbol",       "")).strip(),
                str(sr.get("Company Name", "")).strip(),
                str(sr.get("Industry",     "")).strip(),
                str(sr.get("Series",       "")).strip(),
                str(sr.get("ISIN Code",    "")).strip(),
                idx_name, cat,
            ]
            for c, v in enumerate(vals, 1):
                _w(ws, row, c, v,
                   bold=(c == 2), clr=dk if c == 2 else CLR["txt"],
                   bg=bg, center=(c in (1, 2, 5, 6)), sz=9)
            ws.row_dimensions[row].height = 15

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(NC)}{row}"
    print(f"  All_Stocks: {serial} total rows")


# ── Orchestrate ───────────────────────────────────────────────────────────────
def build_excel(data: dict, out_path: str):
    print("\n[2/2]  Building Excel …")
    wb = openpyxl.Workbook()

    _build_summary(wb, data)
    print("  ✓ Summary sheet")

    for cat in ["Broad Based", "Sectoral", "Thematic", "Strategy"]:
        entries = [(n, c, s, f) for n, c, s, f in INDICES if c == cat]
        _build_cat_list(wb, cat, entries, data)
        n = _build_cat_stocks(wb, cat, entries, data)
        print(f"  ✓ {cat}: {len(entries)} indices, {n} stock rows")

    _build_all_stocks(wb, data)

    wb.save(out_path)
    print(f"\n✅  Saved → {Path(out_path).resolve()}")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    out = "NIFTY_Indices_Master.xlsx"

    print("=" * 62)
    print("  NIFTY Indices Scraper + Excel Builder")
    print(f"  {len(INDICES)} indices  |  4 categories")
    print("=" * 62)

    print("\n[1/2]  Downloading constituent CSVs …\n")
    data = fetch_all_csvs()

    fetched = sum(1 for v in data.values() if v is not None)
    print(f"\n  Downloaded: {fetched}/{len(INDICES)}")
    if fetched < len(INDICES):
        missing = [n for n, *_ in INDICES if n not in data]
        print("  Missing   :", missing)

    build_excel(data, out)
    print("\nDone! Open", out)
    print("=" * 62)
