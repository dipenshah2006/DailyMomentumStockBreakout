"""
Refresh india/NSE/NIFTY_Indices_Master.xlsx with today's constituent data.

Downloads all NIFTY index CSV files from niftyindices.com in parallel,
then rewrites the All_Stocks sheet (and per-category sheets) so the
sector/industry mapping used by generate_summary.py is always current.

Exit codes:
  0 — success (or graceful fallback: existing XLSX kept unchanged)
  1 — fatal error (openpyxl import failed, etc.)
"""

import csv
import io
import os
import sys
import time
import urllib.request
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

XLSX_PATH   = "india/NSE/NIFTY_Indices_Master.xlsx"
TIMEOUT_SEC = 20
MAX_WORKERS = 8
RETRY_COUNT = 2

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer":         "https://www.niftyindices.com/",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

# ── All indices to fetch, with category labels ───────────────────────────────
# Priority order for dedup: Sectoral (industry) > Broad Based > Thematic > Strategy
INDEX_CATALOG = [
    # category, index name, csv url
    # ── Broad Based ────────────────────────────────────────────────────────
    ("Broad Based", "NIFTY 50",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty50list.csv"),
    ("Broad Based", "NIFTY NEXT 50",
     "https://www.niftyindices.com/IndexConstituent/ind_niftynext50list.csv"),
    ("Broad Based", "NIFTY 100",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty100list.csv"),
    ("Broad Based", "NIFTY 200",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty200list.csv"),
    ("Broad Based", "NIFTY 500",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty500list.csv"),
    ("Broad Based", "NIFTY TOTAL MARKET",
     "https://www.niftyindices.com/IndexConstituent/ind_niftytotalmarket_list.csv"),
    ("Broad Based", "NIFTY MIDCAP 50",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymidcap50list.csv"),
    ("Broad Based", "NIFTY MIDCAP 100",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymidcap100list.csv"),
    ("Broad Based", "NIFTY MIDCAP 150",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymidcap150list.csv"),
    ("Broad Based", "NIFTY SMALLCAP 50",
     "https://www.niftyindices.com/IndexConstituent/ind_niftysmallcap50list.csv"),
    ("Broad Based", "NIFTY SMALLCAP 100",
     "https://www.niftyindices.com/IndexConstituent/ind_niftysmallcap100list.csv"),
    ("Broad Based", "NIFTY SMALLCAP 250",
     "https://www.niftyindices.com/IndexConstituent/ind_niftysmallcap250list.csv"),
    ("Broad Based", "NIFTY LARGEMIDCAP 250",
     "https://www.niftyindices.com/IndexConstituent/ind_niftylargemidcap250list.csv"),
    ("Broad Based", "NIFTY MIDSMALLCAP 400",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymidsmallcap400list.csv"),
    ("Broad Based", "NIFTY MICROCAP 250",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymicrocap250_list.csv"),
    # ── Sectoral ───────────────────────────────────────────────────────────
    ("Sectoral", "NIFTY AUTO",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyautolist.csv"),
    ("Sectoral", "NIFTY BANK",
     "https://www.niftyindices.com/IndexConstituent/ind_niftybanklist.csv"),
    ("Sectoral", "NIFTY FINANCIAL SERVICES",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyfinancelist.csv"),
    ("Sectoral", "NIFTY FINANCIAL SERVICES EX-BANK",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyfinancialservicesexbank_list.csv"),
    ("Sectoral", "NIFTY FMCG",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyfmcglist.csv"),
    ("Sectoral", "NIFTY IT",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyitlist.csv"),
    ("Sectoral", "NIFTY MEDIA",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymedialist.csv"),
    ("Sectoral", "NIFTY METAL",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymetallist.csv"),
    ("Sectoral", "NIFTY PHARMA",
     "https://www.niftyindices.com/IndexConstituent/ind_niftypharmalist.csv"),
    ("Sectoral", "NIFTY PSU BANK",
     "https://www.niftyindices.com/IndexConstituent/ind_niftypsubanklist.csv"),
    ("Sectoral", "NIFTY PRIVATE BANK",
     "https://www.niftyindices.com/IndexConstituent/ind_niftypvtbanklist.csv"),
    ("Sectoral", "NIFTY REALTY",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyrealtylist.csv"),
    ("Sectoral", "NIFTY HEALTHCARE",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyhealthcarelist.csv"),
    ("Sectoral", "NIFTY OIL & GAS",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyoilgaslist.csv"),
    ("Sectoral", "NIFTY CONSUMER DURABLES",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyconsumerdurableslist.csv"),
    ("Sectoral", "NIFTY CAPITAL MARKETS",
     "https://www.niftyindices.com/IndexConstituent/ind_niftycapitalmarketslist.csv"),
    # ── Thematic ───────────────────────────────────────────────────────────
    ("Thematic", "NIFTY COMMODITIES",
     "https://www.niftyindices.com/IndexConstituent/ind_niftycommoditieslist.csv"),
    ("Thematic", "NIFTY INDIA CONSUMPTION",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyconsumptionlist.csv"),
    ("Thematic", "NIFTY CPSE",
     "https://www.niftyindices.com/IndexConstituent/ind_niftycpselist.csv"),
    ("Thematic", "NIFTY ENERGY",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyenergylist.csv"),
    ("Thematic", "NIFTY INFRASTRUCTURE",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyinfralist.csv"),
    ("Thematic", "NIFTY MNC",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymnclist.csv"),
    ("Thematic", "NIFTY PSE",
     "https://www.niftyindices.com/IndexConstituent/ind_niftypselist.csv"),
    ("Thematic", "NIFTY SERVICES SECTOR",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyservicesectorlist.csv"),
    ("Thematic", "NIFTY INDIA DIGITAL",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyindiadigital_list.csv"),
    ("Thematic", "NIFTY INDIA MANUFACTURING",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyindiamanufacturing_list.csv"),
    ("Thematic", "NIFTY INDIA DEFENCE",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyindiadefence_list.csv"),
    ("Thematic", "NIFTY MOBILITY",
     "https://www.niftyindices.com/IndexConstituent/ind_niftymobility_list.csv"),
    ("Thematic", "NIFTY INDIA TOURISM",
     "https://www.niftyindices.com/IndexConstituent/ind_niftytourism_list.csv"),
    # ── Strategy ───────────────────────────────────────────────────────────
    ("Strategy", "NIFTY DIVIDEND OPPORTUNITIES 50",
     "https://www.niftyindices.com/IndexConstituent/ind_niftydividendopportunities50list.csv"),
    ("Strategy", "NIFTY50 EQUAL WEIGHT",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty50EWlist.csv"),
    ("Strategy", "NIFTY100 EQUAL WEIGHT",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty100EWlist.csv"),
    ("Strategy", "NIFTY ALPHA 50",
     "https://www.niftyindices.com/IndexConstituent/ind_niftyalpha50list.csv"),
    ("Strategy", "NIFTY200 MOMENTUM 30",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty200momentum30list.csv"),
    ("Strategy", "NIFTY500 MOMENTUM 50",
     "https://www.niftyindices.com/IndexConstituent/ind_nifty500momentum50list.csv"),
]

# Category priority for industry assignment (Sectoral data is most specific)
CATEGORY_PRIORITY = {"Sectoral": 0, "Broad Based": 1, "Thematic": 2, "Strategy": 3}


# ── Download ──────────────────────────────────────────────────────────────────

def _fetch_csv(category: str, index_name: str, url: str) -> tuple[str, str, str, list[dict]]:
    """Download one index CSV. Returns (category, index_name, url, rows)."""
    last_err = None
    for attempt in range(RETRY_COUNT + 1):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=TIMEOUT_SEC) as resp:
                raw = resp.read().decode("utf-8-sig", errors="replace")
            rows = list(csv.DictReader(io.StringIO(raw)))
            # Normalise column names and values (strip whitespace, guard None)
            rows = [
                {(k.strip() if k else ""): (v.strip() if v else "")
                 for k, v in r.items() if k}
                for r in rows
                if any(v for v in r.values())
            ]
            return category, index_name, url, rows
        except Exception as e:
            last_err = e
            if attempt < RETRY_COUNT:
                time.sleep(2 ** attempt)
    print(f"  ⚠️  {index_name}: {last_err}")
    return category, index_name, url, []


def fetch_all() -> dict[tuple[str, str], list[dict]]:
    """Download all CSVs in parallel. Returns {(category, index_name): rows}."""
    results = {}
    ok = failed = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_fetch_csv, cat, name, url): name
                   for cat, name, url in INDEX_CATALOG}
        for fut in as_completed(futures):
            cat, name, url, rows = fut.result()
            results[(cat, name)] = rows
            if rows:
                ok += 1
            else:
                failed += 1
    print(f"  Downloaded: {ok} OK, {failed} failed")
    return results


# ── Build in-memory data structures ──────────────────────────────────────────

def build_all_stocks(results: dict) -> list[dict]:
    """
    Deduplicate across all indices.
    Each symbol appears once, assigned to the index/industry of the highest-priority category
    (Sectoral first, then Broad Based, Thematic, Strategy).
    Returns list of dicts ready to write to All_Stocks sheet.
    """
    # sym → (priority, category, index_name, company, industry, series, isin)
    best: dict[str, tuple] = {}
    for (cat, idx_name), rows in results.items():
        prio = CATEGORY_PRIORITY.get(cat, 99)
        for row in rows:
            sym     = row.get("Symbol", "").strip()
            company = row.get("Company Name", "").strip()
            industry= row.get("Industry", "").strip()
            series  = row.get("Series", "EQ").strip()
            isin    = row.get("ISIN Code", "").strip()
            if not sym:
                continue
            if sym not in best or prio < best[sym][0]:
                best[sym] = (prio, cat, idx_name, company, industry, series, isin)

    all_stocks = []
    for rank, (sym, (prio, cat, idx_name, company, industry, series, isin)) in \
            enumerate(sorted(best.items(), key=lambda x: (CATEGORY_PRIORITY.get(x[1][1], 99), x[0])), 1):
        all_stocks.append({
            "#":           rank,
            "Symbol":      sym,
            "Company Name":company,
            "Industry":    industry,
            "Series":      series,
            "ISIN Code":   isin,
            "Index Name":  idx_name,
            "Category":    cat,
        })
    return all_stocks


def build_category_stocks(results: dict, target_cat: str) -> list[dict]:
    """All stocks for a given category (one row per index membership)."""
    rows_out = []
    rank = 1
    for (cat, idx_name), rows in results.items():
        if cat != target_cat:
            continue
        for row in rows:
            sym = row.get("Symbol", "").strip()
            if not sym:
                continue
            rows_out.append({
                "#":           rank,
                "Index Name":  idx_name,
                "Symbol":      sym,
                "Company Name":row.get("Company Name", "").strip(),
                "Industry":    row.get("Industry", "").strip(),
                "Series":      row.get("Series", "EQ").strip(),
                "ISIN Code":   row.get("ISIN Code", "").strip(),
            })
            rank += 1
    return rows_out


# ── XLSX writer ───────────────────────────────────────────────────────────────

def _header_style(ws, row_idx: int, ncols: int):
    """Bold the header row."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1E3A5F")
        font = Font(bold=True, color="FFFFFF")
        align = Alignment(horizontal="left")
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font  = font
            cell.fill  = fill
            cell.alignment = align
    except Exception:
        pass


def write_sheet(ws, title_line: str, headers: list[str], rows: list[dict]):
    today_str = datetime.now().strftime("%d %b %Y")
    ws.append([title_line])
    ws.append([f"Source: www.niftyindices.com  |  Generated: {today_str}"])
    ws.append([])
    ws.append(headers)
    _header_style(ws, 4, len(headers))
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def update_xlsx(results: dict):
    """Rebuild all sheets in the XLSX with fresh data."""
    import openpyxl

    all_stocks      = build_all_stocks(results)
    broad_stocks    = build_category_stocks(results, "Broad Based")
    sectoral_stocks = build_category_stocks(results, "Sectoral")
    thematic_stocks = build_category_stocks(results, "Thematic")
    strategy_stocks = build_category_stocks(results, "Strategy")

    today_str = datetime.now().strftime("%d %b %Y")

    # ── Load or create workbook ──────────────────────────────────────────
    if os.path.exists(XLSX_PATH):
        wb = openpyxl.load_workbook(XLSX_PATH)
        # Remove sheets we'll recreate
        for sheet_name in ["All_Stocks", "Broad_Based_Stocks", "Sectoral_Stocks",
                           "Thematic_Stocks", "Strategy_Stocks"]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    stock_headers = ["#", "Symbol", "Company Name", "Industry", "Series", "ISIN Code",
                     "Index Name", "Category"]
    member_headers = ["#", "Index Name", "Symbol", "Company Name", "Industry", "Series",
                      "ISIN Code"]

    # ── Write per-category sheets ────────────────────────────────────────
    ws_bb = wb.create_sheet("Broad_Based_Stocks")
    write_sheet(ws_bb, "NSE INDIA — BROAD BASED | ALL CONSTITUENT STOCKS",
                member_headers, broad_stocks)

    ws_sec = wb.create_sheet("Sectoral_Stocks")
    write_sheet(ws_sec, "NSE INDIA — SECTORAL | ALL CONSTITUENT STOCKS",
                member_headers, sectoral_stocks)

    ws_them = wb.create_sheet("Thematic_Stocks")
    write_sheet(ws_them, "NSE INDIA — THEMATIC | ALL CONSTITUENT STOCKS",
                member_headers, thematic_stocks)

    ws_strat = wb.create_sheet("Strategy_Stocks")
    write_sheet(ws_strat, "NSE INDIA — STRATEGY | ALL CONSTITUENT STOCKS",
                member_headers, strategy_stocks)

    # ── Write All_Stocks (deduplicated) ──────────────────────────────────
    ws_all = wb.create_sheet("All_Stocks")
    write_sheet(ws_all,
                "ALL NIFTY INDEX CONSTITUENTS — COMBINED (ALL CATEGORIES)",
                stock_headers, all_stocks)

    os.makedirs(os.path.dirname(XLSX_PATH), exist_ok=True)
    wb.save(XLSX_PATH)
    return len(all_stocks)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("❌ openpyxl not installed — cannot refresh NIFTY master.")
        sys.exit(1)

    print(f"🔄 Refreshing {XLSX_PATH} from niftyindices.com ...")
    print(f"   Fetching {len(INDEX_CATALOG)} index CSVs ({MAX_WORKERS} parallel workers) ...")
    t0 = time.time()

    results = fetch_all()

    total_downloaded = sum(len(v) for v in results.values())
    if total_downloaded == 0:
        print("⚠️  All downloads failed — keeping existing XLSX unchanged.")
        sys.exit(0)

    print(f"  Total constituent rows fetched: {total_downloaded}")
    print("📝 Writing updated XLSX ...")

    try:
        n = update_xlsx(results)
        elapsed = time.time() - t0
        print(f"✅ {XLSX_PATH} updated: {n} unique symbols in {elapsed:.1f}s")
    except Exception as e:
        print(f"⚠️  Failed to write XLSX ({e}) — keeping existing file unchanged.")
        sys.exit(0)


if __name__ == "__main__":
    main()
